from __future__ import annotations

import base64
import io
import json
import os
import time
import uuid
import zipfile
from types import SimpleNamespace
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timezone

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from pypdf import PdfReader
from openpyxl import Workbook
from sqlalchemy import create_engine, event, func, select, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import app.models  # noqa: F401
from app.core.db import Base, get_db
from app.core.security import create_access_token
from app.main import app
from app.models.folio_sequence import InstitutionalFolioSequence
from app.models.client import Client
from app.models.audit_log import AuditLog
from app.models.lab_work_order import (
    LabWorkOrder,
    LabWorkOrderEquipment,
    LabWorkOrderGroupRequest,
    LabWorkOrderSignatureSession,
)
from app.models.lab_work_order_revision import LabWorkOrderRevision
from app.models.linked_company import LinkedCompany
from app.models.notification import Notification
from app.models.communication import CommunicationConversation, CommunicationMessage
from app.models.operational_ticket import OperationalTicket
from app.models.user import Role, User
from app.schemas.operational_ticket import TicketReject, TicketReview
from app.schemas.lab_work_order import LabSignatureGroupWrite, LabWorkOrderGroupCreate
from app.services.lab_work_order_pdfs import generate_lab_work_order_pdf
from app.services.work_order_pdfs import _build_equipment_lines
from app.services.lab_work_orders import (
    _allocate_folio,
    create_additional_work_order,
    sign_individual,
)
from app.services.lab_work_orders import delete_work_order
from app.services.lab_work_orders import create_group_request, claim_group_request, create_work_order_group, reject_group_request, approve_group_request
from app.services.operational_tickets import approve_reopen_ticket, reject_ticket


PNG_DATA_URL = "data:image/png;base64," + base64.b64encode(
    base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
    )
).decode()


@pytest.fixture()
def lab_context():
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine, "connect")
    def enable_sqlite_foreign_keys(dbapi_connection, _connection_record):
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, expire_on_commit=False)
    with factory() as db:
        technician_role = Role(name="Tecnico", description="Técnico")
        capture_role = Role(name="Captura", description="Captura")
        admin_role = Role(name="Administrador", description="Administrador")
        db.add_all([technician_role, capture_role, admin_role])
        db.flush()
        users = []
        for key, role in (
            ("tech", technician_role),
            ("capture", capture_role),
            ("admin", admin_role),
        ):
            user = User(
                username=f"lab-{key}",
                email=f"lab-{key}@example.test",
                full_name=f"LAB {key}",
                hashed_password="unused",
                account_type="internal",
                status="active",
                is_active=True,
                role_id=role.id,
                roles=[role],
            )
            users.append(user)
        db.add_all(users)
        db.commit()

    def override_db():
        with factory() as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    client = TestClient(app)
    tokens = {
        key: create_access_token(
            str(user.id),
            extra_claims={"roles": [user.roles[0].name], "auth_context": "internal"},
        )
        for key, user in zip(("tech", "capture", "admin"), users, strict=True)
    }
    try:
        yield client, factory, tokens
    finally:
        app.dependency_overrides.clear()


def test_staff_direct_group_materializes_consecutive_real_orders(lab_context):
    client, factory, tokens = lab_context
    payload = {**create_payload("Cliente final documental"), "quantity": 3}
    response = client.post(
        "/api/lab-work-order-groups", json=payload, headers=auth(tokens["admin"])
    )
    assert response.status_code == 201, response.text
    assert [item["folio"] for item in response.json()["related_work_orders"]] == [6400, 6401, 6402]
    with factory() as db:
        rows = list(db.scalars(select(LabWorkOrder).order_by(LabWorkOrder.sequence_number)).all())
        assert len(rows) == 3
        assert {item.root_work_order_id for item in rows} == {rows[0].id}
        assert all(item.client_name == "Cliente final documental" for item in rows)


def test_group_request_approval_is_idempotent_and_pending_consumes_no_folios(lab_context):
    client, factory, tokens = lab_context
    with factory() as db:
        operator = Client(legal_name="Operador", commercial_name="Operador")
        db.add(operator)
        db.flush()
        requester = db.scalar(select(User).where(User.username == "lab-tech"))
        request = create_group_request(
            db,
            LabWorkOrderGroupCreate(**create_payload("Cliente final"), quantity=2),
            requester,
            operator_client_id=operator.id,
        )
        request_id = request.id
        assert request.conversation_id is None
        assert db.scalar(select(func.count(LabWorkOrder.id))) == 0
        assert db.scalar(select(func.count(CommunicationConversation.id))) == 0
        admin_notification = db.scalar(select(Notification).where(Notification.entity_type == "work_order_group_request"))
        assert admin_notification.entity_id == request_id
    headers = auth(tokens["admin"])
    inbox = client.get("/api/lab-work-order-groups/requests", headers=headers)
    assert inbox.status_code == 200
    assert inbox.json()[0]["id"] == request_id
    assert inbox.json()[0]["operator_client_name"] == "Operador"
    assert inbox.json()[0]["requested_by_name"] == "LAB tech"
    claimed = client.post(f"/api/lab-work-order-groups/requests/{request_id}/claim", headers=headers)
    assert claimed.status_code == 200
    conversation_id = claimed.json()["conversation_id"]
    assert conversation_id is not None
    first = client.post(f"/api/lab-work-order-groups/requests/{request_id}/approve", headers=headers)
    second = client.post(f"/api/lab-work-order-groups/requests/{request_id}/approve", headers=headers)
    assert first.status_code == second.status_code == 200
    assert first.json()["root_work_order_id"] == second.json()["root_work_order_id"]
    with factory() as db:
        assert db.scalar(select(func.count(LabWorkOrder.id))) == 2
        assert list(db.scalars(select(LabWorkOrder.folio).order_by(LabWorkOrder.folio))) == [6400, 6401]
        conversation = db.get(CommunicationConversation, conversation_id)
        assert {participant.username for participant in conversation.participants} == {"lab-tech", "lab-admin"}
        messages = list(db.scalars(select(CommunicationMessage).where(CommunicationMessage.conversation_id == conversation_id).order_by(CommunicationMessage.sequence)))
        assert any("solicitó un grupo de 2 órdenes" in message.body for message in messages)
        assert any("está atendiendo" in message.body for message in messages)
        assert any("Folios asignados: 6400, 6401" in message.body for message in messages)


def test_group_request_has_one_handler_and_rejection_message_without_folios(lab_context):
    _client, factory, _tokens = lab_context
    with factory() as db:
        operator = Client(legal_name="Operador B", commercial_name="Operador B")
        db.add(operator); db.flush()
        requester = db.scalar(select(User).where(User.username == "lab-tech"))
        admin = db.scalar(select(User).where(User.username == "lab-admin"))
        competitor = db.scalar(select(User).where(User.username == "lab-capture"))
        request = create_group_request(db, LabWorkOrderGroupCreate(**create_payload("Cliente rechazado"), quantity=3), requester, operator_client_id=operator.id)
        claimed = claim_group_request(db, request.id, admin)
        with pytest.raises(HTTPException) as missing_reason:
            reject_group_request(db, request.id, admin, "")
        assert missing_reason.value.status_code == 422
        with pytest.raises(HTTPException) as conflict:
            claim_group_request(db, request.id, competitor)
        assert conflict.value.status_code == 409
        rejected = reject_group_request(db, request.id, admin, "Capacidad no disponible")
        assert rejected.status == "rejected"
        assert rejected.handled_by_user_id == claimed.handled_by_user_id == admin.id
        assert rejected.folios == []
        assert db.scalar(select(func.count(LabWorkOrder.id))) == 0
        messages = list(db.scalars(select(CommunicationMessage).where(CommunicationMessage.conversation_id == rejected.conversation_id)))
        assert any("rechazó la solicitud. Motivo: Capacidad no disponible" in item.body for item in messages)


def test_requester_cannot_approve_or_reject_their_own_group_request(lab_context):
    _client, factory, _tokens = lab_context
    with factory() as db:
        operator = Client(legal_name="Operador C", commercial_name="Operador C")
        db.add(operator)
        db.flush()
        requester = db.scalar(select(User).where(User.username == "lab-tech"))
        request = create_group_request(
            db,
            LabWorkOrderGroupCreate(**create_payload("Cliente autoaprobado"), quantity=1),
            requester,
            operator_client_id=operator.id,
        )
        with pytest.raises(HTTPException) as approve_exc:
            approve_group_request(db, request.id, requester)
        assert approve_exc.value.status_code == 403
        assert approve_exc.value.detail == "TICKET_SELF_APPROVAL_FORBIDDEN"
        with pytest.raises(HTTPException) as reject_exc:
            reject_group_request(db, request.id, requester, "No aplica")
        assert reject_exc.value.status_code == 403
        assert reject_exc.value.detail == "TICKET_SELF_APPROVAL_FORBIDDEN"


def test_postgresql_concurrent_direct_groups_do_not_duplicate_folios(postgres_lab_context):
    _client, factory, _tokens = postgres_lab_context
    payload = LabWorkOrderGroupCreate(**create_payload("Cliente concurrente"), quantity=2)

    def create_group(index: int) -> list[int]:
        with factory() as db:
            user = db.scalar(select(User).where(User.username == "postgres-admin"))
            created = create_work_order_group(db, payload, user, operator_client_id=None)
            return [item.folio for item in created.related_work_orders]

    with ThreadPoolExecutor(max_workers=2) as pool:
        groups = list(pool.map(create_group, range(2)))

    assert all(group[1] == group[0] + 1 for group in groups)
    assert sorted(folio for group in groups for folio in group) == [6400, 6401, 6402, 6403]


@pytest.fixture()
def postgres_lab_context():
    database_url = os.getenv("LAB_POSTGRES_TEST_URL")
    if not database_url:
        pytest.skip("requiere LAB_POSTGRES_TEST_URL para probar locks PostgreSQL reales")

    schema = f"ticket_lock_{uuid.uuid4().hex}"
    admin_engine = create_engine(database_url)
    with admin_engine.begin() as connection:
        connection.execute(text(f'CREATE SCHEMA "{schema}"'))

    engine = create_engine(
        database_url,
        connect_args={"options": f"-csearch_path={schema}"},
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, expire_on_commit=False)
    with factory() as db:
        roles = {
            name: Role(name=name, description=name)
            for name in ("Tecnico", "Captura", "Administrador")
        }
        db.add_all(roles.values())
        db.flush()
        users = {}
        for key, role_name in (
            ("tech", "Tecnico"),
            ("capture", "Captura"),
            ("admin", "Administrador"),
        ):
            role = roles[role_name]
            user = User(
                username=f"postgres-{key}",
                email=f"postgres-{key}@example.test",
                full_name=f"PostgreSQL {key}",
                hashed_password="unused",
                account_type="internal",
                status="active",
                is_active=True,
                role_id=role.id,
                roles=[role],
            )
            users[key] = user
            db.add(user)
        db.commit()

    def override_db():
        with factory() as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    client = TestClient(app)
    tokens = {
        key: create_access_token(
            str(user.id),
            extra_claims={"roles": [user.roles[0].name], "auth_context": "internal"},
        )
        for key, user in users.items()
    }
    try:
        yield client, factory, tokens
    finally:
        app.dependency_overrides.clear()
        engine.dispose()
        with admin_engine.begin() as connection:
            connection.execute(text(f'DROP SCHEMA "{schema}" CASCADE'))
        admin_engine.dispose()


def auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def create_payload(client_name: str = "Cliente LAB") -> dict:
    return {
        "reception_date": "2026-08-13",
        "client_name": client_name,
        "address": "Av. Prueba 123",
        "contact_name": "Persona Cliente",
        "contact_phone": "3312345678",
        "contact_email": "cliente@example.com",
        "postal_code": "45601",
        "city": "Tlaquepaque",
        "state_name": "Jalisco",
        "purchase_order": "OC-123",
        "notes": "Recepción LAB",
    }


def equipment_payload(index: int, **extra) -> dict:
    return {
        "instrument": f"Instrumento {index}",
        "brand": "MYC Test",
        "identification": f"ID-{index}",
        "serial_number": f"SER-{index}",
        "report_number": None,
        "is_good_condition": index % 2 == 0,
        **extra,
    }


def configure_default_services(
    client: TestClient, headers: dict[str, str], work_order_id: int, *, service_type: str = "traceable"
) -> None:
    """Fase 3: la recepción sólo puede firmarse cuando cada equipo trae una
    configuración operacional coherente (servicio elegido y, si aplica,
    folio ya reservado). 'traceable' reserva un folio MYCT real vía el
    allocator existente sin requerir LinkedCompany, así que sirve como
    configuración por defecto para pruebas que no ejercitan el servicio en
    sí, sólo el ciclo de firma/cierre."""
    detail = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{work_order_id}", headers=headers
    ).json()
    for item in detail["equipment"]:
        if item["service_type"] is not None:
            continue
        response = client.put(
            f"/api/mobile/v1/technician/lab-work-orders/{work_order_id}/equipment/{item['id']}/service",
            json={"service_type": service_type, "linked_company_id": None},
            headers=headers,
        )
        assert response.status_code == 200, response.text


def signatures_payload() -> dict:
    signed_at = datetime.now(timezone.utc).isoformat()
    return {
        "technician": {
            "signer_name": "Técnico LAB",
            "signed_at": signed_at,
            "version": 1,
            "signature_data_url": PNG_DATA_URL,
        },
        "client": {
            "signer_name": "Cliente LAB",
            "signed_at": signed_at,
            "version": 1,
            "signature_data_url": PNG_DATA_URL,
        },
    }


def test_lab_security_and_initial_folio(lab_context):
    client, _factory, tokens = lab_context
    url = "/api/mobile/v1/technician/lab-work-orders"
    assert client.get(url).status_code == 401
    capture_response = client.get(url, headers=auth(tokens["capture"]))
    assert capture_response.status_code == 200
    assert capture_response.json() == []

    first = client.post(url, json=create_payload(), headers=auth(tokens["tech"]))
    second = client.post(url, json=create_payload("Otro"), headers=auth(tokens["tech"]))
    assert first.status_code == 201, first.text
    assert second.status_code == 201, second.text
    assert [first.json()["folio"], second.json()["folio"]] == [6400, 6401]
    assert first.json()["root_work_order_id"] == first.json()["id"]


def test_lab_client_xlsx_import_normalizes_exact_identity_without_collapsing_company(lab_context):
    client, _factory, tokens = lab_context
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["CLIENTE", "CONTACTO", "DIRECCIÓN"])
    sheet.append(["Honda de México", "Ing. Juan Pérez", "Av. Industria 123"])
    sheet.append(["  HÓNDA de mexico ", "ING JUAN PEREZ", "Av Industria 123"])
    sheet.append(["Honda de México", "Lic. Ana Pérez", "Av. Industria 123"])
    sheet.append(["Honda de México", "Ing. Juan Pérez", "Av. Industria 456"])
    sheet.append(["", "Contacto inválido", "Sin cliente"])
    content = io.BytesIO()
    workbook.save(content)

    response = client.post(
        "/api/mobile/v1/technician/lab-clients/import",
        files={"upload": ("clientes.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth(tokens["admin"]),
    )
    assert response.status_code == 200, response.text
    assert response.json() == {
        "new": 3,
        "skipped": 1,
        "invalid": 1,
        "errors": [{"row": 6, "reason": "Falta Empresa"}],
    }
    listed = client.get(
        "/api/mobile/v1/technician/lab-clients?search=Honda",
        headers=auth(tokens["tech"]),
    )
    assert listed.status_code == 200
    assert len(listed.json()) == 3
    assert {item["attention"] for item in listed.json()} == {"Ing. Juan Pérez", "Lic. Ana Pérez"}


def test_lab_client_xlsx_import_keeps_rows_with_blank_address_or_attention(lab_context):
    client, _factory, tokens = lab_context
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["CLIENTE", "CONTACTO", "DIRECCIÓN"])
    sheet.append(["Solo Empresa SA", "", ""])
    sheet.append(["Sin Dirección SA", "Ing. Contacto", ""])
    sheet.append(["Sin Contacto SA", "", "Av. Siempre 1"])
    sheet.append(["", "Contacto huérfano", "Dirección huérfana"])
    content = io.BytesIO()
    workbook.save(content)

    response = client.post(
        "/api/mobile/v1/technician/lab-clients/import",
        files={"upload": ("clientes.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth(tokens["admin"]),
    )
    assert response.status_code == 200, response.text
    assert response.json() == {
        "new": 3,
        "skipped": 0,
        "invalid": 1,
        "errors": [{"row": 5, "reason": "Falta Empresa"}],
    }
    listed = client.get(
        "/api/mobile/v1/technician/lab-clients?search=Empresa",
        headers=auth(tokens["tech"]),
    )
    assert listed.status_code == 200
    assert len(listed.json()) == 1
    assert listed.json()[0]["address"] == ""
    assert listed.json()[0]["attention"] == ""


@pytest.mark.parametrize(
    ("attention_header", "postal_header"),
    [
        ("ATENCIÓN", "CÓDIGO POSTAL"),
        ("ATENCION", "CODIGO POSTAL"),
        ("CONTACTO", "CP"),
        (" contacto ", "C.P."),
    ],
)
def test_lab_client_xlsx_import_accepts_structured_header_aliases_and_ignores_auxiliary_columns(
    lab_context,
    attention_header,
    postal_header,
):
    client, _factory, tokens = lab_context
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "  cliente  ",
        attention_header,
        " DIRECCION ",
        postal_header,
        " ciudad ",
        "ESTADO",
        "DIRECCIÓN ORIGINAL",
        "REVISAR",
    ])
    sheet.append([
        "  Empresa Estructurada SA  ",
        "  Ing. Contacto  ",
        "  Calle Institucional 123  ",
        "01234",
        "  Guadalajara  ",
        "  Jalisco  ",
        "Dirección histórica que no debe importarse",
        "Sí",
    ])
    content = io.BytesIO()
    workbook.save(content)

    response = client.post(
        "/api/mobile/v1/technician/lab-clients/import",
        files={
            "upload": (
                "clientes-estructurados.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth(tokens["admin"]),
    )
    assert response.status_code == 200, response.text
    assert response.json() == {"new": 1, "skipped": 0, "invalid": 0, "errors": []}

    listed = client.get(
        "/api/mobile/v1/technician/lab-clients?search=Estructurada&limit=5",
        headers=auth(tokens["tech"]),
    )
    assert listed.status_code == 200, listed.text
    assert len(listed.json()) == 1
    imported = listed.json()[0]
    assert {
        field: imported[field]
        for field in ("company", "attention", "address", "postal_code", "city", "state")
    } == {
        "company": "Empresa Estructurada SA",
        "attention": "Ing. Contacto",
        "address": "Calle Institucional 123",
        "postal_code": "01234",
        "city": "Guadalajara",
        "state": "Jalisco",
    }


def test_work_order_snapshot_completes_blank_catalog_fields_without_overwriting_payload(lab_context):
    """A LabClient with blank address/attention (allowed since only company is
    required) must not force an OT snapshot back to blank when the OT payload
    already supplies its own address/contact_name."""
    client, factory, tokens = lab_context
    headers = auth(tokens["tech"])
    created_client = client.post(
        "/api/mobile/v1/technician/lab-clients",
        json={"company": "Cliente sin datos", "address": "", "attention": ""},
        headers=headers,
    )
    assert created_client.status_code == 201, created_client.text

    payload = create_payload("Se reemplaza por snapshot de empresa")
    payload["lab_client_id"] = created_client.json()["id"]
    payload["address"] = "Dirección capturada en la OT"
    payload["contact_name"] = "Atención capturada en la OT"
    order = client.post(
        "/api/mobile/v1/technician/lab-work-orders", json=payload, headers=headers
    )
    assert order.status_code == 201, order.text
    body = order.json()
    assert body["client_name"] == "Cliente sin datos"
    assert body["address"] == "Dirección capturada en la OT"
    assert body["contact_name"] == "Atención capturada en la OT"

    updated = client.patch(
        f"/api/mobile/v1/technician/lab-work-orders/{body['id']}",
        json={"lab_client_id": created_client.json()["id"], "address": "Dirección editada de nuevo"},
        headers=headers,
    )
    assert updated.status_code == 200, updated.text
    assert updated.json()["address"] == "Dirección editada de nuevo"
    assert updated.json()["contact_name"] == "Atención capturada en la OT"


def test_modern_lab_flow_uses_independent_sequences_and_linked_manual_folio(lab_context):
    client, factory, tokens = lab_context
    headers = auth(tokens["tech"])
    created_client = client.post(
        "/api/mobile/v1/technician/lab-clients",
        json={"company": "Cliente moderno", "address": "Calle Uno 10", "attention": "Ing. Responsable"},
        headers=headers,
    )
    assert created_client.status_code == 201, created_client.text
    with factory() as db:
        linked = LinkedCompany(
            name="Laboratorio vinculado",
            abbreviation="LV-TEST",
            default_certificate_prefix="LVT",
            is_enabled=True,
        )
        db.add(linked)
        db.commit()
        linked_id = linked.id

    payload = create_payload("Este valor se reemplaza por snapshot")
    payload["lab_client_id"] = created_client.json()["id"]
    order = client.post(
        "/api/mobile/v1/technician/lab-work-orders", json=payload, headers=headers
    )
    assert order.status_code == 201, order.text
    assert order.json()["client_name"] == "Cliente moderno"
    order_id = order.json()["id"]
    equipment_ids = []
    for index in range(1, 4):
        added = client.post(
            f"/api/mobile/v1/technician/lab-work-orders/{order_id}/equipment",
            json=equipment_payload(index),
            headers=headers,
        )
        assert added.status_code == 201
        equipment_ids.append(added.json()["equipment"][-1]["id"])

    accredited = client.put(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/equipment/{equipment_ids[0]}/service",
        json={"service_type": "accredited", "linked_company_id": None}, headers=headers,
    )
    traceable = client.put(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/equipment/{equipment_ids[1]}/service",
        json={"service_type": "traceable", "linked_company_id": None}, headers=headers,
    )
    linked_response = client.put(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/equipment/{equipment_ids[2]}/service",
        json={"service_type": "linked", "linked_company_id": linked_id}, headers=headers,
    )
    assert accredited.status_code == traceable.status_code == linked_response.status_code == 200
    month_year = date.today().strftime("%m-%y")
    assert accredited.json()["equipment"][0]["certificate_folio"] == f"MYCA-{month_year}-4700"
    assert traceable.json()["equipment"][1]["certificate_folio"] == f"MYCT-{month_year}-1640"
    linked_equipment = linked_response.json()["equipment"][2]
    assert linked_equipment["certificate_folio"] is None
    assert linked_equipment["folio_status"] == "pending"
    assert linked_equipment["linked_company_name_snapshot"] == "Laboratorio vinculado"
    assert linked_equipment["linked_company_prefix_snapshot"] == "LVT"
    # Cierre UX 2026-09 (item D): PUT .../service ya materializa la solicitud
    # linked_folio automatica -- igual que create/update_configured_equipment.
    # El POST manual a /tickets/folio para esta misma solicitud ahora es
    # redundante y responde 409 (ver test_manual_linked_folio_endpoint_does_not_
    # duplicate_automatic_request en test_lab_phase2_integrated_alta.py), asi
    # que esta prueba resuelve directamente el ticket automatico ya creado.
    linked_ticket_id = linked_equipment["folio_ticket_id"]
    assert linked_ticket_id is not None
    with factory() as db:
        auto_ticket = db.get(OperationalTicket, linked_ticket_id)
        assert auto_ticket.type == "linked_folio"
        assert auto_ticket.status == "pending"
        assert auto_ticket.conversation_id is not None
    authorized_literal = "cap-y/26 001-a"
    resolved = client.post(
        f"/api/mobile/v1/technician/tickets/{linked_ticket_id}/resolve",
        json={"authorized_folio": authorized_literal, "comment": "Autorizado por Admin"},
        headers=auth(tokens["admin"]),
    )
    assert resolved.status_code == 200, resolved.text
    assert resolved.json()["authorized_folio"] == authorized_literal
    refreshed = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}", headers=headers
    ).json()
    assert refreshed["equipment"][2]["certificate_folio"] == authorized_literal
    assert refreshed["equipment"][2]["folio_status"] == "authorized"
    with factory() as db:
        sequences = {
            row.prefix: row.next_value
            for row in db.scalars(
                select(InstitutionalFolioSequence).where(
                    InstitutionalFolioSequence.document_type == "lab_certificate"
                )
            )
        }
    assert sequences == {"MYCA": 4701, "MYCT": 1641}

    # Fase 3: los tres equipos ya están coherentes (folio reservado/vinculado
    # congelado, el vinculado con folio autorizado aparte) -- la recepción sí
    # puede firmarse; lo que queda bloqueado es el CIERRE, porque ninguna
    # FieldSheet fue capturada todavía.
    signed = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/signatures",
        json=signatures_payload(), headers=headers,
    )
    assert signed.status_code == 200, signed.text
    assert signed.json()["status"] == "received_signed"
    blocked = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/complete",
        headers=headers,
    )
    assert blocked.status_code == 409
    assert blocked.json()["detail"]["code"] == "LAB_FIELD_SHEETS_INCOMPLETE"
    assert [item["equipment_position"] for item in blocked.json()["detail"]["items"]] == [1, 2, 3]

    created_sheet = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/equipment/{equipment_ids[0]}/field-sheet",
        json={"template_key": "general"}, headers=headers,
    )
    assert created_sheet.status_code == 201, created_sheet.text
    sheet_json = created_sheet.json()
    assert sheet_json["company"] == "Cliente moderno"
    assert sheet_json["attention"] == "Ing. Responsable"
    assert sheet_json["capture_values"]["instrument"] == "Instrumento 1"
    assert sheet_json["template_definition"]["template_key"] == "general"
    original_row_count = len(sheet_json["results_rows"])
    section = sheet_json["template_definition"]["result_sections"][0]
    rows = [
        {
            "id": row["id"],
            "section_key": row["section_key"],
            "row_number": row["row_number"],
            "row_data": {"result": "1.00"} if index == 0 else row["row_data"],
        }
        for index, row in enumerate(sheet_json["results_rows"])
    ]
    if section.get("allow_add_rows"):
        rows.append({
            "section_key": section["key"],
            "row_number": max(row["row_number"] for row in rows if row["section_key"] == section["key"]) + 1,
            "row_data": {"result": "2.00"},
        })
    updated_sheet = client.patch(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/equipment/{equipment_ids[0]}/field-sheet",
        json={"final_condition": "BUENA", "observations": "Sin observaciones", "results_rows": rows},
        headers=headers,
    )
    assert updated_sheet.status_code == 200, updated_sheet.text
    if section.get("allow_add_rows"):
        assert len(updated_sheet.json()["results_rows"]) == original_row_count + 1
    completed_sheet = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/equipment/{equipment_ids[0]}/field-sheet/complete",
        headers=headers,
    )
    assert completed_sheet.status_code == 200, completed_sheet.text
    assert completed_sheet.json()["status"] == "completed"
    with factory() as db:
        assert db.get(LabWorkOrder, order_id).status == "in_progress"
    blocked_after_one_sheet = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/complete",
        headers=headers,
    )
    assert blocked_after_one_sheet.status_code == 409
    assert blocked_after_one_sheet.json()["detail"]["code"] == "LAB_FIELD_SHEETS_INCOMPLETE"
    assert [item["equipment_position"] for item in blocked_after_one_sheet.json()["detail"]["items"]] == [2, 3]

    capture_write = client.put(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/equipment/{equipment_ids[0]}/service",
        json={"service_type": "traceable", "linked_company_id": None},
        headers=auth(tokens["capture"]),
    )
    assert capture_write.status_code == 403
    package = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/package",
        headers=auth(tokens["capture"]),
    )
    assert package.status_code == 200
    assert package.content.startswith(b"%PDF")


def test_equipment_crud_limit_and_model_accepted_but_not_range_or_capacity(lab_context):
    """Cierre UX 2026-09: model sigue siendo identidad propia del equipo LAB
    (mismo criterio que Equipment productivo), pero range_or_capacity
    ("Alcance / capacidad") ya NO es un dato de alta de equipo -- el schema
    (extra='forbid') lo rechaza. Ese dato se captura en la FieldSheet, no en
    la recepción/alta."""
    client, _factory, tokens = lab_context
    headers = auth(tokens["tech"])
    created = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload(),
        headers=headers,
    ).json()
    base = f"/api/mobile/v1/technician/lab-work-orders/{created['id']}"
    rejected = client.post(
        f"{base}/equipment",
        json=equipment_payload(1, model="Modelo X-100", range_or_capacity="0-100 kg"),
        headers=headers,
    )
    assert rejected.status_code == 422, rejected.text
    with_model = client.post(
        f"{base}/equipment",
        json=equipment_payload(1, model="Modelo X-100"),
        headers=headers,
    )
    assert with_model.status_code == 201, with_model.text
    created_equipment = with_model.json()["equipment"][-1]
    assert created_equipment["model"] == "Modelo X-100"
    assert "range_or_capacity" not in created_equipment
    for index in range(2, 11):
        response = client.post(f"{base}/equipment", json=equipment_payload(index), headers=headers)
        assert response.status_code == 201, response.text
    eleventh = client.post(f"{base}/equipment", json=equipment_payload(11), headers=headers)
    assert eleventh.status_code == 409
    detail = client.get(base, headers=headers).json()
    assert len(detail["equipment"]) == 10
    equipment_id = detail["equipment"][0]["id"]
    updated = client.patch(
        f"{base}/equipment/{equipment_id}",
        json=equipment_payload(99),
        headers=headers,
    )
    assert updated.status_code == 200
    deleted = client.delete(f"{base}/equipment/{equipment_id}", headers=headers)
    assert deleted.status_code == 200
    assert [item["position"] for item in deleted.json()["equipment"]] == list(range(1, 10))


def test_additional_work_orders_inherit_and_keep_group_chain(lab_context):
    client, _factory, tokens = lab_context
    headers = auth(tokens["tech"])
    root = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload(),
        headers=headers,
    ).json()
    for index in range(1, 11):
        client.post(
            f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/equipment",
            json=equipment_payload(index),
            headers=headers,
        )
    additional = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/additional",
        headers=headers,
    )
    assert additional.status_code == 201, additional.text
    extra = additional.json()
    assert extra["folio"] == 6401
    assert extra["root_work_order_id"] == root["id"]
    assert extra["previous_work_order_id"] == root["id"]
    assert extra["sequence_number"] == 2
    assert extra["client_name"] == root["client_name"]
    assert extra["equipment"] == []

    for index in range(11, 21):
        client.post(
            f"/api/mobile/v1/technician/lab-work-orders/{extra['id']}/equipment",
            json=equipment_payload(index),
            headers=headers,
        )
    third = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{extra['id']}/additional",
        headers=headers,
    ).json()
    assert (third["folio"], third["root_work_order_id"], third["previous_work_order_id"]) == (
        6402,
        root["id"],
        extra["id"],
    )
    assert [item["folio"] for item in third["related_work_orders"]] == [6400, 6401, 6402]


def test_one_signature_session_completes_and_locks_entire_group(lab_context):
    client, factory, tokens = lab_context
    headers = auth(tokens["tech"])
    mapping_payload = create_payload("CLIENTE PRUEBA")
    mapping_payload.update(
        address="Avenida Ejemplo 123",
        contact_name="Persona Prueba",
        postal_code="45601",
        city="Tlaquepaque",
        state_name="Jalisco",
        purchase_order="OC-TEST-001",
    )
    root = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=mapping_payload,
        headers=headers,
    ).json()
    for index in range(1, 11):
        client.post(
            f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/equipment",
            json=equipment_payload(index),
            headers=headers,
        )
    extra = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/additional",
        headers=headers,
    ).json()
    client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{extra['id']}/equipment",
        json=equipment_payload(11),
        headers=headers,
    )
    configure_default_services(client, headers, root["id"])
    configure_default_services(client, headers, extra["id"])

    signed = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{extra['id']}/signatures",
        json=signatures_payload(),
        headers=headers,
    )
    assert signed.status_code == 200, signed.text
    signature_session_id = signed.json()["signature_session_id"]
    with factory() as db:
        group = list(db.scalars(select(LabWorkOrder).order_by(LabWorkOrder.folio)))
        assert {item.signature_session_id for item in group} == {signature_session_id}
        assert all(item.status == "received_signed" for item in group)

    assert client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{extra['id']}/equipment",
        json=equipment_payload(12),
        headers=headers,
    ).status_code == 409
    completed = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/complete",
        headers=headers,
    )
    assert completed.status_code == 200, completed.text
    assert all(item["status"] == "completed" for item in completed.json()["related_work_orders"])
    for work_order_id, expected_folio, expected_instrument in (
            (root["id"], "6400", "INSTRUMENTO 1"),
            (extra["id"], "6401", "INSTRUMENTO 11"),
    ):
        pdf = client.get(
            f"/api/mobile/v1/technician/lab-work-orders/{work_order_id}/pdf",
            headers=headers,
        )
        assert pdf.status_code == 200
        assert pdf.content.startswith(b"%PDF")
        rendered_text = "\n".join(
            page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf.content)).pages
        )
        assert expected_folio in rendered_text
        assert expected_instrument in rendered_text
        assert "TÉCNICO LAB" in rendered_text
        assert "CLIENTE LAB" in rendered_text
        assert "CLIENTE PRUEBA" in rendered_text
        assert rendered_text.count("AVENIDA EJEMPLO 123") == 1
        assert "PERSONA PRUEBA" in rendered_text
        assert "45601" in rendered_text
        assert "Tlaquepaque" in rendered_text
        assert "Jalisco" in rendered_text
        assert "OC-TEST-001" in rendered_text
        assert "Avenida Ejemplo 123, Tlaquepaque" not in rendered_text


def _create_anticipated_lab_group(
    client: TestClient, tokens: dict[str, str], *, quantity: int = 3
) -> tuple[dict, list[dict]]:
    created = client.post(
        "/api/lab-work-order-groups",
        json={**create_payload("Cohortes LAB"), "quantity": quantity},
        headers=auth(tokens["admin"]),
    )
    assert created.status_code == 201, created.text
    root = created.json()
    return root, root["related_work_orders"]


def test_partial_individual_then_open_group_closure_preserves_historical_group(
    lab_context,
):
    client, factory, tokens = lab_context
    headers = auth(tokens["tech"])
    root, related = _create_anticipated_lab_group(client, tokens)
    root_id, second_id, third_id = [item["id"] for item in related]
    root_endpoint = f"/api/mobile/v1/technician/lab-work-orders/{root_id}"
    second_endpoint = f"/api/mobile/v1/technician/lab-work-orders/{second_id}"
    third_endpoint = f"/api/mobile/v1/technician/lab-work-orders/{third_id}"

    added = client.post(
        f"{root_endpoint}/equipment", json=equipment_payload(1), headers=headers
    )
    assert added.status_code == 201, added.text
    group_sign = client.post(
        f"{root_endpoint}/signatures", json=signatures_payload(), headers=headers
    )
    assert group_sign.status_code == 409

    configure_default_services(client, headers, root_id)
    individual_sign = client.post(
        f"{root_endpoint}/signatures/individual",
        json=signatures_payload(),
        headers=headers,
    )
    assert individual_sign.status_code == 200, individual_sign.text
    session_a = individual_sign.json()["signature_session_id"]
    assert session_a is not None
    assert individual_sign.json()["signature_scope"] == "individual"
    assert client.post(f"{root_endpoint}/complete", headers=headers).status_code == 409
    assert client.get(f"{second_endpoint}/pdf", headers=headers).status_code == 409

    individual_complete = client.post(
        f"{root_endpoint}/complete/individual", headers=headers
    )
    assert individual_complete.status_code == 200, individual_complete.text
    completed_root = individual_complete.json()
    assert completed_root["status"] == "completed"
    assert client.post(
        f"{root_endpoint}/complete/individual", headers=headers
    ).status_code == 200
    assert client.post(
        f"{root_endpoint}/signatures/individual",
        json=signatures_payload(),
        headers=headers,
    ).status_code == 409
    assert client.patch(
        root_endpoint, json={"notes": "No debe editarse"}, headers=headers
    ).status_code == 409

    frozen = {
        "signature_session_id": completed_root["signature_session_id"],
        "final_pdf_sha256": completed_root["final_pdf_sha256"],
        "client_name": completed_root["client_name"],
        "root_work_order_id": completed_root["root_work_order_id"],
        "sequence_number": completed_root["sequence_number"],
    }
    frozen_pdf_at = datetime.fromisoformat(
        completed_root["final_pdf_generated_at"].replace("Z", "+00:00")
    ).replace(tzinfo=None)
    frozen_completed_at = datetime.fromisoformat(
        completed_root["completed_at"].replace("Z", "+00:00")
    ).replace(tzinfo=None)
    with factory() as db:
        frozen_pdf = db.get(LabWorkOrder, root_id).final_pdf

    second = client.get(second_endpoint, headers=headers).json()
    changed = client.patch(
        second_endpoint,
        json={
            "client_name": "Cohortes LAB abiertas",
            "expected_edit_version": second["edit_version"],
        },
        headers=headers,
    )
    assert changed.status_code == 200, changed.text
    assert client.post(
        f"{second_endpoint}/equipment",
        json={
            **equipment_payload(2),
            "expected_edit_version": changed.json()["edit_version"],
        },
        headers=headers,
    ).status_code == 201
    third = client.get(third_endpoint, headers=headers).json()
    assert client.post(
        f"{third_endpoint}/equipment",
        json={
            **equipment_payload(3),
            "expected_edit_version": third["edit_version"],
        },
        headers=headers,
    ).status_code == 201

    root_after_edits = client.get(root_endpoint, headers=headers).json()
    assert {key: root_after_edits[key] for key in frozen} == frozen
    assert datetime.fromisoformat(
        root_after_edits["final_pdf_generated_at"].replace("Z", "+00:00")
    ).replace(tzinfo=None) == frozen_pdf_at
    assert datetime.fromisoformat(
        root_after_edits["completed_at"].replace("Z", "+00:00")
    ).replace(tzinfo=None) == frozen_completed_at

    configure_default_services(client, headers, second_id)
    configure_default_services(client, headers, third_id)
    signed_open_cohort = client.post(
        f"{second_endpoint}/signatures",
        json=signatures_payload(),
        headers=headers,
    )
    assert signed_open_cohort.status_code == 200, signed_open_cohort.text
    session_b = signed_open_cohort.json()["signature_session_id"]
    assert session_b not in {None, session_a}
    assert signed_open_cohort.json()["signature_scope"] == "group"
    related_after_sign = signed_open_cohort.json()["related_work_orders"]
    assert [item["status"] for item in related_after_sign] == [
        "completed",
        "received_signed",
        "received_signed",
    ]
    assert [item["signature_session_id"] for item in related_after_sign] == [
        session_a,
        session_b,
        session_b,
    ]
    assert client.post(
        f"{second_endpoint}/complete/individual", headers=headers
    ).status_code == 409

    completed_open_cohort = client.post(
        f"{second_endpoint}/complete", headers=headers
    )
    assert completed_open_cohort.status_code == 200, completed_open_cohort.text
    assert all(
        item["status"] == "completed"
        for item in completed_open_cohort.json()["related_work_orders"]
    )
    root_final = client.get(root_endpoint, headers=headers).json()
    assert {key: root_final[key] for key in frozen} == frozen
    assert datetime.fromisoformat(
        root_final["final_pdf_generated_at"].replace("Z", "+00:00")
    ).replace(tzinfo=None) == frozen_pdf_at
    assert datetime.fromisoformat(
        root_final["completed_at"].replace("Z", "+00:00")
    ).replace(tzinfo=None) == frozen_completed_at

    with factory() as db:
        group = list(
            db.scalars(
                select(LabWorkOrder).order_by(LabWorkOrder.sequence_number)
            )
        )
        assert len(group) == 3
        assert {item.root_work_order_id for item in group} == {root_id}
        assert [item.signature_session_id for item in group] == [
            session_a,
            session_b,
            session_b,
        ]
        sessions = list(
            db.scalars(
                select(LabWorkOrderSignatureSession)
                .where(LabWorkOrderSignatureSession.root_work_order_id == root_id)
                .order_by(LabWorkOrderSignatureSession.version)
            )
        )
        assert [(item.id, item.version) for item in sessions] == [
            (session_a, 1),
            (session_b, 2),
        ]
        assert group[0].final_pdf == frozen_pdf
        audits = list(
            db.scalars(
                select(AuditLog).where(
                    AuditLog.action.in_(
                        (
                            "lab_work_order.individual_signed",
                            "lab_work_order.individual_completed",
                            "lab_work_order.group_signed",
                            "lab_work_order.group_completed",
                        )
                    )
                )
            )
        )
        assert {item.new_values["scope"] for item in audits} == {
            "individual",
            "group",
        }

    exported = client.get(
        "/api/mobile/v1/technician/lab-work-orders/export",
        headers=auth(tokens["admin"]),
    )
    assert exported.status_code == 200, exported.text
    with zipfile.ZipFile(io.BytesIO(exported.content)) as archive:
        names = archive.namelist()
        exported_orders = json.loads(archive.read("work_orders.json"))
    assert len([name for name in names if name.startswith("pdf/")]) == 3
    assert len(
        [
            name
            for name in names
            if name.startswith("signatures/session-") and name.endswith(".json")
        ]
    ) == 2
    assert len(
        [
            name
            for name in names
            if name.startswith("signatures/session-") and name.endswith(".png")
        ]
    ) == 4
    assert [item["signature_session_id"] for item in exported_orders] == [
        session_a,
        session_b,
        session_b,
    ]


def test_individual_closure_requires_equipment_and_reopens_only_its_cohort(
    lab_context,
):
    client, _factory, tokens = lab_context
    tech_headers = auth(tokens["tech"])
    root, related = _create_anticipated_lab_group(client, tokens)
    root_id, second_id, _third_id = [item["id"] for item in related]
    root_endpoint = f"/api/mobile/v1/technician/lab-work-orders/{root_id}"
    second_endpoint = f"/api/mobile/v1/technician/lab-work-orders/{second_id}"

    assert client.post(
        f"{root_endpoint}/signatures/individual",
        json=signatures_payload(),
        headers=tech_headers,
    ).status_code == 409
    client.post(
        f"{root_endpoint}/equipment", json=equipment_payload(1), headers=tech_headers
    )
    configure_default_services(client, tech_headers, root_id)
    assert client.post(
        f"{root_endpoint}/signatures/individual",
        json=signatures_payload(),
        headers=tech_headers,
    ).status_code == 200
    closed = client.post(
        f"{root_endpoint}/complete/individual", headers=tech_headers
    )
    assert closed.status_code == 200, closed.text

    second_before = client.get(second_endpoint, headers=tech_headers).json()
    ticket = client.post(
        "/api/mobile/v1/technician/tickets",
        json={
            "work_order_id": root_id,
            "reason": "Corrección individual",
            "description": "Reabrir sólo la cohorte individual cerrada.",
            "requested_signature_policy": "preserve",
        },
        headers=tech_headers,
    )
    assert ticket.status_code == 201, ticket.text
    approved = client.post(
        f"/api/mobile/v1/technician/tickets/{ticket.json()['id']}/approve",
        json={"signature_policy": "preserve"},
        headers=auth(tokens["admin"]),
    )
    assert approved.status_code == 200, approved.text

    reopened_root = client.get(root_endpoint, headers=tech_headers).json()
    second_after = client.get(second_endpoint, headers=tech_headers).json()
    assert reopened_root["status"] == "draft"
    assert reopened_root["revision_number"] == 2
    assert reopened_root["signature_preserved"] is True
    assert second_after["status"] == "draft"
    assert second_after["revision_number"] == second_before["revision_number"]
    assert second_after["edit_version"] == second_before["edit_version"]

    reclosed = client.post(
        f"{root_endpoint}/complete/individual", headers=tech_headers
    )
    assert reclosed.status_code == 200, reclosed.text
    assert reclosed.json()["status"] == "completed"


@pytest.mark.skipif(
    not os.getenv("LAB_POSTGRES_TEST_URL"),
    reason="requiere LAB_POSTGRES_TEST_URL para probar versiones concurrentes",
)
def test_postgresql_concurrent_individual_cohorts_get_distinct_versions(
    postgres_lab_context,
):
    client, factory, tokens = postgres_lab_context
    root, related = _create_anticipated_lab_group(client, tokens, quantity=2)
    ids = [item["id"] for item in related]
    for index, work_order_id in enumerate(ids, start=1):
        response = client.post(
            f"/api/mobile/v1/technician/lab-work-orders/{work_order_id}/equipment",
            json=equipment_payload(index),
            headers=auth(tokens["tech"]),
        )
        assert response.status_code == 201, response.text
    payload = LabSignatureGroupWrite(**signatures_payload())

    def sign(work_order_id: int) -> tuple[int, int]:
        with factory() as db:
            user = db.scalar(select(User).where(User.username == "postgres-tech"))
            assert user is not None
            result = sign_individual(db, work_order_id, payload, user)
            assert result.signature_session is not None
            return result.signature_session.id, result.signature_session.version

    with ThreadPoolExecutor(max_workers=2) as pool:
        sessions = sorted(pool.map(sign, ids), key=lambda item: item[1])

    assert [version for _session_id, version in sessions] == [1, 2]
    assert len({session_id for session_id, _version in sessions}) == 2
    assert root["root_work_order_id"] == root["id"]


def test_requires_equipment_and_both_valid_signatures(lab_context):
    client, _factory, tokens = lab_context
    headers = auth(tokens["tech"])
    root = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload(),
        headers=headers,
    ).json()
    endpoint = f"/api/mobile/v1/technician/lab-work-orders/{root['id']}"
    assert client.post(f"{endpoint}/signatures", json=signatures_payload(), headers=headers).status_code == 409
    client.post(f"{endpoint}/equipment", json=equipment_payload(1), headers=headers)
    configure_default_services(client, headers, root["id"])
    invalid = signatures_payload()
    invalid["client"]["signature_data_url"] = "data:image/png;base64,bm90LXBuZw=="
    assert client.post(f"{endpoint}/signatures", json=invalid, headers=headers).status_code == 422
    assert client.post(f"{endpoint}/complete", headers=headers).status_code == 409


def test_individual_endpoint_keeps_single_work_order_flow_available(lab_context):
    client, _factory, tokens = lab_context
    headers = auth(tokens["tech"])
    created = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload("OT individual"),
        headers=headers,
    ).json()
    endpoint = f"/api/mobile/v1/technician/lab-work-orders/{created['id']}"
    assert client.post(
        f"{endpoint}/equipment", json=equipment_payload(1), headers=headers
    ).status_code == 201
    configure_default_services(client, headers, created["id"])
    signed = client.post(
        f"{endpoint}/signatures/individual",
        json=signatures_payload(),
        headers=headers,
    )
    assert signed.status_code == 200, signed.text
    completed = client.post(f"{endpoint}/complete/individual", headers=headers)
    assert completed.status_code == 200, completed.text
    assert completed.json()["status"] == "completed"
    assert completed.json()["root_work_order_id"] == created["id"]


def test_folio_never_exceeds_6999(lab_context):
    client, factory, tokens = lab_context
    with factory() as db:
        db.add(
            InstitutionalFolioSequence(
                document_type="lab_work_order",
                prefix="LAB",
                year=0,
                next_value=6999,
            )
        )
        db.commit()
    headers = auth(tokens["tech"])
    first = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload(),
        headers=headers,
    )
    assert first.status_code == 201
    assert first.json()["folio"] == 6999
    exhausted = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload(),
        headers=headers,
    )
    assert exhausted.status_code == 409


def test_lab_certificate_folio_never_exceeds_7999(lab_context):
    client, factory, tokens = lab_context
    headers = auth(tokens["tech"])
    with factory() as db:
        db.add(
            InstitutionalFolioSequence(
                document_type="lab_certificate",
                prefix="MYCA",
                year=0,
                next_value=7999,
            )
        )
        db.commit()
    order = client.post(
        "/api/mobile/v1/technician/lab-work-orders", json=create_payload(), headers=headers
    )
    order_id = order.json()["id"]
    equipment_ids = []
    for index in range(1, 3):
        added = client.post(
            f"/api/mobile/v1/technician/lab-work-orders/{order_id}/equipment",
            json=equipment_payload(index),
            headers=headers,
        )
        equipment_ids.append(added.json()["equipment"][-1]["id"])
    first = client.put(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/equipment/{equipment_ids[0]}/service",
        json={"service_type": "accredited", "linked_company_id": None},
        headers=headers,
    )
    assert first.status_code == 200, first.text
    month_year = date.today().strftime("%m-%y")
    assert first.json()["equipment"][0]["certificate_folio"] == f"MYCA-{month_year}-7999"
    exhausted = client.put(
        f"/api/mobile/v1/technician/lab-work-orders/{order_id}/equipment/{equipment_ids[1]}/service",
        json={"service_type": "accredited", "linked_company_id": None},
        headers=headers,
    )
    assert exhausted.status_code == 409, exhausted.text


def test_postgresql_concurrent_lab_certificate_folio_allocation_is_unique():
    database_url = os.getenv("LAB_POSTGRES_TEST_URL")
    if not database_url:
        pytest.skip("requiere LAB_POSTGRES_TEST_URL para probar locks PostgreSQL reales")

    schema = f"lab_cert_lock_{uuid.uuid4().hex}"
    admin_engine = create_engine(database_url)
    with admin_engine.begin() as connection:
        connection.execute(text(f'CREATE SCHEMA "{schema}"'))
    engine = create_engine(database_url, connect_args={"options": f"-csearch_path={schema}"})
    factory = sessionmaker(bind=engine)
    Base.metadata.create_all(engine)

    from app.services.lab_work_orders import _allocate_lab_certificate_folio

    def allocate() -> str:
        with factory() as db:
            folio = _allocate_lab_certificate_folio(db, "MYCA")
            db.commit()
            return folio

    try:
        with ThreadPoolExecutor(max_workers=2) as pool:
            folios = sorted(pool.map(lambda _index: allocate(), range(2)))
        month_year = date.today().strftime("%m-%y")
        assert folios == [f"MYCA-{month_year}-4700", f"MYCA-{month_year}-4701"]
    finally:
        with admin_engine.begin() as connection:
            connection.execute(text(f'DROP SCHEMA "{schema}" CASCADE'))


def test_export_manifest_matches_persisted_counts(lab_context):
    client, _factory, tokens = lab_context
    tech_headers = auth(tokens["tech"])
    root = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload(),
        headers=tech_headers,
    ).json()
    client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/equipment",
        json=equipment_payload(1),
        headers=tech_headers,
    )
    assert client.get(
        "/api/mobile/v1/technician/lab-work-orders/export",
        headers=tech_headers,
    ).status_code == 403
    exported = client.get(
        "/api/mobile/v1/technician/lab-work-orders/export",
        headers=auth(tokens["admin"]),
    )
    assert exported.status_code == 200, exported.text
    with zipfile.ZipFile(io.BytesIO(exported.content)) as archive:
        manifest = json.loads(archive.read("manifest.json"))
        work_orders = json.loads(archive.read("work_orders.json"))
        equipment = json.loads(archive.read("equipment.json"))
    assert manifest["work_order_count"] == len(work_orders) == 1
    assert manifest["equipment_count"] == len(equipment) == 1


def test_work_order_structured_filters_are_combinable_paginated_and_protected(lab_context):
    client, _factory, tokens = lab_context
    url = "/api/mobile/v1/technician/lab-work-orders"
    headers = auth(tokens["tech"])
    for client_name in ("Susana Industrial", "SUSANA Metrología", "Cliente Distinto"):
        response = client.post(url, json=create_payload(client_name), headers=headers)
        assert response.status_code == 201

    assert client.get(url, params={"folio": "6401"}, headers=headers).json()[0]["folio"] == 6401
    assert {item["folio"] for item in client.get(url, params={"folio": "640"}, headers=headers).json()} == {
        6400, 6401, 6402
    }
    assert len(client.get(url, params={"client": "Susana Industrial"}, headers=headers).json()) == 1
    assert len(client.get(url, params={"client": "susana"}, headers=headers).json()) == 2
    combined = client.get(
        url,
        params={"folio": "6401", "client": "susana"},
        headers=headers,
    )
    assert [item["folio"] for item in combined.json()] == [6401]
    assert client.get(url, params={"client": "inexistente"}, headers=headers).json() == []
    first_page = client.get(url, params={"limit": 2, "offset": 0}, headers=headers).json()
    second_page = client.get(url, params={"limit": 2, "offset": 2}, headers=headers).json()
    assert [item["folio"] for item in first_page] == [6402, 6401]
    assert [item["folio"] for item in second_page] == [6400]
    assert client.get(url).status_code == 401
    assert client.get(url, headers=auth(tokens["capture"])).status_code == 200


def _completed_work_order(client: TestClient, token: str, name: str = "Cliente Ticket") -> dict:
    headers = auth(token)
    root = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload(name),
        headers=headers,
    ).json()
    endpoint = f"/api/mobile/v1/technician/lab-work-orders/{root['id']}"
    client.post(f"{endpoint}/equipment", json=equipment_payload(1), headers=headers)
    configure_default_services(client, headers, root["id"])
    signed = client.post(f"{endpoint}/signatures", json=signatures_payload(), headers=headers)
    assert signed.status_code == 200, signed.text
    completed = client.post(f"{endpoint}/complete", headers=headers)
    assert completed.status_code == 200, completed.text
    return completed.json()


def test_lab_delivery_is_independent_from_technical_closure(lab_context):
    client, _factory, tokens = lab_context
    headers = auth(tokens["tech"])
    created = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload("Cliente entrega"),
        headers=headers,
    )
    assert created.status_code == 201, created.text
    assert created.json()["departure_date"] is None
    forbidden_write = client.patch(
        f"/api/mobile/v1/technician/lab-work-orders/{created.json()['id']}",
        json={"departure_date": "2026-09-03"},
        headers=headers,
    )
    assert forbidden_write.status_code == 422

    completed = _completed_work_order(client, tokens["tech"], "Cliente entrega cerrada")
    assert completed["departure_date"] is None
    delivery_status = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{completed['id']}/delivery", headers=headers
    )
    assert delivery_status.status_code == 200
    assert delivery_status.json()["group_complete"] is False
    technical_pdf = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{completed['id']}/pdf",
        headers=headers,
    )
    assert technical_pdf.status_code == 200
    technical_text = "\n".join(
        page.extract_text() or "" for page in PdfReader(io.BytesIO(technical_pdf.content)).pages
    )
    assert "PENDIENTE" in technical_text


def test_ticket_preserves_minor_change_and_versions_pdf(lab_context):
    client, _factory, tokens = lab_context
    tech_headers = auth(tokens["tech"])
    admin_headers = auth(tokens["admin"])
    work_order = _completed_work_order(client, tokens["tech"])
    ticket_url = "/api/mobile/v1/technician/tickets"
    payload = {
        "work_order_id": work_order["id"],
        "reason": "Folio de certificado",
        "description": "Se recibió el folio después del cierre.",
        "requested_signature_policy": "preserve",
    }
    assert client.post(ticket_url, json=payload).status_code == 401
    assert client.post(ticket_url, json=payload, headers=auth(tokens["capture"])).status_code == 403
    created = client.post(ticket_url, json=payload, headers=tech_headers)
    assert created.status_code == 201, created.text
    ticket = created.json()
    assert ticket["status"] == "pending"
    still_closed = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}", headers=tech_headers
    ).json()
    assert still_closed["status"] == "completed"

    approved = client.post(
        f"{ticket_url}/{ticket['id']}/approve",
        json={"signature_policy": "preserve", "comment": "Cambio administrativo"},
        headers=admin_headers,
    )
    assert approved.status_code == 200, approved.text
    assert approved.json()["status"] == "in_progress"
    reopened = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}", headers=tech_headers
    ).json()
    assert reopened["revision_number"] == 2
    assert reopened["signature_preserved"] is True
    assert reopened["signature_session_id"] == work_order["signature_session_id"]

    changed = client.patch(
        f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}",
        json={"notes": "Folio de certificado CERT-2026-1", "expected_edit_version": reopened["edit_version"]},
        headers=tech_headers,
    )
    assert changed.status_code == 200, changed.text
    assert changed.json()["signature_preserved"] is True
    assert client.patch(
        f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}",
        json={"notes": "request obsoleto", "expected_edit_version": reopened["edit_version"]},
        headers=tech_headers,
    ).status_code == 409

    closed = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}/complete",
        headers=tech_headers,
    )
    assert closed.status_code == 200, closed.text
    assert closed.json()["revision_number"] == 2
    assert client.get(f"{ticket_url}/{ticket['id']}", headers=tech_headers).json()["status"] == "resolved"
    history = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}/revisions",
        headers=tech_headers,
    ).json()
    assert [item["revision_number"] for item in history] == [1, 2]
    historical_pdf = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}/revisions/1/pdf",
        headers=tech_headers,
    )
    assert historical_pdf.status_code == 200
    assert historical_pdf.content.startswith(b"%PDF")


def test_reopen_preserve_edit_general_critical_field_keeps_signature(lab_context):
    """MOB-001 CASO A: editar un campo general CRÍTICO (client_name) que ya
    existía antes de la reapertura, con requested_signature_policy=preserve,
    no debe invalidar la firma histórica ni exigir una nueva."""
    client, _factory, tokens = lab_context
    tech_headers = auth(tokens["tech"])
    admin_headers = auth(tokens["admin"])
    work_order = _completed_work_order(client, tokens["tech"], "Cliente Original")
    ticket = client.post(
        "/api/mobile/v1/technician/tickets",
        json={
            "work_order_id": work_order["id"],
            "reason": "Corrección de razón social",
            "description": "El cliente solicitó corregir su nombre legal.",
            "requested_signature_policy": "preserve",
        },
        headers=tech_headers,
    ).json()
    approved = client.post(
        f"/api/mobile/v1/technician/tickets/{ticket['id']}/approve",
        json={"signature_policy": "preserve"},
        headers=admin_headers,
    )
    assert approved.status_code == 200, approved.text

    endpoint = f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}"
    reopened = client.get(endpoint, headers=tech_headers).json()
    assert reopened["signature_required"] is False
    assert reopened["signature_preserved"] is True
    assert reopened["signature_session_id"] == work_order["signature_session_id"]

    changed = client.patch(
        endpoint,
        json={
            "client_name": "Cliente Original Corregido",
            "expected_edit_version": reopened["edit_version"],
        },
        headers=tech_headers,
    )
    assert changed.status_code == 200, changed.text
    body = changed.json()
    assert body["client_name"] == "Cliente Original Corregido"
    assert body["signature_preserved"] is True
    assert body["signature_required"] is False
    assert body["signature_session_id"] == work_order["signature_session_id"]

    closed = client.post(f"{endpoint}/complete", headers=tech_headers)
    assert closed.status_code == 200, closed.text
    closed_body = closed.json()
    assert closed_body["status"] == "completed"
    assert closed_body["revision_number"] == 2
    assert closed_body["signature_session_id"] == work_order["signature_session_id"]
    assert closed_body["signature_required"] is False


def test_reopen_preserve_edit_existing_equipment_keeps_signature(lab_context):
    """MOB-001 CASO B: editar datos de un equipo YA EXISTENTE (instrument,
    brand, identification, serial_number, is_good_condition) durante una
    reapertura preserve no debe invalidar la firma ni solicitar una nueva."""
    client, _factory, tokens = lab_context
    tech_headers = auth(tokens["tech"])
    admin_headers = auth(tokens["admin"])
    work_order = _completed_work_order(client, tokens["tech"], "Cliente Equipo")
    endpoint = f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}"
    equipment_id = client.get(endpoint, headers=tech_headers).json()["equipment"][0]["id"]

    ticket = client.post(
        "/api/mobile/v1/technician/tickets",
        json={
            "work_order_id": work_order["id"],
            "reason": "Corrección de número de serie",
            "description": "El número de serie capturado tenía un error de dedo.",
            "requested_signature_policy": "preserve",
        },
        headers=tech_headers,
    ).json()
    approved = client.post(
        f"/api/mobile/v1/technician/tickets/{ticket['id']}/approve",
        json={"signature_policy": "preserve"},
        headers=admin_headers,
    )
    assert approved.status_code == 200, approved.text
    reopened = client.get(endpoint, headers=tech_headers).json()
    assert reopened["signature_required"] is False

    changed = client.patch(
        f"{endpoint}/equipment/{equipment_id}",
        json={
            **equipment_payload(1, serial_number="SER-1-CORREGIDO"),
            "expected_edit_version": reopened["edit_version"],
        },
        headers=tech_headers,
    )
    assert changed.status_code == 200, changed.text
    body = changed.json()
    assert body["equipment"][0]["serial_number"] == "SER-1-CORREGIDO"
    assert body["signature_preserved"] is True
    assert body["signature_required"] is False
    assert body["signature_session_id"] == work_order["signature_session_id"]

    closed = client.post(f"{endpoint}/complete", headers=tech_headers)
    assert closed.status_code == 200, closed.text
    closed_body = closed.json()
    assert closed_body["status"] == "completed"
    assert closed_body["revision_number"] == 2
    assert closed_body["signature_session_id"] == work_order["signature_session_id"]
    assert closed_body["signature_required"] is False


def test_ready_for_signatures_remains_locked_until_formal_reopening(lab_context):
    """Una OT firmada no vuelve a draft por una edición ordinaria."""
    client, _factory, tokens = lab_context
    headers = auth(tokens["tech"])
    root = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload("Cliente Pre-Cierre"),
        headers=headers,
    ).json()
    endpoint = f"/api/mobile/v1/technician/lab-work-orders/{root['id']}"
    client.post(f"{endpoint}/equipment", json=equipment_payload(1), headers=headers)
    configure_default_services(client, headers, root["id"])
    signed = client.post(f"{endpoint}/signatures", json=signatures_payload(), headers=headers)
    assert signed.status_code == 200, signed.text
    signature_session_id = signed.json()["signature_session_id"]

    changed = client.patch(
        endpoint,
        json={"client_name": "Cliente Renombrado"},
        headers=headers,
    )
    assert changed.status_code == 409, changed.text
    unchanged = client.get(endpoint, headers=headers).json()
    assert unchanged["status"] == "received_signed"
    assert unchanged["signature_session_id"] == signature_session_id
    assert unchanged["signature_required"] is False


def test_structural_change_invalidates_signature_and_requires_new_signature(lab_context):
    client, _factory, tokens = lab_context
    tech_headers = auth(tokens["tech"])
    admin_headers = auth(tokens["admin"])
    work_order = _completed_work_order(client, tokens["tech"], "Cliente Estructural")
    ticket = client.post(
        "/api/mobile/v1/technician/tickets",
        json={
            "work_order_id": work_order["id"],
            "reason": "Agregar equipo",
            "description": "El cliente entregó un equipo adicional.",
            "requested_signature_policy": "preserve",
        },
        headers=tech_headers,
    ).json()
    reopened = client.post(
        f"/api/mobile/v1/technician/tickets/{ticket['id']}/approve",
        json={"signature_policy": "preserve"},
        headers=admin_headers,
    )
    assert reopened.status_code == 200
    detail = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}", headers=tech_headers
    ).json()
    changed = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}/equipment",
        json={**equipment_payload(2), "expected_edit_version": detail["edit_version"]},
        headers=tech_headers,
    )
    assert changed.status_code == 201, changed.text
    assert changed.json()["signature_required"] is True
    assert changed.json()["signature_session_id"] is None
    endpoint = f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}"
    assert client.post(f"{endpoint}/complete", headers=tech_headers).status_code == 409
    configure_default_services(client, tech_headers, work_order["id"])
    resigned = client.post(
        f"{endpoint}/signatures", json=signatures_payload(), headers=tech_headers
    )
    assert resigned.status_code == 200, resigned.text
    assert resigned.json()["signature_session_id"] != work_order["signature_session_id"]
    assert client.post(f"{endpoint}/complete", headers=tech_headers).status_code == 200


def test_ticket_rejection_and_invalid_lifecycle(lab_context):
    client, _factory, tokens = lab_context
    tech_headers = auth(tokens["tech"])
    admin_headers = auth(tokens["admin"])
    open_order = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload("Abierta"),
        headers=tech_headers,
    ).json()
    payload = {
        "work_order_id": open_order["id"],
        "reason": "No aplica",
        "description": "La orden todavía está abierta.",
        "requested_signature_policy": "invalidate",
    }
    assert client.post(
        "/api/mobile/v1/technician/tickets", json=payload, headers=tech_headers
    ).status_code == 409
    closed = _completed_work_order(client, tokens["tech"], "Para rechazo")
    payload["work_order_id"] = closed["id"]
    ticket = client.post(
        "/api/mobile/v1/technician/tickets", json=payload, headers=tech_headers
    ).json()
    rejected = client.post(
        f"/api/mobile/v1/technician/tickets/{ticket['id']}/reject",
        json={"comment": "No se acreditó la necesidad"},
        headers=admin_headers,
    )
    assert rejected.status_code == 200
    assert rejected.json()["status"] == "rejected"
    assert client.post(
        f"/api/mobile/v1/technician/tickets/{ticket['id']}/approve",
        json={"signature_policy": "invalidate"},
        headers=admin_headers,
    ).status_code == 409


@pytest.mark.parametrize("signature_policy", ["preserve", "invalidate"])
def test_postgresql_ticket_approval_locks_only_ticket_row(
    postgres_lab_context, signature_policy
):
    client, _factory, tokens = postgres_lab_context
    work_order = _completed_work_order(
        client, tokens["tech"], f"PostgreSQL {signature_policy}"
    )
    ticket = client.post(
        "/api/mobile/v1/technician/tickets",
        json={
            "work_order_id": work_order["id"],
            "reason": "Corrección posterior al cierre",
            "description": "Validación del bloqueo de ticket en PostgreSQL.",
            "requested_signature_policy": signature_policy,
        },
        headers=auth(tokens["tech"]),
    ).json()

    approved = client.post(
        f"/api/mobile/v1/technician/tickets/{ticket['id']}/approve",
        json={"signature_policy": signature_policy},
        headers=auth(tokens["admin"]),
    )

    assert approved.status_code == 200, approved.text
    assert approved.json()["status"] == "in_progress"
    reopened = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{work_order['id']}",
        headers=auth(tokens["tech"]),
    ).json()
    assert reopened["signature_preserved"] is (signature_policy == "preserve")
    assert reopened["signature_required"] is (signature_policy == "invalidate")
    assert reopened["signature_session_id"] == (
        work_order["signature_session_id"] if signature_policy == "preserve" else None
    )


def test_postgresql_ticket_rejection_cannot_be_resolved_again(postgres_lab_context):
    client, _factory, tokens = postgres_lab_context
    work_order = _completed_work_order(client, tokens["tech"], "PostgreSQL rechazo")
    ticket = client.post(
        "/api/mobile/v1/technician/tickets",
        json={
            "work_order_id": work_order["id"],
            "reason": "Solicitud improcedente",
            "description": "Se valida el rechazo definitivo del ticket.",
            "requested_signature_policy": "invalidate",
        },
        headers=auth(tokens["tech"]),
    ).json()
    endpoint = f"/api/mobile/v1/technician/tickets/{ticket['id']}"

    rejected = client.post(
        f"{endpoint}/reject",
        json={"comment": "No procede la reapertura"},
        headers=auth(tokens["admin"]),
    )

    assert rejected.status_code == 200, rejected.text
    assert rejected.json()["status"] == "rejected"
    assert client.post(
        f"{endpoint}/reject",
        json={"comment": "Segundo rechazo"},
        headers=auth(tokens["admin"]),
    ).status_code == 409
    assert client.post(
        f"{endpoint}/approve",
        json={"signature_policy": "preserve"},
        headers=auth(tokens["admin"]),
    ).status_code == 409


def test_postgresql_concurrent_ticket_resolution_allows_one_winner(
    postgres_lab_context,
):
    client, factory, tokens = postgres_lab_context
    work_order = _completed_work_order(client, tokens["tech"], "PostgreSQL carrera")
    ticket = client.post(
        "/api/mobile/v1/technician/tickets",
        json={
            "work_order_id": work_order["id"],
            "reason": "Resolución concurrente",
            "description": "Dos revisores intentan resolver el mismo ticket.",
            "requested_signature_policy": "preserve",
        },
        headers=auth(tokens["tech"]),
    ).json()

    def resolve(action: str) -> tuple[str, int]:
        with factory() as db:
            admin = db.scalar(select(User).where(User.username == "postgres-admin"))
            assert admin is not None
            try:
                if action == "approve":
                    approve_reopen_ticket(
                        db,
                        ticket["id"],
                        TicketReview(signature_policy="preserve"),
                        admin,
                    )
                else:
                    reject_ticket(
                        db,
                        ticket["id"],
                        TicketReject(comment="Resolución concurrente rechazada"),
                        admin,
                    )
                return action, 200
            except HTTPException as exc:
                db.rollback()
                return action, exc.status_code

    with ThreadPoolExecutor(max_workers=2) as pool:
        outcomes = list(pool.map(resolve, ("approve", "reject")))

    assert sorted(status for _action, status in outcomes) == [200, 409]
    detail = client.get(
        f"/api/mobile/v1/technician/tickets/{ticket['id']}",
        headers=auth(tokens["admin"]),
    ).json()
    assert detail["status"] in {"in_progress", "rejected"}


def test_lab_pdf_leaves_missing_purchase_order_empty():
    payload = create_payload("CLIENTE SIN ORDEN")
    payload["purchase_order"] = None
    payload["reception_date"] = date.fromisoformat(payload["reception_date"])
    work_order = LabWorkOrder(
        folio=6400,
        sequence_number=1,
        created_by_user_id=1,
        status="draft",
        **payload,
    )
    work_order.equipment = [
        LabWorkOrderEquipment(position=1, **equipment_payload(1))
    ]

    pdf, _filename = generate_lab_work_order_pdf(work_order)
    rendered_text = "\n".join(
        page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf)).pages
    )

    assert "ORDEN DE COMPRA /COTIZACIÓN" in rendered_text
    assert "ORDEN DE COMPRA /COTIZACIÓN 0" not in rendered_text


def test_lab_pdf_uses_certificate_folio_and_preserves_legacy_precedence():
    payload = create_payload("CLIENTE CON FOLIO")
    payload["reception_date"] = date.fromisoformat(payload["reception_date"])
    work_order = LabWorkOrder(
        folio=6401,
        sequence_number=1,
        created_by_user_id=1,
        status="draft",
        **payload,
    )
    equipment = LabWorkOrderEquipment(position=1, **equipment_payload(1))
    equipment.certificate_folio = "VIN-2026-001"
    equipment.report_number = "REPORTE-LEGACY"
    work_order.equipment = [equipment]
    pdf, _filename = generate_lab_work_order_pdf(work_order)
    rendered_text = "\n".join(
        page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf)).pages
    )
    assert "VIN-2026-001" in rendered_text

    base = {
        "name": "Equipo",
        "brand": "Marca",
        "internal_id": "ID",
        "serial_number": "Serie",
        "is_good_condition": True,
    }
    preferred = _build_equipment_lines([
        SimpleNamespace(
            **base,
            certificate_folio="CERT-PRIORITARIO",
            report_number="REPORTE",
            certificates=[SimpleNamespace(is_active=True, expected_folio="CERTIFICADO", folio=None)],
        )
    ])[0]
    report_legacy = _build_equipment_lines([
        SimpleNamespace(**base, report_number="REPORTE", certificates=[])
    ])[0]
    certificate_legacy = _build_equipment_lines([
        SimpleNamespace(
            **base,
            report_number=None,
            certificates=[SimpleNamespace(is_active=True, expected_folio=None, folio="CERTIFICADO")],
        )
    ])[0]
    assert preferred.certificate_folio == "CERT-PRIORITARIO"
    assert report_legacy.certificate_folio == "REPORTE"
    assert certificate_legacy.certificate_folio == "CERTIFICADO"


@pytest.mark.skipif(
    not os.getenv("LAB_POSTGRES_TEST_URL"),
    reason="requiere PostgreSQL temporal para probar advisory/row lock real",
)
def test_postgresql_concurrent_folio_allocation_is_unique():
    engine = create_engine(os.environ["LAB_POSTGRES_TEST_URL"])
    factory = sessionmaker(bind=engine)

    def allocate() -> int:
        with factory() as db:
            with db.begin():
                value = _allocate_folio(db)
                time.sleep(0.1)
                return value

    with ThreadPoolExecutor(max_workers=2) as pool:
        values = sorted(pool.map(lambda _index: allocate(), range(2)))

    assert values == [6400, 6401]

    with factory() as db:
        role = Role(name="Tecnico LAB concurrente", description="Prueba LAB")
        user = User(
            username="lab-concurrent-tech",
            email="lab-concurrent-tech@example.test",
            full_name="Técnico LAB concurrente",
            hashed_password="unused",
            account_type="internal",
            status="active",
            is_active=True,
            role=role,
            roles=[role],
        )
        db.add(user)
        db.flush()
        root = LabWorkOrder(
            folio=6402,
            sequence_number=1,
            created_by_user_id=user.id,
            **create_payload(),
        )
        db.add(root)
        db.flush()
        root.root_work_order_id = root.id
        root.equipment = [
            LabWorkOrderEquipment(position=index, **equipment_payload(index))
            for index in range(1, 11)
        ]
        db.commit()
        root_id = root.id
        user_id = user.id

    def create_additional() -> tuple[str, int]:
        with factory() as db:
            user = db.get(User, user_id)
            assert user is not None
            try:
                result = create_additional_work_order(db, root_id, user)
                return "created", result.folio
            except HTTPException as exc:
                db.rollback()
                return "rejected", exc.status_code

    with ThreadPoolExecutor(max_workers=2) as pool:
        outcomes = list(pool.map(lambda _index: create_additional(), range(2)))

    assert sorted(outcomes) == [("created", 6403), ("rejected", 409)]
    with factory() as db:
        folios = list(db.scalars(select(LabWorkOrder.folio).order_by(LabWorkOrder.folio)))
    assert folios == [6402, 6403]


def test_delete_lab_work_order_is_admin_only_and_removes_exclusive_data(lab_context):
    client, factory, tokens = lab_context
    root = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload("Eliminar LAB"),
        headers=auth(tokens["tech"]),
    ).json()
    client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/equipment",
        json=equipment_payload(1),
        headers=auth(tokens["tech"]),
    )
    endpoint = f"/api/mobile/v1/technician/lab-work-orders/{root['id']}"

    assert client.delete(endpoint, headers=auth(tokens["tech"])).status_code == 403
    assert client.delete(endpoint, headers=auth(tokens["capture"])).status_code == 403
    assert client.delete(endpoint, headers=auth(tokens["admin"])).status_code == 204
    assert client.delete(endpoint, headers=auth(tokens["admin"])).status_code == 404

    with factory() as db:
        assert db.get(LabWorkOrder, root["id"]) is None
        assert db.scalar(
            select(func.count(LabWorkOrderEquipment.id)).where(
                LabWorkOrderEquipment.work_order_id == root["id"]
            )
        ) == 0
        audit = db.scalar(
            select(AuditLog).where(
                AuditLog.action == "lab_work_order.deleted",
                AuditLog.entity_id == root["id"],
            )
        )
        assert audit is not None
        assert audit.previous_values["folio"] == 6400


def _signed_lab_group(client: TestClient, token: str) -> tuple[dict, dict, int]:
    headers = auth(token)
    root = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload("Grupo compartido"),
        headers=headers,
    ).json()
    for index in range(1, 11):
        client.post(
            f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/equipment",
            json=equipment_payload(index),
            headers=headers,
        )
    extra = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/additional",
        headers=headers,
    ).json()
    client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{extra['id']}/equipment",
        json=equipment_payload(11),
        headers=headers,
    )
    configure_default_services(client, headers, root["id"])
    configure_default_services(client, headers, extra["id"])
    signed = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/signatures",
        json=signatures_payload(),
        headers=headers,
    ).json()
    return root, extra, signed["signature_session_id"]


def test_delete_additional_lab_order_preserves_root_and_shared_signatures(lab_context):
    client, factory, tokens = lab_context
    root, extra, signature_session_id = _signed_lab_group(client, tokens["tech"])

    response = client.delete(
        f"/api/mobile/v1/technician/lab-work-orders/{extra['id']}",
        headers=auth(tokens["admin"]),
    )
    assert response.status_code == 204, response.text
    remaining = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}",
        headers=auth(tokens["tech"]),
    ).json()
    assert [item["id"] for item in remaining["related_work_orders"]] == [root["id"]]
    assert remaining["signature_session_id"] == signature_session_id

    with factory() as db:
        assert db.get(LabWorkOrder, extra["id"]) is None
        assert db.get(LabWorkOrderSignatureSession, signature_session_id) is not None

    response = client.delete(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}",
        headers=auth(tokens["admin"]),
    )
    assert response.status_code == 204, response.text
    with factory() as db:
        assert db.get(LabWorkOrder, root["id"]) is None
        assert db.get(LabWorkOrderSignatureSession, signature_session_id) is None


def test_delete_root_lab_order_reparents_group_and_shared_session(lab_context):
    client, factory, tokens = lab_context
    root, extra, signature_session_id = _signed_lab_group(client, tokens["tech"])

    response = client.delete(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}",
        headers=auth(tokens["admin"]),
    )
    assert response.status_code == 204, response.text
    remaining = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{extra['id']}",
        headers=auth(tokens["tech"]),
    ).json()
    assert remaining["root_work_order_id"] == extra["id"]
    assert remaining["previous_work_order_id"] is None
    assert remaining["sequence_number"] == 1
    assert [item["id"] for item in remaining["related_work_orders"]] == [extra["id"]]

    with factory() as db:
        session = db.get(LabWorkOrderSignatureSession, signature_session_id)
        assert session is not None
        assert session.root_work_order_id == extra["id"]


def _approved_lab_group_request(
    client: TestClient,
    factory: sessionmaker,
    tokens: dict[str, str],
    *,
    quantity: int = 2,
) -> tuple[dict, dict]:
    with factory() as db:
        operator = Client(
            legal_name="Operador histórico LAB",
            commercial_name="Operador histórico LAB",
        )
        db.add(operator)
        db.flush()
        requester = db.scalar(select(User).where(User.username == "lab-tech"))
        assert requester is not None
        request = create_group_request(
            db,
            LabWorkOrderGroupCreate(
                **create_payload("Cliente histórico LAB"), quantity=quantity
            ),
            requester,
            operator_client_id=operator.id,
        )
        request_id = request.id

    headers = auth(tokens["admin"])
    claimed = client.post(
        f"/api/lab-work-order-groups/requests/{request_id}/claim", headers=headers
    )
    assert claimed.status_code == 200, claimed.text
    approved = client.post(
        f"/api/lab-work-order-groups/requests/{request_id}/approve", headers=headers
    )
    assert approved.status_code == 200, approved.text
    root = client.get(
        f"/api/mobile/v1/technician/lab-work-orders/{approved.json()['root_work_order_id']}",
        headers=auth(tokens["tech"]),
    )
    assert root.status_code == 200, root.text
    return approved.json(), root.json()


def test_delete_requested_group_additional_then_root_preserves_request_history_and_folios(
    lab_context,
):
    client, factory, tokens = lab_context
    request, root = _approved_lab_group_request(client, factory, tokens)
    additional = root["related_work_orders"][1]
    request_id = request["id"]
    conversation_id = request["conversation_id"]
    requested_by_user_id = request["requested_by_user_id"]
    handled_by_user_id = request["handled_by_user_id"]
    created_at = request["created_at"]
    decided_at = request["decided_at"]

    additional_delete = client.delete(
        f"/api/mobile/v1/technician/lab-work-orders/{additional['id']}",
        headers=auth(tokens["admin"]),
    )
    assert additional_delete.status_code == 204, additional_delete.text
    with factory() as db:
        stored = db.get(LabWorkOrderGroupRequest, request_id)
        assert stored is not None
        assert stored.root_work_order_id == root["id"]

    root_delete = client.delete(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}",
        headers=auth(tokens["admin"]),
    )
    assert root_delete.status_code == 204, root_delete.text

    with factory() as db:
        stored = db.get(LabWorkOrderGroupRequest, request_id)
        assert stored is not None
        assert stored.root_work_order_id is None
        assert stored.status == "approved"
        assert stored.conversation_id == conversation_id
        assert stored.requested_by_user_id == requested_by_user_id
        assert stored.handled_by_user_id == handled_by_user_id
        assert stored.quantity == 2
        assert stored.client_name == "Cliente histórico LAB"
        assert stored.created_at.isoformat() == created_at
        assert stored.decided_at.isoformat() == decided_at
        assert db.get(CommunicationConversation, conversation_id) is not None
        materialization = db.scalar(
            select(AuditLog).where(
                AuditLog.action == "lab_work_order.group_materialized",
                AuditLog.entity_id == root["id"],
            )
        )
        assert materialization is not None
        assert materialization.new_values["folios"] == [6400, 6401]

    next_group = client.post(
        "/api/lab-work-order-groups",
        json={**create_payload("Folio posterior"), "quantity": 1},
        headers=auth(tokens["admin"]),
    )
    assert next_group.status_code == 201, next_group.text
    assert next_group.json()["folio"] == 6402


def test_delete_requested_group_root_reparents_request_then_nulls_it_on_last_delete(
    lab_context,
):
    client, factory, tokens = lab_context
    request, root = _approved_lab_group_request(client, factory, tokens)
    replacement = root["related_work_orders"][1]

    root_delete = client.delete(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}",
        headers=auth(tokens["admin"]),
    )
    assert root_delete.status_code == 204, root_delete.text

    with factory() as db:
        stored = db.get(LabWorkOrderGroupRequest, request["id"])
        survivor = db.get(LabWorkOrder, replacement["id"])
        assert stored is not None
        assert stored.root_work_order_id == replacement["id"]
        assert stored.status == "approved"
        assert survivor is not None
        assert survivor.root_work_order_id == replacement["id"]
        assert survivor.previous_work_order_id is None
        assert survivor.sequence_number == 1

    last_delete = client.delete(
        f"/api/mobile/v1/technician/lab-work-orders/{replacement['id']}",
        headers=auth(tokens["admin"]),
    )
    assert last_delete.status_code == 204, last_delete.text
    with factory() as db:
        stored = db.get(LabWorkOrderGroupRequest, request["id"])
        assert stored is not None
        assert stored.root_work_order_id is None
        assert stored.status == "approved"
        assert stored.conversation_id == request["conversation_id"]


def test_delete_middle_lab_order_repairs_chain_without_deleting_sisters(lab_context):
    client, factory, tokens = lab_context
    headers = auth(tokens["tech"])
    root = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload("Cadena LAB"),
        headers=headers,
    ).json()
    for index in range(1, 11):
        client.post(
            f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/equipment",
            json=equipment_payload(index),
            headers=headers,
        )
    middle = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/additional",
        headers=headers,
    ).json()
    for index in range(11, 21):
        client.post(
            f"/api/mobile/v1/technician/lab-work-orders/{middle['id']}/equipment",
            json=equipment_payload(index),
            headers=headers,
        )
    last = client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{middle['id']}/additional",
        headers=headers,
    ).json()

    response = client.delete(
        f"/api/mobile/v1/technician/lab-work-orders/{middle['id']}",
        headers=auth(tokens["admin"]),
    )
    assert response.status_code == 204, response.text

    with factory() as db:
        remaining = list(db.scalars(select(LabWorkOrder).order_by(LabWorkOrder.sequence_number)))
        assert [item.id for item in remaining] == [root["id"], last["id"]]
        assert [item.sequence_number for item in remaining] == [1, 2]
        assert remaining[1].previous_work_order_id == root["id"]
        assert {item.root_work_order_id for item in remaining} == {root["id"]}


def test_delete_completed_single_lab_order_removes_inline_pdf_and_signature_session(lab_context):
    client, factory, tokens = lab_context
    completed = _completed_work_order(client, tokens["tech"], "Finalizada eliminable")
    signature_session_id = completed["signature_session_id"]

    response = client.delete(
        f"/api/mobile/v1/technician/lab-work-orders/{completed['id']}",
        headers=auth(tokens["admin"]),
    )
    assert response.status_code == 204, response.text
    with factory() as db:
        assert db.get(LabWorkOrder, completed["id"]) is None
        assert db.get(LabWorkOrderSignatureSession, signature_session_id) is None


def test_delete_root_preserves_shared_ticket_revision_and_repairs_notification(lab_context):
    client, factory, tokens = lab_context
    root, extra, _signature_session_id = _signed_lab_group(client, tokens["tech"])
    client.post(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}/complete",
        headers=auth(tokens["tech"]),
    )
    ticket = client.post(
        "/api/mobile/v1/technician/tickets",
        json={
            "work_order_id": root["id"],
            "reason": "Corrección compartida",
            "description": "Conservar la historia de la OT hermana.",
            "requested_signature_policy": "preserve",
        },
        headers=auth(tokens["tech"]),
    ).json()
    approved = client.post(
        f"/api/mobile/v1/technician/tickets/{ticket['id']}/approve",
        json={"signature_policy": "preserve"},
        headers=auth(tokens["admin"]),
    )
    assert approved.status_code == 200, approved.text

    response = client.delete(
        f"/api/mobile/v1/technician/lab-work-orders/{root['id']}",
        headers=auth(tokens["admin"]),
    )
    assert response.status_code == 204, response.text

    with factory() as db:
        stored_ticket = db.get(OperationalTicket, ticket["id"])
        assert stored_ticket is not None
        assert stored_ticket.work_order_id == extra["id"]
        revisions = list(
            db.scalars(
                select(LabWorkOrderRevision).where(
                    LabWorkOrderRevision.reopen_ticket_id == ticket["id"]
                )
            )
        )
        assert {revision.work_order_id for revision in revisions} == {extra["id"]}
        notifications = list(
            db.scalars(
                select(Notification).where(
                    Notification.entity_type == "ticket",
                    Notification.entity_id == ticket["id"],
                )
            )
        )
        assert notifications
        assert all(
            notification.metadata_json["work_order_id"] == extra["id"]
            for notification in notifications
        )


def test_delete_lab_work_order_rolls_back_on_commit_failure(lab_context, monkeypatch):
    client, factory, tokens = lab_context
    created = client.post(
        "/api/mobile/v1/technician/lab-work-orders",
        json=create_payload("Rollback LAB"),
        headers=auth(tokens["tech"]),
    ).json()

    with factory() as db:
        admin = db.scalar(
            select(User).where(User.email == "lab-admin@example.test")
        )
        assert admin is not None

        def fail_commit() -> None:
            raise RuntimeError("fallo simulado")

        monkeypatch.setattr(db, "commit", fail_commit)
        with pytest.raises(HTTPException) as exc_info:
            delete_work_order(db, created["id"], admin)
        assert exc_info.value.status_code == 409

    with factory() as db:
        assert db.get(LabWorkOrder, created["id"]) is not None
