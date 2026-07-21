import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.services.master_template_fingerprints import (
    canonical_service_type,
    detect_service_type,
)


def build_master(*, institutional_block: str, dynamic_folio: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Temperatura TLD"
    sheet.print_area = "A1:K80"
    sheet.merge_cells("C2:I2")
    sheet.merge_cells("C6:H6")
    sheet.merge_cells("E15:K16")
    sheet["C2"] = "METROLOGIA Y SERVICIOS MYC S.A. DE C.V."
    sheet["C6"] = "CERTIFICADO DE CALIBRACIÓN"
    sheet["D7"] = "Calibration Certificate"
    sheet["I7"] = institutional_block
    sheet["B10"] = "Orden de trabajo:"
    sheet["H10"] = "Folio:"
    sheet["I10"] = dynamic_folio
    sheet["A12"] = "Datos del usuario / User data"
    sheet["E17"] = "Instrumento a calibrar"
    sheet["E18"] = "Instrument under calibration"
    sheet["E35"] = "Instrumento de referencia"
    sheet["E36"] = "Reference instrument"
    sheet["J33"] = "=J31+365"
    for coordinate in ("C2", "C6", "D7", "I7", "A12", "E17", "E18", "E35", "E36"):
        sheet[coordinate].font = Font(name="Arial", bold=True)
        sheet[coordinate].fill = PatternFill("solid", fgColor="D9EAF7")
    clients = workbook.create_sheet("Registro de clientes")
    clients["B15"] = "Cliente"
    clients["C15"] = "contacto"
    clients["D15"] = "direccion"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class MasterTemplateFingerprintTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.reference_path = Path(self.temp_dir.name) / "master.xlsx"

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_canonical_service_types_do_not_use_document_labels(self):
        self.assertEqual(canonical_service_type("accredited_iso_17025"), "accredited")
        self.assertEqual(canonical_service_type("traceable"), "traceable")
        self.assertIsNone(canonical_service_type("Acreditación / Accreditation: 88795"))

    def test_accredited_template_survives_changed_text_and_accreditation_number(self):
        self.reference_path.write_bytes(build_master(
            institutional_block="Acreditación / Accreditation: 88795",
            dynamic_folio="MYCA-OLD",
        ))
        candidate = build_master(
            institutional_block="Reconocimiento institucional vigente: 99999",
            dynamic_folio="MYCA-07-2026-0001",
        )
        result = detect_service_type(
            candidate,
            extension=".xlsx",
            expected_template_path=self.reference_path,
            expected_service_type="accredited_iso_17025",
        )
        self.assertEqual(result["status"], "coincide")
        self.assertEqual(result["detected"], "accredited")
        self.assertGreaterEqual(result["template_match"]["evidence_groups"], 3)

    def test_registered_traceable_template_is_detected_without_literal_search(self):
        raw = build_master(institutional_block="Declaración institucional", dynamic_folio="MYCT-BASE")
        self.reference_path.write_bytes(raw)
        result = detect_service_type(
            build_master(institutional_block="Texto documental actualizado", dynamic_folio="MYCT-2026-1"),
            extension=".xlsx",
            expected_template_path=self.reference_path,
            expected_service_type="traceable",
        )
        self.assertEqual(result["status"], "coincide")
        self.assertEqual(result["detected"], "traceable")

    def test_unrelated_workbook_is_a_blocking_mismatch(self):
        self.reference_path.write_bytes(build_master(
            institutional_block="Acreditación / Accreditation: 88795",
            dynamic_folio="MYCA-BASE",
        ))
        unrelated = Workbook()
        unrelated.active.title = "Inventario"
        unrelated.active["A1"] = "Archivo distinto"
        buffer = BytesIO()
        unrelated.save(buffer)
        result = detect_service_type(
            buffer.getvalue(),
            extension=".xlsx",
            expected_template_path=self.reference_path,
            expected_service_type="accredited_iso_17025",
        )
        self.assertEqual(result["status"], "mismatch")
        self.assertIsNone(result["detected"])

    def test_detector_does_not_modify_workbook_payload(self):
        raw = build_master(institutional_block="Acreditación 88795", dynamic_folio="MYCA-1")
        self.reference_path.write_bytes(raw)
        detect_service_type(
            raw,
            extension=".xlsx",
            expected_template_path=self.reference_path,
            expected_service_type="accredited_iso_17025",
        )
        workbook = load_workbook(BytesIO(raw), data_only=False)
        self.assertEqual(workbook["Temperatura TLD"]["I7"].value, "Acreditación 88795")


if __name__ == "__main__":
    unittest.main()
