import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.services.sat_catalogs.comparison import compare_catalog_sources
from app.services.sat_catalogs.sat_xls_source import SatXlsSourceError, extract_catalog_rows


class SatOfficialXlsSourceTests(unittest.TestCase):
    def write_workbook(self, directory: str, sheet_name: str, rows: list[list[object]]) -> Path:
        source = Path(directory) / "sat_fixture.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for row in rows:
            sheet.append(row)
        workbook.save(source)
        workbook.close()
        return source

    def test_detects_real_header_preserves_zero_padding_and_ignores_blank_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            source = self.write_workbook(directory, "c_FormaPago", [
                ["Catálogo de formas de pago"],
                [],
                ["c_FormaPago", "Descripción", "Fecha inicio de vigencia", "Fecha fin de vigencia"],
                [3, "Transferencia", "2022-01-01", None],
                [None, None, None, None],
            ])
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "c_FormaPago"
            sheet.append(["Catálogo de formas de pago"])
            sheet.append([])
            sheet.append(["c_FormaPago", "Descripción", "Fecha inicio de vigencia", "Fecha fin de vigencia"])
            sheet.append([3, "Transferencia", "2022-01-01", None])
            sheet["A4"].number_format = "00"
            workbook.save(source)
            workbook.close()
            rows, sheets = extract_catalog_rows(source, "payment_forms")
        self.assertEqual(sheets[0]["header_row"], 3)
        self.assertEqual(rows, [{"code": "03", "name": "Transferencia", "valid_from": "2022-01-01", "valid_until": None, "c_formapago": "03", "descripcion": "Transferencia", "fecha_inicio_de_vigencia": "2022-01-01"}])

    def test_tax_rate_rows_use_maximum_for_fixed_rate_and_disambiguate_repeated_rates(self):
        with tempfile.TemporaryDirectory() as directory:
            source = self.write_workbook(directory, "c_TasaOCuota", [
                ["Rango o Fijo", "c_TasaOCuota", None, "Impuesto", "Factor", "Traslado", "Retención", "Fecha inicio de vigencia"],
                [None, "Valor mínimo", "Valor máximo"],
                ["Fijo", None, 0.16, "IVA", "Tasa", "Sí", "No", "2022-01-01"],
                ["Fijo", None, 0.16, "IEPS", "Tasa", "Sí", "No", "2022-01-01"],
            ])
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "c_TasaOCuota"
            sheet.append(["Rango o Fijo", "c_TasaOCuota", None, "Impuesto", "Factor", "Traslado", "Retención", "Fecha inicio de vigencia"])
            sheet.append([None, "Valor mínimo", "Valor máximo"])
            sheet.append(["Fijo", None, 0.16, "IVA", "Tasa", "Sí", "No", "2022-01-01"])
            sheet.append(["Fijo", None, 0.16, "IEPS", "Tasa", "Sí", "No", "2022-01-01"])
            sheet["C3"].number_format = "0.000000"
            sheet["C4"].number_format = "0.000000"
            workbook.save(source)
            workbook.close()
            rows, _ = extract_catalog_rows(source, "tax_rates")
        self.assertEqual(rows[0]["code"], "0.160000")
        self.assertEqual(rows[1]["code"], "0.160000 | IEPS | Tasa | Sí | No")

    def test_comparison_reports_source_only_catalog_and_sqlite_difference(self):
        with tempfile.TemporaryDirectory() as directory:
            source = self.write_workbook(directory, "c_Moneda", [
                ["Catálogo de moneda"],
                ["c_Moneda", "Descripción", "Fecha inicio de vigencia"],
                ["MXN", "Peso mexicano", "2022-01-01"],
                ["USD", "Dólar americano", "2022-01-01"],
            ])
            sqlite_source = Path(directory) / "catalogs.db"
            connection = sqlite3.connect(sqlite_source)
            connection.execute("CREATE TABLE cfdi_40_monedas (id TEXT, texto TEXT, vigencia_desde TEXT)")
            connection.execute("INSERT INTO cfdi_40_monedas VALUES ('MXN', 'Peso mexicano', '2022-01-01')")
            connection.commit()
            connection.close()
            report = compare_catalog_sources(source, sqlite_source, "currencies")
        self.assertEqual(report["result"], "different")
        self.assertEqual(report["missing_in_sqlite"], ["USD"])

    def test_rejects_missing_required_sheet(self):
        with tempfile.TemporaryDirectory() as directory:
            source = self.write_workbook(directory, "otra", [["clave"]])
            with self.assertRaises(SatXlsSourceError):
                extract_catalog_rows(source, "voucher_types")
