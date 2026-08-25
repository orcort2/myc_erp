from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from unittest.mock import patch
import zipfile

import pytest
from fastapi import HTTPException
from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from starlette.datastructures import UploadFile

import app.models  # noqa: F401
from app.core.db import Base
from app.models.audit_log import AuditLog
from app.models.certificate import Certificate, CertificateCaptureFile, CertificatePdfVersion
from app.models.client import Client
from app.models.controlled_document import (
    ControlledDocument,
    ControlledDocumentVersion,
    DocumentInterpretation,
)
from app.models.equipment import Equipment
from app.models.field_sheet import FieldSheet
from app.models.service_order import ServiceOrderItem
from app.models.user import User
from app.schemas.certificate import CertificateCreate, CertificateUpdate
from app.schemas.equipment import EquipmentCreate, EquipmentUpdate
from app.schemas.service_order import ServiceOrderCreate, ServiceOrderItemCreate
from app.services.certificates import (
    ALLOWED_TRANSITIONS,
    create_certificate,
    quality_approve,
    release_to_client,
    send_to_quality,
    start_capture,
    update_certificate,
)
from app.services import certificate_authentication
from app.services.capture_packages import (
    _registered_verification_master_match,
    upload_capture_files,
)
from app.services.catalog_items import _prepare_values
from app.services.equipment import create_equipment, update_equipment
from app.services.institutional_folios import build_certificate_folio
from app.services.folio_engine import suggest_certificate_folio
from app.services.service_orders import create_service_order


@pytest.fixture()
def db():
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, expire_on_commit=False)
    with factory() as session:
        yield session
    engine.dispose()


@pytest.fixture()
def mixed_order(db):
    actor = User(
        username="verification-actor",
        email="verification@example.test",
        full_name="Verification Actor",
        hashed_password="unused",
    )
    client = Client(legal_name="Cliente metrológico")
    db.add_all([actor, client])
    db.commit()
    order = create_service_order(
        db,
        ServiceOrderCreate(
            client_id=client.id,
            requires_payment=False,
            items=[
                ServiceOrderItemCreate(
                    service_name="Calibración trazable",
                    operational_category="calibration",
                    calibration_scope="traceable",
                    quantity=1,
                ),
                ServiceOrderItemCreate(
                    service_name="Verificación dimensional",
                    operational_category="verification",
                    calibration_scope=None,
                    quantity=1,
                ),
            ],
        ),
        user_id=actor.id,
    )
    items = {item.operational_category: item for item in order.items}
    return actor, order, items


def test_verification_ets_has_work_order_equipment_and_distinct_certificate(db, mixed_order):
    actor, order, items = mixed_order
    assert len(order.work_orders) == 1
    assert order.work_orders[0].equipment_limit == 10

    verification = create_equipment(
        db,
        EquipmentCreate(
            service_order_id=order.id,
            service_order_item_id=items["verification"].id,
            calibration_scope=None,
            name="Micrómetro",
        ),
        user_id=actor.id,
    )
    calibration = create_equipment(
        db,
        EquipmentCreate(
            service_order_id=order.id,
            service_order_item_id=items["calibration"].id,
            calibration_scope="traceable",
            name="Termómetro",
        ),
        user_id=actor.id,
    )

    verification_certificate = db.scalar(
        select(Certificate).where(Certificate.equipment_id == verification.id)
    )
    calibration_certificate = db.scalar(
        select(Certificate).where(Certificate.equipment_id == calibration.id)
    )
    assert verification.calibration_scope is None
    assert verification.service_order_item_id == items["verification"].id
    assert verification_certificate.certificate_type == "verification"
    assert verification_certificate.title == "Certificado de Verificación"
    assert verification_certificate.folio.startswith("MYCV-")
    assert calibration_certificate.certificate_type == "trazable"
    assert calibration_certificate.folio.startswith("MYCT")
    with pytest.raises(HTTPException, match="no admite edición manual"):
        update_certificate(
            db,
            verification_certificate.id,
            CertificateUpdate(expected_folio="MANUAL-1"),
            user_id=actor.id,
        )


def test_verification_only_ets_creates_its_own_equipment_and_certificate(db):
    actor = User(
        username="verification-only",
        email="verification-only@example.test",
        full_name="Verification Only",
        hashed_password="unused",
    )
    client = Client(legal_name="Cliente sólo Verificación")
    db.add_all([actor, client])
    db.commit()
    order = create_service_order(
        db,
        ServiceOrderCreate(
            client_id=client.id,
            requires_payment=False,
            items=[
                ServiceOrderItemCreate(
                    service_name="Verificación única",
                    operational_category="verification",
                    calibration_scope=None,
                    quantity=1,
                )
            ],
        ),
        user_id=actor.id,
    )
    equipment = create_equipment(
        db,
        EquipmentCreate(service_order_id=order.id, name="Regla patrón"),
        user_id=actor.id,
    )
    certificate = db.scalar(select(Certificate).where(Certificate.equipment_id == equipment.id))
    assert equipment.service_order_item_id == order.items[0].id
    assert equipment.calibration_scope is None
    assert certificate.certificate_type == "verification"
    assert certificate.title == "Certificado de Verificación"


def test_mixed_ets_requires_exact_item_and_verification_rejects_calibration_scope(db, mixed_order):
    actor, order, items = mixed_order
    with pytest.raises(HTTPException, match="Selecciona la partida metrológica"):
        create_equipment(
            db,
            EquipmentCreate(service_order_id=order.id, name="Sin proceso"),
            user_id=actor.id,
        )
    with pytest.raises(HTTPException, match="Verificación no admite alcance"):
        create_equipment(
            db,
            EquipmentCreate(
                service_order_id=order.id,
                service_order_item_id=items["verification"].id,
                calibration_scope="traceable",
                name="Proceso falso",
            ),
            user_id=actor.id,
        )


def test_verification_rejects_calibration_certificate_types(db, mixed_order):
    actor, order, items = mixed_order
    equipment = create_equipment(
        db,
        EquipmentCreate(
            service_order_id=order.id,
            service_order_item_id=items["verification"].id,
            name="Balanza",
        ),
        user_id=actor.id,
    )
    for certificate_type in ("acreditado", "trazable", "vinculado"):
        with pytest.raises(HTTPException, match="no corresponde al proceso"):
            create_certificate(
                db,
                CertificateCreate(
                    service_order_id=order.id,
                    equipment_id=equipment.id,
                    certificate_type=certificate_type,
                ),
                user_id=actor.id,
            )


def test_verification_folio_format_and_institutional_sequence(db):
    first = build_certificate_folio(
        db, service_type="verification", issued_on=date(2026, 8, 24)
    )
    second = build_certificate_folio(
        db, service_type="verification", issued_on=date(2026, 8, 24)
    )
    third = build_certificate_folio(
        db, service_type="verification", issued_on=date(2026, 9, 1)
    )
    next_cycle = build_certificate_folio(
        db, service_type="verification", issued_on=date(2027, 1, 2)
    )
    assert first == "MYCV-08-26-0001"
    assert second == "MYCV-08-26-0002"
    assert third == "MYCV-09-26-0003"
    assert next_cycle == "MYCV-01-27-0001"
    suggested = suggest_certificate_folio(
        db,
        certificate_type="verification",
        issued_on=date(2026, 8, 24),
        sequence=9123,
    )
    assert suggested.suggested_folio == "MYCV-08-26-9123"
    with pytest.raises(HTTPException, match="secuencia institucional"):
        suggest_certificate_folio(
            db,
            certificate_type="verification",
            issued_on=date(2026, 8, 24),
            manual_folio="MYCV-MANUAL",
        )


def test_verification_catalog_preserves_generic_master_without_scope():
    values = _prepare_values(
        {
            "item_type": "service",
            "service_kind": "simple",
            "commodity": "verification",
            "operational_category": "verification",
            "category": "Verificacion",
            "name": "Verificación genérica",
            "origin_currency": "MXN",
            "exchange_rate": 1,
            "margin_percent": 0,
            "expected_certificate_master_id": 44,
            "calibration_scope": None,
        }
    )
    assert values["expected_certificate_master_id"] == 44
    assert values["calibration_scope"] is None


def test_verification_uses_shared_capture_quality_authentication_release_and_versions(db, mixed_order):
    actor, order, items = mixed_order
    equipment = create_equipment(
        db,
        EquipmentCreate(
            service_order_id=order.id,
            service_order_item_id=items["verification"].id,
            name="Vernier",
        ),
        user_id=actor.id,
    )
    certificate = db.scalar(select(Certificate).where(Certificate.equipment_id == equipment.id))
    sheet = FieldSheet(
        equipment_id=equipment.id,
        work_order_id=equipment.work_order_id,
        work_order_number=equipment.work_order_number,
        status="completed",
        calibration_date=date(2026, 8, 24),
    )
    db.add(sheet)
    db.flush()
    certificate.field_sheet_id = sheet.id
    equipment.certificate_master_document_id = 1
    equipment.certificate_master_version_id = 1
    equipment.certificate_template_path_snapshot = "masters/verification.xlsx"
    db.add(
        CertificateCaptureFile(
            certificate_id=certificate.id,
            service_order_id=order.id,
            original_filename="Master_MYCV-08-26-0001.xlsx",
            stored_path="capture/Master_MYCV-08-26-0001.xlsx",
            identification_status="identified",
            validation_results={},
            uploaded_by_id=actor.id,
        )
    )
    db.commit()

    assert start_capture(db, certificate.id, user_id=actor.id).status == "capture_in_progress"
    assert send_to_quality(db, certificate.id, user_id=actor.id).status == "quality_review"
    assert quality_approve(db, certificate.id, user_id=actor.id).status == "quality_approved"
    assert "authenticated" in ALLOWED_TRANSITIONS["quality_approved"]
    assert "released_to_client" in ALLOWED_TRANSITIONS["authenticated"]

    def authenticate(_db, item, *, user_id, origin):
        item.status = "authenticated"
        item.authentication_code = "MYC-AUTH-VERIFICATION-1"
        item.authenticated_pdf_path = "certificates/verification-authenticated.pdf"
        item.authenticated_by_id = user_id
        return item

    with patch.object(
        certificate_authentication,
        "_authenticate_certificate_pdf",
        side_effect=authenticate,
    ):
        authenticated = certificate_authentication.authenticate_certificate(
            db,
            certificate.id,
            user_id=actor.id,
            origin="quality",
        )
    assert authenticated.status == "authenticated"
    assert authenticated.certificate_type == "verification"
    with patch("app.services.certificates._authenticated_document_exists", return_value=True):
        released = release_to_client(db, certificate.id, user_id=actor.id)
    assert released.status == "released_to_client"

    now = datetime.now(timezone.utc)
    db.add_all(
        [
            CertificatePdfVersion(
                certificate_id=certificate.id,
                version_number=1,
                file_path="certificates/verification-v1.pdf",
                uploaded_at=now,
                is_current=False,
            ),
            CertificatePdfVersion(
                certificate_id=certificate.id,
                version_number=2,
                file_path="certificates/verification-v2.pdf",
                uploaded_at=now,
                is_current=True,
            ),
        ]
    )
    db.commit()
    versions = db.scalars(
        select(CertificatePdfVersion)
        .where(CertificatePdfVersion.certificate_id == certificate.id)
        .order_by(CertificatePdfVersion.version_number)
    ).all()
    assert [version.version_number for version in versions] == [1, 2]
    assert [version.is_current for version in versions] == [False, True]


def _active_master(db, *, code, path):
    document = ControlledDocument(
        code=code,
        name=code,
        document_type="certificate_master",
        status="active",
    )
    db.add(document)
    db.flush()
    version = ControlledDocumentVersion(
        document_id=document.id,
        revision="1",
        file_path=str(path),
        original_filename=path.name,
        status="active",
    )
    db.add(version)
    db.flush()
    return document, version


def _master_selection_audit_count(db, equipment_id):
    return len(
        db.scalars(
            select(AuditLog.id).where(
                AuditLog.action == "equipment.certificate_master_selected",
                AuditLog.entity_id == equipment_id,
            )
        ).all()
    )


def test_verification_rejects_ambiguous_registered_master_fingerprint(
    db,
    mixed_order,
    tmp_path,
):
    actor, order, items = mixed_order
    generic_workbook = Workbook()
    generic_sheet = generic_workbook.active
    generic_sheet.title = "Master genérico"
    generic_sheet["A1"] = "Referencia descargable"
    generic_path = tmp_path / "verification-generic.xlsx"
    generic_workbook.save(generic_path)
    generic_workbook.close()
    generic, _ = _active_master(db, code="MYC-VER-GENERIC", path=generic_path)
    items["verification"].expected_certificate_master_id = generic.id
    db.commit()
    equipment = create_equipment(
        db,
        EquipmentCreate(
            service_order_id=order.id,
            service_order_item_id=items["verification"].id,
            name="Equipo ambiguo",
        ),
        user_id=actor.id,
    )
    certificate = db.scalar(select(Certificate).where(Certificate.equipment_id == equipment.id))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Certificado"
    sheet["A1"] = "Estructura técnica de Verificación"
    first_path = tmp_path / "verification-a.xlsx"
    second_path = tmp_path / "verification-b.xlsx"
    workbook.save(first_path)
    workbook.save(second_path)
    raw = first_path.read_bytes()
    workbook.close()

    documents = []
    for code, path in (("MYC-VER-A", first_path), ("MYC-VER-B", second_path)):
        document, version = _active_master(db, code=code, path=path)
        documents.append(document)
        db.add(
            DocumentInterpretation(
                document_id=document.id,
                document_version_id=version.id,
                name=f"Identidad {code}",
                interpretation_type="certificate_template_source",
                service_type="verification",
                status="approved",
                version=1,
            )
        )
    db.commit()

    with patch(
        "app.services.capture_packages.resolve_storage_path",
        side_effect=lambda value: Path(value) if value else None,
    ):
        document_id, template_path, validation = _registered_verification_master_match(
            db,
            raw,
            ".xlsx",
        )

    assert document_id is None
    assert template_path is None
    assert validation["status"] == "ambiguous"
    assert validation["candidate_document_ids"] == sorted(document.id for document in documents)

    with patch(
        "app.services.capture_packages.resolve_storage_path",
        side_effect=lambda value: Path(value) if value else None,
    ), patch("app.services.capture_packages.save_validated_content") as save:
        save.return_value.relative_path = "capture/ambiguous.xlsx"
        save.return_value.checksum_sha256 = "ambiguous"
        result = upload_capture_files(
            db,
            order.id,
            [
                UploadFile(
                    filename=f"Master_{certificate.folio}.xlsx",
                    file=BytesIO(raw),
                )
            ],
            user_id=actor.id,
        )

    db.refresh(equipment)
    assert result["summary"]["unidentified"] == 1
    assert result["processed"][0]["validation"]["master_registration"]["status"] == "ambiguous"
    assert equipment.certificate_master_document_id == generic.id
    assert "final_certificate_master_document_id" not in equipment.certificate_operational_context_snapshot


def test_verification_freezes_generic_then_specific_master_and_item_identity(db, mixed_order, tmp_path):
    actor, order, items = mixed_order
    generic_path = tmp_path / "generic.xlsx"
    final_path = tmp_path / "specific.xlsx"
    generic_path.touch()
    final_path.touch()
    generic, generic_version = _active_master(db, code="MYC-VER-GEN", path=generic_path)
    specific, specific_version = _active_master(db, code="MYC-VER-ESP", path=final_path)
    items["verification"].expected_certificate_master_id = generic.id
    db.commit()

    equipment = create_equipment(
        db,
        EquipmentCreate(
            service_order_id=order.id,
            service_order_item_id=items["verification"].id,
            name="Comparador",
        ),
        user_id=actor.id,
    )
    assert equipment.certificate_master_document_id == generic.id
    assert equipment.certificate_master_version_id == generic_version.id
    assert equipment.certificate_operational_context_snapshot["initial_certificate_master_document_id"] == generic.id

    equipment = update_equipment(
        db,
        equipment.id,
        EquipmentUpdate(certificate_master_document_id=specific.id),
        user_id=actor.id,
    )
    context = equipment.certificate_operational_context_snapshot
    assert equipment.certificate_master_document_id == specific.id
    assert equipment.certificate_master_version_id == specific_version.id
    assert context["initial_certificate_master_document_id"] == generic.id
    assert context["final_certificate_master_document_id"] == specific.id
    assert context["final_certificate_master_version_id"] == specific_version.id
    assert context["certificate_master_selection_history"][0]["previous_document_id"] == generic.id

    history_count = len(context["certificate_master_selection_history"])
    audit_count = _master_selection_audit_count(db, equipment.id)
    equipment = update_equipment(
        db,
        equipment.id,
        EquipmentUpdate(certificate_master_document_id=specific.id),
        user_id=actor.id,
    )
    assert len(equipment.certificate_operational_context_snapshot["certificate_master_selection_history"]) == history_count
    assert _master_selection_audit_count(db, equipment.id) == audit_count

    with pytest.raises(HTTPException, match="Master final de Verificación ya está congelado"):
        update_equipment(
            db,
            equipment.id,
            EquipmentUpdate(certificate_master_document_id=generic.id),
            user_id=actor.id,
        )
    db.refresh(equipment)
    assert equipment.certificate_master_document_id == specific.id
    assert equipment.certificate_master_version_id == specific_version.id

    with pytest.raises(HTTPException, match="partida metrológica.*no puede cambiarse"):
        update_equipment(
            db,
            equipment.id,
            EquipmentUpdate(service_order_item_id=items["calibration"].id),
            user_id=actor.id,
        )


def test_verification_identified_legacy_capture_still_blocks_master_change(
    db,
    mixed_order,
    tmp_path,
):
    actor, order, items = mixed_order
    equipment = create_equipment(
        db,
        EquipmentCreate(
            service_order_id=order.id,
            service_order_item_id=items["verification"].id,
            name="Equipo legacy con evidencia",
        ),
        user_id=actor.id,
    )
    certificate = db.scalar(select(Certificate).where(Certificate.equipment_id == equipment.id))
    db.add(
        CertificateCaptureFile(
            certificate_id=certificate.id,
            service_order_id=order.id,
            original_filename="evidencia-legacy.xlsx",
            identification_status="identified",
            validation_results={},
            uploaded_by_id=actor.id,
        )
    )
    candidate_path = tmp_path / "legacy-candidate.xlsx"
    candidate_path.touch()
    candidate, _ = _active_master(db, code="MYC-VER-LEGACY-CANDIDATE", path=candidate_path)
    db.commit()

    with pytest.raises(HTTPException, match="no puede cambiar después de identificar"):
        update_equipment(
            db,
            equipment.id,
            EquipmentUpdate(certificate_master_document_id=candidate.id),
            user_id=actor.id,
        )


def test_verification_replaces_generic_in_bundle_and_auto_freezes_registered_real_master(
    db, mixed_order, tmp_path
):
    actor, order, items = mixed_order
    generic_workbook = Workbook()
    generic_sheet = generic_workbook.active
    generic_sheet.title = "Master genérico"
    generic_sheet["A1"] = "Referencia genérica"
    generic_path = tmp_path / "master-generico.xlsx"
    generic_workbook.save(generic_path)
    generic_workbook.close()
    generic, generic_version = _active_master(
        db,
        code="MYC-VER-GEN-CAPTURA",
        path=generic_path,
    )
    items["verification"].expected_certificate_master_id = generic.id
    db.commit()
    equipment = create_equipment(
        db,
        EquipmentCreate(
            service_order_id=order.id,
            service_order_item_id=items["verification"].id,
            name="Vernier",
            internal_id="VER-REAL-01",
            serial_number="SN-REAL-01",
        ),
        user_id=actor.id,
    )
    certificate = db.scalar(select(Certificate).where(Certificate.equipment_id == equipment.id))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Certificado"
    sheet["A1"] = "Certificado de Verificación"
    sheet["A2"] = certificate.folio
    sheet["A3"] = equipment.internal_id
    sheet["A4"] = equipment.serial_number
    sheet["B6"] = 0
    master_path = tmp_path / "master-especifico.xlsx"
    workbook.save(master_path)
    specific, specific_version = _active_master(
        db,
        code="MYC-VER-ESP-CAPTURA",
        path=master_path,
    )
    db.add(
        DocumentInterpretation(
            document_id=specific.id,
            document_version_id=specific_version.id,
            name="Master técnico estructurado de Verificación",
            interpretation_type="certificate_template_source",
            service_type="verification",
            status="approved",
            version=1,
        )
    )
    db.commit()
    sheet["B6"] = 12.34
    real_file = BytesIO()
    workbook.save(real_file)
    workbook.close()

    package = BytesIO()
    delivery_name = f"Master_{certificate.folio}.xlsx"
    with zipfile.ZipFile(package, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            f"ETS-{order.id}/OT-{equipment.work_order_number:04d}/{certificate.folio}/{delivery_name}",
            real_file.getvalue(),
        )
    package.seek(0)

    with patch(
        "app.services.capture_packages.resolve_storage_path",
        side_effect=lambda value: Path(value) if value else None,
    ), patch("app.services.capture_packages.save_validated_content") as save:
        save.return_value.relative_path = "capture/real-B.xlsx"
        save.return_value.checksum_sha256 = "abc123"
        result = upload_capture_files(
            db,
            order.id,
            [UploadFile(filename="bonche-verificacion.zip", file=package)],
            user_id=actor.id,
        )

    db.refresh(equipment)
    assert result["summary"]["identified"] == 1
    assert result["processed"][0]["certificate_id"] == certificate.id
    assert result["processed"][0]["filename"] == delivery_name
    assert result["processed"][0]["validation"]["servicio"]["method"] == "template_fingerprint"
    assert result["processed"][0]["validation"]["servicio"]["status"] == "coincide"
    assert result["processed"][0]["validation"]["master_registration"]["document_id"] == specific.id
    assert equipment.certificate_master_document_id == specific.id
    assert equipment.certificate_master_version_id == specific_version.id
    context = equipment.certificate_operational_context_snapshot
    assert context["initial_certificate_master_document_id"] == generic.id
    assert context["final_certificate_master_document_id"] == specific.id
    assert context["certificate_master_selection_history"][-1]["selection_source"] == "capture_upload_fingerprint"
    assert generic_version.id != specific_version.id


def _verification_with_frozen_master(db, mixed_order, tmp_path):
    actor, order, items = mixed_order
    generic_book = Workbook()
    generic_book.active.title = "Master genérico"
    generic_book.active["A1"] = "Referencia inicial"
    generic_path = tmp_path / "generic-authority.xlsx"
    generic_book.save(generic_path)
    generic_book.close()
    generic, _ = _active_master(db, code="MYC-VER-AUTH-GEN", path=generic_path)
    items["verification"].expected_certificate_master_id = generic.id
    db.commit()

    equipment = create_equipment(
        db,
        EquipmentCreate(
            service_order_id=order.id,
            service_order_item_id=items["verification"].id,
            name="Equipo autoridad final",
            internal_id="AUTH-01",
            serial_number="AUTH-SN-01",
        ),
        user_id=actor.id,
    )
    certificate = db.scalar(select(Certificate).where(Certificate.equipment_id == equipment.id))
    master_book = Workbook()
    master_sheet = master_book.active
    master_sheet.title = "Certificado A"
    master_sheet["A1"] = "Certificado de Verificación A"
    master_sheet["A2"] = certificate.folio
    master_sheet["A3"] = equipment.internal_id
    master_sheet["A4"] = equipment.serial_number
    master_sheet["B6"] = 0
    master_path = tmp_path / "master-final-a.xlsx"
    master_book.save(master_path)
    master, version = _active_master(db, code="MYC-VER-AUTH-A", path=master_path)
    db.add(
        DocumentInterpretation(
            document_id=master.id,
            document_version_id=version.id,
            name="Master final A de Verificación",
            interpretation_type="certificate_template_source",
            service_type="verification",
            status="approved",
            version=1,
        )
    )
    db.commit()
    equipment = update_equipment(
        db,
        equipment.id,
        EquipmentUpdate(certificate_master_document_id=master.id),
        user_id=actor.id,
    )
    master_sheet["B6"] = 10
    real_file = BytesIO()
    master_book.save(real_file)
    master_book.close()
    return actor, order, equipment, certificate, master, version, master_path, real_file.getvalue()


def _upload_verification_file(db, actor, order, certificate, raw):
    with patch(
        "app.services.capture_packages.resolve_storage_path",
        side_effect=lambda value: Path(value) if value else None,
    ), patch(
        "app.services.capture_packages._registered_verification_master_match",
        side_effect=AssertionError("No debe resolverse de nuevo un Master final congelado"),
    ), patch("app.services.capture_packages.save_validated_content") as save:
        save.return_value.relative_path = "capture/final-authority.xlsx"
        save.return_value.checksum_sha256 = "final-authority"
        return upload_capture_files(
            db,
            order.id,
            [
                UploadFile(
                    filename=f"Master_{certificate.folio}.xlsx",
                    file=BytesIO(raw),
                )
            ],
            user_id=actor.id,
        )


def test_verification_final_master_is_idempotent_and_ignores_other_active_candidates(
    db,
    mixed_order,
    tmp_path,
):
    actor, order, equipment, certificate, master_a, version_a, master_path, raw_a = (
        _verification_with_frozen_master(db, mixed_order, tmp_path)
    )
    master_b_path = tmp_path / "master-candidate-b.xlsx"
    master_b_path.write_bytes(master_path.read_bytes())
    master_b, version_b = _active_master(db, code="MYC-VER-AUTH-B", path=master_b_path)
    db.add(
        DocumentInterpretation(
            document_id=master_b.id,
            document_version_id=version_b.id,
            name="Master candidato B de Verificación",
            interpretation_type="certificate_template_source",
            service_type="verification",
            status="approved",
            version=1,
        )
    )
    db.commit()
    context_before = dict(equipment.certificate_operational_context_snapshot)
    history_count = len(context_before["certificate_master_selection_history"])
    audit_count = _master_selection_audit_count(db, equipment.id)

    result = _upload_verification_file(db, actor, order, certificate, raw_a)

    db.refresh(equipment)
    assert result["summary"]["identified"] == 1
    assert result["processed"][0]["validation"]["master_registration"] == {
        "status": "frozen_snapshot",
        "method": "equipment_final_snapshot",
        "document_id": master_a.id,
        "version_id": version_a.id,
    }
    assert equipment.certificate_master_document_id == master_a.id
    assert equipment.certificate_master_version_id == version_a.id
    assert len(equipment.certificate_operational_context_snapshot["certificate_master_selection_history"]) == history_count
    assert _master_selection_audit_count(db, equipment.id) == audit_count


def test_verification_file_for_other_master_cannot_replace_frozen_final(
    db,
    mixed_order,
    tmp_path,
):
    actor, order, equipment, certificate, master_a, version_a, _, _ = (
        _verification_with_frozen_master(db, mixed_order, tmp_path)
    )
    master_b_book = Workbook()
    master_b_sheet = master_b_book.active
    master_b_sheet.title = "Certificado B"
    master_b_sheet["A1"] = "Certificado de Verificación B"
    master_b_sheet["A2"] = certificate.folio
    master_b_sheet["A3"] = equipment.internal_id
    master_b_sheet["A4"] = equipment.serial_number
    master_b_sheet["B9"] = 0
    master_b_path = tmp_path / "master-real-b.xlsx"
    master_b_book.save(master_b_path)
    master_b, version_b = _active_master(db, code="MYC-VER-REAL-B", path=master_b_path)
    db.add(
        DocumentInterpretation(
            document_id=master_b.id,
            document_version_id=version_b.id,
            name="Master real B de Verificación",
            interpretation_type="certificate_template_source",
            service_type="verification",
            status="approved",
            version=1,
        )
    )
    db.commit()
    master_b_sheet["B9"] = 20
    real_b = BytesIO()
    master_b_book.save(real_b)
    master_b_book.close()

    result = _upload_verification_file(
        db,
        actor,
        order,
        certificate,
        real_b.getvalue(),
    )

    db.refresh(equipment)
    assert result["summary"]["unidentified"] == 1
    assert result["processed"][0]["validation"]["servicio"]["status"] == "mismatch"
    assert result["processed"][0]["validation"]["master_registration"]["document_id"] == master_a.id
    assert equipment.certificate_master_document_id == master_a.id
    assert equipment.certificate_master_version_id == version_a.id


def test_verification_frozen_final_version_survives_new_active_revision(
    db,
    mixed_order,
    tmp_path,
):
    actor, order, equipment, certificate, master_a, version_a, _, raw_a = (
        _verification_with_frozen_master(db, mixed_order, tmp_path)
    )
    revision_book = Workbook()
    revision_book.active.title = "Nueva revisión"
    revision_book.active["A1"] = "Estructura posterior incompatible"
    revision_path = tmp_path / "master-final-a-rev2.xlsx"
    revision_book.save(revision_path)
    revision_book.close()
    version_a.status = "obsolete"
    version_a2 = ControlledDocumentVersion(
        document_id=master_a.id,
        revision="2",
        file_path=str(revision_path),
        original_filename=revision_path.name,
        status="active",
    )
    db.add(version_a2)
    master_a.current_revision = "2"
    db.commit()

    result = _upload_verification_file(db, actor, order, certificate, raw_a)

    db.refresh(equipment)
    assert result["summary"]["identified"] == 1
    assert equipment.certificate_master_document_id == master_a.id
    assert equipment.certificate_master_version_id == version_a.id
    assert equipment.certificate_master_version_id != version_a2.id
    assert result["processed"][0]["validation"]["master_registration"]["version_id"] == version_a.id
