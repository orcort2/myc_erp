import shutil
import tempfile
import inspect
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook
from pypdf import PdfReader

from app.core.db import get_db
from app.main import app
from app.models.certificate import Certificate, CertificatePdfVersion
from app.routers.certificates import authenticate_certificate as authenticate_certificate_endpoint
from app.security.api_access import enforce_api_access
from app.services.certificate_authentication import (
    _convert_master_to_pdf,
    authenticate_certificate_pdf,
)


def approved_certificate(**overrides):
    values = {
        "id": 1,
        "status": "quality_approved",
        "expected_folio": "MYCA-07-2026-0001",
        "folio": "MYCA-07-2026-0001",
        "final_pdf_path": None,
        "final_pdf_original_filename": None,
        "final_pdf_uploaded_at": None,
        "final_pdf_uploaded_by_id": None,
        "authenticated_pdf_path": None,
        "authentication_code": None,
        "authentication_hash": None,
        "authenticated_pdf_generated_at": None,
        "authenticated_by_id": None,
        "verification_url": None,
        "match_status": "pending",
        "pdf_versions": [],
        "updated_at": datetime(2026, 7, 21, tzinfo=timezone.utc),
    }
    values.update(overrides)
    return SimpleNamespace(**values)


def identified_master(path: Path):
    return SimpleNamespace(
        id=10,
        certificate_id=1,
        identification_status="identified",
        original_filename="Master_MYCA-07-2026-0001.xlsx",
        stored_path="capture/1/10-Master_MYCA-07-2026-0001.xlsx",
        created_at=datetime(2026, 7, 21, tzinfo=timezone.utc),
        path=path,
    )


def test_authentication_requires_quality_approval_only():
    db = MagicMock()
    with pytest.raises(HTTPException, match="aprobados por calidad"):
        authenticate_certificate_pdf(db, approved_certificate(status="quality_review"), user_id=4)


def test_approved_master_generates_and_authenticates_without_pdf_or_match():
    db = MagicMock()
    certificate = approved_certificate(match_status="pending", final_pdf_path=None)
    with tempfile.TemporaryDirectory() as temporary:
        root = Path(temporary)
        master = root / "master.xlsx"
        master.write_bytes(b"xlsx")
        generated = root / "generated.pdf"
        generated.write_bytes(b"generated-pdf")
        final_target = root / "10-Master_MYCA-07-2026-0001.pdf"
        authenticated_target = root / "10-Master_MYCA-07-2026-0001_authenticated.pdf"
        db.scalar.return_value = identified_master(master)

        def fake_stamp(_source, target, **_kwargs):
            target.write_bytes(b"authenticated-pdf")

        def fake_atomic_write(target, content):
            target.write_bytes(content)

        with (
            patch("app.services.certificate_authentication.resolve_storage_path", return_value=master),
            patch("app.services.certificate_authentication.relative_storage_path", return_value="capture/1/10-Master_MYCA-07-2026-0001.xlsx"),
            patch("app.services.certificate_authentication.build_storage_path", side_effect=[final_target, authenticated_target]),
            patch("app.services.certificate_authentication._convert_master_to_pdf", return_value=generated),
            patch("app.services.certificate_authentication._stamp_pdf", side_effect=fake_stamp),
            patch("app.services.certificate_authentication.atomic_write", side_effect=fake_atomic_write),
            patch("app.services.certificate_authentication.write_audit_log") as audit,
        ):
            updated = authenticate_certificate_pdf(db, certificate, user_id=4)

    assert updated.status == "authenticated"
    assert updated.match_status == "pending"
    assert updated.authenticated_by_id == 4
    assert updated.authenticated_pdf_generated_at is not None
    assert updated.final_pdf_path == str(final_target)
    assert updated.authenticated_pdf_path == str(authenticated_target)
    versions = [call.args[0] for call in db.add.call_args_list if isinstance(call.args[0], CertificatePdfVersion)]
    assert len(versions) == 1
    assert versions[0].source_status == "quality_approved"
    audit_values = audit.call_args.kwargs["new_values"]
    assert audit.call_args.kwargs["user_id"] == 4
    assert audit.call_args.kwargs["previous_values"]["status"] == "quality_approved"
    assert audit_values["status"] == "authenticated"
    assert audit_values["capture_master_file_id"] == 10
    assert audit_values["match_status"] == "pending"


def test_conversion_failure_does_not_authenticate_or_mutate_pdf_references():
    db = MagicMock()
    certificate = approved_certificate(match_status="pending", final_pdf_path=None)
    with tempfile.TemporaryDirectory() as temporary:
        root = Path(temporary)
        master = root / "master.xlsx"
        master.write_bytes(b"xlsx")
        final_target = root / "10-Master_MYCA-07-2026-0001.pdf"
        authenticated_target = root / "10-Master_MYCA-07-2026-0001_authenticated.pdf"
        db.scalar.return_value = identified_master(master)

        with (
            patch("app.services.certificate_authentication.resolve_storage_path", return_value=master),
            patch("app.services.certificate_authentication.relative_storage_path", return_value="capture/1/10-Master_MYCA-07-2026-0001.xlsx"),
            patch("app.services.certificate_authentication.build_storage_path", side_effect=[final_target, authenticated_target]),
            patch(
                "app.services.certificate_authentication._convert_master_to_pdf",
                side_effect=HTTPException(status_code=422, detail="conversion failed"),
            ),
            patch("app.services.certificate_authentication.write_audit_log") as audit,
        ):
            with pytest.raises(HTTPException, match="conversion failed"):
                authenticate_certificate_pdf(db, certificate, user_id=4)

    assert certificate.status == "quality_approved"
    assert certificate.final_pdf_path is None
    assert certificate.authenticated_pdf_path is None
    assert certificate.authenticated_by_id is None
    audit.assert_not_called()


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice no está disponible")
def test_authenticate_endpoint_returns_200_with_real_converter():
    now = datetime(2026, 7, 21, tzinfo=timezone.utc)
    certificate = Certificate(
        id=1,
        folio="MYCA-07-2026-0001",
        expected_folio="MYCA-07-2026-0001",
        service_order_id=1,
        equipment_id=1,
        field_sheet_id=1,
        certificate_type="acreditado",
        status="quality_approved",
        external_source="excel",
        match_status="pending",
        is_active=True,
        client_visible=False,
        created_at=now,
        updated_at=now,
    )
    certificate.pdf_versions = []
    db = MagicMock()
    permission_dependency = inspect.signature(authenticate_certificate_endpoint).parameters[
        "current_user"
    ].default.dependency
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[enforce_api_access] = lambda: None
    app.dependency_overrides[permission_dependency] = lambda: SimpleNamespace(id=4)

    with tempfile.TemporaryDirectory() as temporary:
        root = Path(temporary)
        master = root / "Master_MYCA-07-2026-0001.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Certificado"
        sheet["A1"] = "METROLOGIA Y SERVICIOS MYC"
        sheet["A2"] = "CERTIFICADO DE CALIBRACIÓN"
        sheet.print_area = "A1:H40"
        workbook.save(master)
        final_target = root / "Master_MYCA-07-2026-0001.pdf"
        authenticated_target = root / "Master_MYCA-07-2026-0001_authenticated.pdf"
        capture_file = identified_master(master)

        try:
            with (
                patch("app.routers.certificates.get_certificate", return_value=certificate),
                patch("app.services.certificate_authentication._approved_capture_master", return_value=capture_file),
                patch("app.services.certificate_authentication.resolve_storage_path", return_value=master),
                patch("app.services.certificate_authentication.relative_storage_path", return_value="capture/1/Master_MYCA-07-2026-0001.xlsx"),
                    patch("app.services.certificate_authentication.build_storage_path", side_effect=[final_target, authenticated_target]),
                    patch("app.services.certificate_authentication.atomic_write", side_effect=lambda target, content: target.write_bytes(content)),
                    patch("app.services.certificate_authentication.write_audit_log") as audit,
                TestClient(app) as client,
            ):
                response = client.post("/api/certificates/1/authenticate")
        finally:
            app.dependency_overrides.clear()

        assert response.status_code == 200
        assert response.json()["status"] == "authenticated"
        assert response.json()["authenticated_by_id"] == 4
        assert response.json()["match_status"] == "pending"
        assert final_target.read_bytes().startswith(b"%PDF")
        assert authenticated_target.read_bytes().startswith(b"%PDF")
        assert audit.call_args.kwargs["user_id"] == 4
        assert audit.call_args.kwargs["new_values"]["capture_master_file_id"] == 10


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice no está disponible")
def test_real_xlsx_converter_produces_an_accessible_pdf():
    with tempfile.TemporaryDirectory() as temporary:
        root = Path(temporary)
        source = root / "Master_MYCA-TEST.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Certificado"
        sheet["A1"] = "METROLOGIA Y SERVICIOS MYC"
        sheet["A2"] = "CERTIFICADO DE CALIBRACIÓN"
        sheet.print_area = "A1:H40"
        auxiliary = workbook.create_sheet("Registro de clientes")
        for row in range(1, 200):
            auxiliary.cell(row=row, column=1, value=f"Cliente auxiliar {row}")
        workbook.save(source)
        output = root / "output"
        output.mkdir()
        pdf = _convert_master_to_pdf(source, output)
        assert pdf.is_file()
        assert pdf.read_bytes().startswith(b"%PDF")
        reader = PdfReader(str(pdf))
        assert len(reader.pages) == 1
        assert "Cliente auxiliar" not in " ".join(page.extract_text() or "" for page in reader.pages)
