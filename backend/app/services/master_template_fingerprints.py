"""Semantic service detection based on registered Master XLSX templates.

The document text is presentation, not the service identifier.  Detection
therefore compares the uploaded workbook with the immutable Master snapshot
assigned to the equipment and only then assigns the snapshot's canonical
service type.
"""
from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any
from unicodedata import normalize

from openpyxl import load_workbook


FINGERPRINT_THRESHOLD = 0.72


def canonical_service_type(value: str | None) -> str | None:
    normalized = (value or "").strip().casefold()
    if normalized in {"accredited", "accredited_iso_17025", "accredited_linked_lab", "acreditado", "vinculado"}:
        return "accredited"
    if normalized in {"traceable", "trazable"}:
        return "traceable"
    return None


def _normalized_text(value: Any) -> str:
    text = normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return " ".join(text.casefold().split())


def _jaccard(left: set[Any], right: set[Any], *, both_empty: float = 1.0) -> float:
    if not left and not right:
        return both_empty
    union = left | right
    return len(left & right) / len(union) if union else both_empty


def _ratio_close(left: int, right: int) -> float:
    maximum = max(left, right, 1)
    return max(0.0, 1.0 - abs(left - right) / maximum)


def _image_anchor(image: Any) -> str:
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is not None:
        return f"{marker.row}:{marker.col}"
    return str(anchor or "unknown")


@dataclass(frozen=True)
class SheetFingerprint:
    title: str
    rows: int
    columns: int
    merged_ranges: frozenset[str]
    styled_cells: frozenset[str]
    formulas: frozenset[str]
    positioned_labels: frozenset[tuple[str, str]]
    image_anchors: frozenset[str]
    print_area: str


@dataclass(frozen=True)
class WorkbookFingerprint:
    sheets: tuple[SheetFingerprint, ...]


def workbook_fingerprint(workbook: Any) -> WorkbookFingerprint:
    sheets = []
    for sheet in workbook.worksheets:
        labels: set[tuple[str, str]] = set()
        styled_cells: set[str] = set()
        formulas: set[str] = set()
        for row in sheet.iter_rows():
            for cell in row:
                if cell.has_style:
                    styled_cells.add(cell.coordinate)
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas.add(cell.coordinate)
                elif isinstance(value, str):
                    normalized = _normalized_text(value)
                    if len(normalized) >= 4:
                        labels.add((cell.coordinate, normalized))
        sheets.append(SheetFingerprint(
            title=_normalized_text(sheet.title),
            rows=sheet.max_row,
            columns=sheet.max_column,
            merged_ranges=frozenset(str(item) for item in sheet.merged_cells.ranges),
            styled_cells=frozenset(styled_cells),
            formulas=frozenset(formulas),
            positioned_labels=frozenset(labels),
            image_anchors=frozenset(_image_anchor(image) for image in sheet._images),
            print_area=str(sheet.print_area or ""),
        ))
    return WorkbookFingerprint(tuple(sheets))


def _sheet_map(fingerprint: WorkbookFingerprint) -> dict[str, SheetFingerprint]:
    return {sheet.title: sheet for sheet in fingerprint.sheets}


def compare_fingerprints(reference: WorkbookFingerprint, candidate: WorkbookFingerprint) -> dict[str, Any]:
    reference_sheets = _sheet_map(reference)
    candidate_sheets = _sheet_map(candidate)
    shared_names = set(reference_sheets) & set(candidate_sheets)
    sheet_names = _jaccard(set(reference_sheets), set(candidate_sheets), both_empty=0.0)
    if not shared_names:
        return {
            "matched": False,
            "score": 0.0,
            "threshold": FINGERPRINT_THRESHOLD,
            "indicators": {"sheet_names": 0.0},
            "evidence_groups": 0,
        }

    dimensions = []
    merged_regions = []
    styled_layout = []
    formula_layout = []
    positioned_labels = []
    image_layout = []
    print_layout = []
    for name in shared_names:
        left = reference_sheets[name]
        right = candidate_sheets[name]
        dimensions.append((_ratio_close(left.rows, right.rows) + _ratio_close(left.columns, right.columns)) / 2)
        merged_regions.append(_jaccard(set(left.merged_ranges), set(right.merged_ranges)))
        styled_layout.append(_jaccard(set(left.styled_cells), set(right.styled_cells)))
        formula_layout.append(_jaccard(set(left.formulas), set(right.formulas)))
        positioned_labels.append(_jaccard(set(left.positioned_labels), set(right.positioned_labels)))
        image_layout.append(_jaccard(set(left.image_anchors), set(right.image_anchors)))
        print_layout.append(1.0 if left.print_area == right.print_area else 0.0)

    average = lambda values: sum(values) / len(values) if values else 0.0
    indicators = {
        "sheet_names": sheet_names,
        "dimensions": average(dimensions),
        "merged_regions": average(merged_regions),
        "styled_layout": average(styled_layout),
        "formula_layout": average(formula_layout),
        "positioned_labels": average(positioned_labels),
        "image_layout": average(image_layout),
        "print_layout": average(print_layout),
    }
    weights = {
        "sheet_names": 0.12,
        "dimensions": 0.08,
        "merged_regions": 0.16,
        "styled_layout": 0.16,
        "formula_layout": 0.12,
        "positioned_labels": 0.18,
        "image_layout": 0.10,
        "print_layout": 0.08,
    }
    score = sum(indicators[key] * weight for key, weight in weights.items())
    structural_groups = ("merged_regions", "styled_layout", "formula_layout", "positioned_labels", "image_layout")
    evidence_groups = sum(indicators[key] >= 0.60 for key in structural_groups)
    matched = (
        score >= FINGERPRINT_THRESHOLD
        and indicators["sheet_names"] >= 0.80
        and indicators["dimensions"] >= 0.80
        and evidence_groups >= 3
    )
    return {
        "matched": matched,
        "score": round(score, 4),
        "threshold": FINGERPRINT_THRESHOLD,
        "indicators": {key: round(value, 4) for key, value in indicators.items()},
        "evidence_groups": evidence_groups,
    }


def detect_service_type(
    raw: bytes,
    *,
    extension: str,
    expected_template_path: Path | None,
    expected_service_type: str | None,
) -> dict[str, Any]:
    expected = canonical_service_type(expected_service_type)
    if expected is None:
        return {
            "status": "no_aplicable",
            "expected": expected_service_type,
            "detected": None,
            "method": "template_fingerprint",
            "reason": "Tipo de servicio ERP sin equivalencia canónica",
        }
    if expected_template_path is None or not expected_template_path.is_file():
        return {
            "status": "no_encontrado",
            "expected": expected,
            "detected": None,
            "method": "template_fingerprint",
            "reason": "Snapshot Master esperado no disponible",
        }
    try:
        candidate_workbook = load_workbook(
            BytesIO(raw), read_only=False, data_only=False, keep_links=False, keep_vba=extension == ".xlsm"
        )
        reference_workbook = load_workbook(
            expected_template_path, read_only=False, data_only=False, keep_links=False,
            keep_vba=expected_template_path.suffix.lower() == ".xlsm",
        )
        comparison = compare_fingerprints(
            workbook_fingerprint(reference_workbook), workbook_fingerprint(candidate_workbook)
        )
        candidate_workbook.close()
        reference_workbook.close()
    except Exception as exc:
        return {
            "status": "mismatch",
            "expected": expected,
            "detected": None,
            "method": "template_fingerprint",
            "reason": f"No fue posible analizar la estructura del Master: {type(exc).__name__}",
        }
    detected = expected if comparison["matched"] else None
    return {
        "status": "coincide" if detected == expected else "mismatch",
        "expected": expected,
        "detected": detected,
        "method": "template_fingerprint",
        "template_match": comparison,
    }
