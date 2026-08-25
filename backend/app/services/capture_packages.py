"""Generation and ingestion of calibration capture packages.

All selection is intentionally server-side: a browser never decides whether an
instrument is eligible or which template/version it receives.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timezone
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from unicodedata import normalize
import zipfile

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models.certificate import Certificate, CertificateCaptureFile
from app.models.equipment import Equipment
from app.models.field_sheet import FieldSheet
from app.models.controlled_document import (
    ControlledDocument,
    ControlledDocumentVersion,
    DocumentInterpretation,
    TechnicalProfile,
)
from app.models.service_order import ServiceOrder, ServiceWorkOrder
from app.models.service_order import ServiceOrderItem
from app.services.field_sheet_pdfs import generate_field_sheet_pdf
from app.services.audit_logs import write_audit_log
from app.services.certificates import CAPTURE_READY_STATUSES
from app.services.master_template_fingerprints import detect_service_type
from app.services.file_security import POLICIES, validate_content, validate_upload
from app.services.equipment import freeze_selected_certificate_master
from app.services.storage_service import resolve_storage_path, save_validated_content


EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
CAPTURE_PACKAGE_FIELD_SHEET_STATUSES = {"completed", "under_review", "approved"}


def normalized_filename(value: str) -> str:
    value = normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii")
    value = " ".join(value.upper().split())
    return "".join("-" if char in r'\\/:*?\"<>|' else char for char in value).strip()


def _ets_name(order: ServiceOrder) -> str:
    return normalized_filename(order.folio or f"ETS-{order.id}") or f"ETS-{order.id}"


def _work_order_folder(work_order: ServiceWorkOrder) -> str:
    return f"OT-{work_order.work_order_number:04d}"


def _certificate_folder(certificate: Certificate) -> str:
    folio = certificate.expected_folio or certificate.folio
    if not folio:
        raise ValueError("Falta folio de certificado")
    return normalized_filename(folio)


def _master_delivery_name(certificate: Certificate, extension: str) -> str:
    return f"Master_{_certificate_folder(certificate)}{extension.lower()}"


def _package_member_prefix(
    order: ServiceOrder,
    work_order: ServiceWorkOrder,
    certificate: Certificate,
    *,
    include_ets: bool,
) -> str:
    parts = [_work_order_folder(work_order), _certificate_folder(certificate)]
    if include_ets:
        parts.insert(0, _ets_name(order))
    return "/".join(parts)


def equipment_base_name(certificate: Certificate, field_sheet: FieldSheet) -> str:
    equipment = certificate.equipment
    if not certificate.expected_folio and not certificate.folio:
        raise ValueError("Falta folio de certificado")
    if not equipment.name:
        raise ValueError("Falta nombre de equipo")
    if not equipment.internal_id:
        raise ValueError("Falta identificación")
    certificate_type = getattr(certificate, "certificate_type", None)
    is_verification = certificate_type == "verification"
    if not is_verification and not field_sheet.next_calibration_date:
        raise ValueError("Falta fecha de próxima calibración")
    reference_date = (
        field_sheet.calibration_date
        if is_verification
        else field_sheet.next_calibration_date
    )
    if reference_date is None:
        raise ValueError("Falta fecha de ejecución")
    return normalized_filename(
        f"{certificate.expected_folio or certificate.folio} {equipment.name} "
        f"{equipment.internal_id} {reference_date:%Y.%m.%d}"
    )


@dataclass
class EligibleItem:
    equipment: Equipment
    field_sheet: FieldSheet | None
    certificate: Certificate | None
    reason: str | None = None

    @property
    def ready(self) -> bool:
        return self.reason is None


def _load_equipment(db: Session, *, service_order_id: int, work_order_id: int | None = None) -> list[Equipment]:
    query = select(Equipment).where(Equipment.service_order_id == service_order_id, Equipment.is_active.is_(True)).options(
        selectinload(Equipment.certificates), selectinload(Equipment.field_sheets)
    )
    if work_order_id is not None:
        query = query.where(Equipment.work_order_id == work_order_id)
    return list(db.scalars(query).all())


def eligibility_for_equipment(db: Session, equipment: Equipment) -> EligibleItem:
    service_order_item_id = getattr(equipment, "service_order_item_id", None)
    service_item = db.get(ServiceOrderItem, service_order_item_id) if service_order_item_id else None
    operational_category = (
        service_item.operational_category
        if service_item is not None
        else "calibration" if equipment.calibration_scope else None
    )
    if operational_category not in {"calibration", "verification"}:
        return EligibleItem(equipment, None, None, "El servicio no pertenece a un proceso metrológico")
    field_sheet = next((item for item in equipment.field_sheets if item.is_active), None)
    certificate = next((item for item in equipment.certificates if item.is_active), None)
    if field_sheet is None:
        return EligibleItem(equipment, None, certificate, "No tiene Hoja de Campo")
    if field_sheet.status not in CAPTURE_PACKAGE_FIELD_SHEET_STATUSES:
        return EligibleItem(equipment, field_sheet, certificate, "La Hoja de Campo no está completada")
    if certificate is None or not (certificate.expected_folio or certificate.folio):
        return EligibleItem(equipment, field_sheet, certificate, "Falta folio de certificado asignado")
    if not equipment.name:
        return EligibleItem(equipment, field_sheet, certificate, "Falta nombre de equipo")
    if not equipment.internal_id:
        return EligibleItem(equipment, field_sheet, certificate, "Falta identificación")
    if operational_category == "calibration" and not field_sheet.next_calibration_date:
        return EligibleItem(equipment, field_sheet, certificate, "Falta fecha de próxima calibración")
    if not equipment.certificate_master_document_id:
        return EligibleItem(equipment, field_sheet, certificate, "Falta plantilla esperada de certificado")
    if not equipment.certificate_master_version_id or not equipment.certificate_template_path_snapshot:
        return EligibleItem(equipment, field_sheet, certificate, "La versión o snapshot de plantilla no está disponible")
    master = db.get(ControlledDocument, equipment.certificate_master_document_id)
    version = db.get(ControlledDocumentVersion, equipment.certificate_master_version_id)
    if master is None or master.status != "active":
        return EligibleItem(equipment, field_sheet, certificate, "La plantilla está inactiva")
    if version is None or version.status != "active":
        return EligibleItem(equipment, field_sheet, certificate, "Falta versión activa de plantilla")
    if version.expires_on and version.expires_on < date.today():
        return EligibleItem(equipment, field_sheet, certificate, "La plantilla está caducada")
    path = resolve_storage_path(equipment.certificate_template_path_snapshot)
    if path is None or not path.is_file():
        return EligibleItem(equipment, field_sheet, certificate, "El archivo snapshot de plantilla no está disponible")
    if path.suffix.lower() != ".xlsx":
        return EligibleItem(equipment, field_sheet, certificate, "El archivo snapshot no es XLSX")
    if equipment.certificate_template_checksum_snapshot and sha256(path.read_bytes()).hexdigest() != equipment.certificate_template_checksum_snapshot:
        return EligibleItem(equipment, field_sheet, certificate, "El hash del archivo snapshot no coincide")
    return EligibleItem(equipment, field_sheet, certificate)


def package_summary(db: Session, service_order_id: int) -> dict:
    order = db.get(ServiceOrder, service_order_id)
    if order is None or not order.is_active:
        raise HTTPException(status_code=404, detail="ETS no encontrado")
    groups = []
    for work_order in db.scalars(select(ServiceWorkOrder).where(ServiceWorkOrder.service_order_id == order.id, ServiceWorkOrder.is_active.is_(True)).order_by(ServiceWorkOrder.sequence)).all():
        items = _eligible_with_pdf(db, _load_equipment(db, service_order_id=order.id, work_order_id=work_order.id))
        groups.append({"work_order_id": work_order.id, "work_order_number": work_order.work_order_number,
                       "ready": sum(item.ready for item in items), "pending": sum(not item.ready for item in items),
                       "blocked": [{"equipment_id": item.equipment.id, "equipment_name": item.equipment.name, "reason": item.reason} for item in items if not item.ready]})
    return {"service_order_id": order.id, "folio": order.folio, "work_orders": groups,
            "ready_total": sum(group["ready"] for group in groups)}


def _render_pair(db: Session, item: EligibleItem) -> tuple[str, bytes, str, bytes]:
    assert item.field_sheet and item.certificate
    equipment_base_name(item.certificate, item.field_sheet)
    pdf, _ = generate_field_sheet_pdf(db, item.field_sheet.id)
    template = resolve_storage_path(item.equipment.certificate_template_path_snapshot)
    assert template is not None
    extension = template.suffix.lower()
    if extension not in EXCEL_EXTENSIONS:
        raise HTTPException(status_code=422, detail="La plantilla snapshot no tiene una extensión Excel aceptada")
    excel = template.read_bytes()
    folio = _certificate_folder(item.certificate)
    pdf_name = f"Hoja_Campo_{folio}.pdf"
    excel_name = _master_delivery_name(item.certificate, extension)
    return pdf_name, pdf, excel_name, excel


def _eligible_with_pdf(db: Session, equipment: list[Equipment]) -> list[EligibleItem]:
    rows = []
    for equipment_item in equipment:
        item = eligibility_for_equipment(db, equipment_item)
        if item.ready:
            try:
                generate_field_sheet_pdf(db, item.field_sheet.id)  # genuine final check
            except Exception:
                item.reason = "No se pudo generar el PDF de la Hoja de Campo"
        rows.append(item)
    return rows


def work_order_package(db: Session, service_order_id: int, work_order_id: int) -> tuple[bytes, str, str]:
    order = db.get(ServiceOrder, service_order_id)
    work_order = db.get(ServiceWorkOrder, work_order_id)
    if order is None or work_order is None or work_order.service_order_id != order.id:
        raise HTTPException(status_code=404, detail="ETS u Orden de Trabajo no encontrada")
    ready = [item for item in _eligible_with_pdf(db, _load_equipment(db, service_order_id=order.id, work_order_id=work_order.id)) if item.ready]
    if not ready:
        raise HTTPException(status_code=409, detail="No hay equipos elegibles para el paquete de Captura")
    folder = _work_order_folder(work_order)
    if len(ready) == 1:
        pdf_name, pdf, excel_name, excel = _render_pair(db, ready[0])
        boundary = "MYC-CAPTURE-PACKAGE"
        body = b"".join([
            f"--{boundary}\r\nContent-Type: application/pdf\r\nContent-Disposition: attachment; filename=\"{pdf_name}\"\r\n\r\n".encode(), pdf, b"\r\n",
            f"--{boundary}\r\nContent-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\nContent-Disposition: attachment; filename=\"{excel_name}\"\r\n\r\n".encode(), excel, b"\r\n",
            f"--{boundary}--\r\n".encode(),
        ])
        return body, f"{folder}.multipart", f"multipart/mixed; boundary={boundary}"
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for item in ready:
            pdf_name, pdf, excel_name, excel = _render_pair(db, item)
            prefix = _package_member_prefix(order, work_order, item.certificate, include_ets=False)
            archive.writestr(f"{prefix}/{pdf_name}", pdf)
            archive.writestr(f"{prefix}/{excel_name}", excel)
    return buffer.getvalue(), f"{folder}.zip", "application/zip"


def service_order_package(db: Session, service_order_id: int) -> tuple[bytes, str]:
    order = db.get(ServiceOrder, service_order_id)
    if order is None:
        raise HTTPException(status_code=404, detail="ETS no encontrado")
    buffer = BytesIO()
    any_file = False
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for work_order in db.scalars(select(ServiceWorkOrder).where(ServiceWorkOrder.service_order_id == order.id, ServiceWorkOrder.is_active.is_(True))).all():
            for item in _eligible_with_pdf(db, _load_equipment(db, service_order_id=order.id, work_order_id=work_order.id)):
                if not item.ready:
                    continue
                pdf_name, pdf, excel_name, excel = _render_pair(db, item)
                prefix = _package_member_prefix(order, work_order, item.certificate, include_ets=True)
                archive.writestr(f"{prefix}/{pdf_name}", pdf)
                archive.writestr(f"{prefix}/{excel_name}", excel)
                any_file = True
    if not any_file:
        raise HTTPException(status_code=409, detail="No hay equipos elegibles; consulta el resumen de bloqueos")
    return buffer.getvalue(), f"{_ets_name(order)}.zip"


def _expected_values(certificate: Certificate) -> dict[str, str]:
    sheet = certificate.field_sheet
    equipment = certificate.equipment
    client = certificate.service_order.client
    return {"folio": certificate.expected_folio or certificate.folio, "cliente": client.commercial_name or client.legal_name or "",
            "equipo": equipment.name or "", "identificacion": equipment.internal_id or "", "marca": equipment.brand or "",
            "modelo": equipment.model or "", "serie": equipment.serial_number or "", "fecha_calibracion": str(sheet.calibration_date or "") if sheet else "",
            "proxima_calibracion": str(sheet.next_calibration_date or "") if sheet else "", "servicio": certificate.certificate_type}


def _validate_excel(
    raw: bytes,
    extension: str,
    certificate: Certificate,
    *,
    expected_template_path: Path | None = None,
) -> dict:
    if extension not in {".xlsx", ".xlsm"}:
        return {key: {"status": "no_aplicable"} for key in _expected_values(certificate)}
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True, keep_vba=extension == ".xlsm")
    observed = " ".join(str(cell.value or "") for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row)
    results = {}
    for key, expected in _expected_values(certificate).items():
        if key == "servicio":
            results[key] = detect_service_type(
                raw,
                extension=extension,
                expected_template_path=(
                    expected_template_path
                    if expected_template_path is not None
                    else resolve_storage_path(certificate.equipment.certificate_template_path_snapshot)
                ),
                expected_service_type=expected,
            )
            continue
        if not expected:
            results[key] = {"status": "no_aplicable"}
        elif expected.casefold() in observed.casefold():
            results[key] = {"status": "coincide", "expected": expected}
        else:
            results[key] = {"status": "no_encontrado", "expected": expected}
    return results


def _registered_verification_master_match(
    db: Session,
    raw: bytes,
    suffix: str,
) -> tuple[int | None, Path | None, dict]:
    """Resolve a real Verification Master only from active structured records."""
    if suffix not in {".xlsx", ".xlsm"}:
        return None, None, {"status": "no_aplicable", "method": "registered_master_fingerprint"}
    profile_rows = db.execute(
        select(ControlledDocument, ControlledDocumentVersion)
        .join(
            TechnicalProfile,
            TechnicalProfile.certificate_template_document_id == ControlledDocument.id,
        )
        .join(
            ControlledDocumentVersion,
            ControlledDocumentVersion.document_id == ControlledDocument.id,
        )
        .where(
            TechnicalProfile.status == "active",
            TechnicalProfile.service_type == "verification",
            ControlledDocument.document_type == "certificate_master",
            ControlledDocument.status == "active",
            ControlledDocumentVersion.status == "active",
        )
    ).all()
    interpretation_rows = db.execute(
        select(ControlledDocument, ControlledDocumentVersion)
        .join(
            DocumentInterpretation,
            DocumentInterpretation.document_id == ControlledDocument.id,
        )
        .join(
            ControlledDocumentVersion,
            ControlledDocumentVersion.id == DocumentInterpretation.document_version_id,
        )
        .where(
            DocumentInterpretation.status == "approved",
            DocumentInterpretation.interpretation_type == "certificate_template_source",
            DocumentInterpretation.service_type == "verification",
            ControlledDocument.document_type == "certificate_master",
            ControlledDocument.status == "active",
            ControlledDocumentVersion.status == "active",
        )
    ).all()
    registered_versions = {
        version.id: (document, version)
        for document, version in [*profile_rows, *interpretation_rows]
    }
    matches: dict[int, tuple[Path, dict]] = {}
    for document, version in registered_versions.values():
        if not version.file_path or (
            version.expires_on is not None and version.expires_on < date.today()
        ):
            continue
        template_path = resolve_storage_path(version.file_path)
        result = detect_service_type(
            raw,
            extension=suffix,
            expected_template_path=template_path,
            expected_service_type="verification",
        )
        if result.get("status") == "coincide" and template_path is not None:
            matches[document.id] = (template_path, result)
    if len(matches) == 1:
        document_id, (template_path, result) = next(iter(matches.items()))
        return document_id, template_path, {
            "status": "coincide",
            "method": "registered_master_fingerprint",
            "document_id": document_id,
            "template_match": result.get("template_match"),
        }
    if len(matches) > 1:
        return None, None, {
            "status": "ambiguous",
            "method": "registered_master_fingerprint",
            "candidate_document_ids": sorted(matches),
        }
    return None, None, {
        "status": "no_encontrado",
        "method": "registered_master_fingerprint",
    }


def _is_macos_auxiliary(path_value: str) -> bool:
    path = Path(path_value)
    return "__MACOSX" in path.parts or path.name == ".DS_Store" or path.name.startswith("._")


def _validation_issue_keys(validation: dict | None) -> tuple[list[str], list[str]]:
    warnings = []
    mismatches = []
    for key, result in (validation or {}).items():
        status = result.get("status") if isinstance(result, dict) else None
        if status == "no_encontrado":
            warnings.append(key)
        elif status in {"mismatch", "no_coincide"}:
            mismatches.append(key)
    return warnings, mismatches


def _mark_capture_started(
    db: Session, certificate: Certificate, *, user_id: int, filename: str
) -> None:
    if user_id is None:
        raise ValueError("La mutación de Captura ETS requiere un actor")
    if certificate.status not in CAPTURE_READY_STATUSES:
        return
    now = datetime.now(timezone.utc)
    previous_status = certificate.status
    certificate.status = "capture_in_progress"
    certificate.capture_started_at = certificate.capture_started_at or now
    certificate.capture_started_by_id = certificate.capture_started_by_id or user_id
    write_audit_log(
        db,
        action="certificate.capture_started",
        entity="certificates",
        entity_id=certificate.id,
        user_id=user_id,
        previous_values={"status": previous_status},
        new_values={
            "status": certificate.status,
            "capture_started_at": certificate.capture_started_at.isoformat(),
            "source": "capture_package_upload",
            "filename": filename,
        },
    )


def _match_uploaded_capture(
    raw: bytes,
    suffix: str,
    certificates: list[Certificate],
    *,
    verification_template_path: Path | None = None,
) -> tuple[Certificate | None, dict | None]:
    matches: list[tuple[int, Certificate, dict]] = []
    strong_identity_fields = {"folio", "identificacion", "serie"}
    for certificate in certificates:
        validation = _validate_excel(
            raw,
            suffix,
            certificate,
            expected_template_path=(
                verification_template_path
                if certificate.certificate_type == "verification"
                else None
            ),
        )
        service_result = validation.get("servicio") or {}
        if service_result.get("status") != "coincide":
            continue
        coincident_fields = {
            key
            for key, result in validation.items()
            if isinstance(result, dict) and result.get("status") == "coincide"
        }
        identity_score = len(coincident_fields - {"servicio"})
        if not coincident_fields.intersection(strong_identity_fields):
            continue
        matches.append((identity_score, certificate, validation))
    if not matches:
        return None, None
    matches.sort(key=lambda item: item[0], reverse=True)
    if len(matches) > 1 and matches[0][0] == matches[1][0]:
        return None, None
    return matches[0][1], matches[0][2]


def upload_capture_files(
    db: Session,
    service_order_id: int,
    files: list[UploadFile],
    *,
    user_id: int,
) -> dict:
    if user_id is None:
        raise ValueError("La carga de Captura ETS requiere un actor")
    order = db.get(ServiceOrder, service_order_id)
    if order is None:
        raise HTTPException(status_code=404, detail="ETS no encontrado")
    expected: dict[str, list[Certificate]] = {}
    active_certificates: list[Certificate] = []
    for item in _load_equipment(db, service_order_id=order.id):
        extension = Path(item.certificate_template_path_snapshot or "").suffix.lower() or ".xlsx"
        for certificate in item.certificates:
            if not certificate.is_active:
                continue
            active_certificates.append(certificate)
            if item.certificate_template_filename_snapshot:
                expected.setdefault(item.certificate_template_filename_snapshot, []).append(certificate)
            if certificate.expected_folio or certificate.folio:
                expected.setdefault(_master_delivery_name(certificate, extension), []).append(certificate)
    candidates: list[tuple[str, bytes]] = []
    ignored_auxiliary: list[str] = []
    ignored_unsupported: list[str] = []
    for upload in files:
        validated = validate_upload(upload, "capture_package")
        raw = validated.content
        if validated.extension == ".zip":
            with zipfile.ZipFile(BytesIO(raw)) as archive:
                for info in archive.infolist():
                    if info.is_dir():
                        continue
                    if _is_macos_auxiliary(info.filename):
                        ignored_auxiliary.append(info.filename)
                        continue
                    filename = Path(info.filename).name
                    if Path(filename).suffix.lower() not in EXCEL_EXTENSIONS or filename.startswith("~$"):
                        ignored_unsupported.append(info.filename)
                        continue
                    member = archive.read(info)
                    suffix = Path(filename).suffix.lower()
                    validate_content(member, suffix, POLICIES["capture_package"])
                    candidates.append((filename, member))
        else:
            filename = Path(upload.filename or "").name
            if _is_macos_auxiliary(upload.filename or ""):
                ignored_auxiliary.append(upload.filename or filename)
            elif Path(filename).suffix.lower() not in EXCEL_EXTENSIONS or filename.startswith("~$"):
                ignored_unsupported.append(upload.filename or filename)
            else:
                candidates.append((filename, raw))
    results = []
    has_verification_certificates = any(
        certificate.certificate_type == "verification"
        for certificate in active_certificates
    )
    for filename, raw in candidates:
        suffix = Path(filename).suffix.lower()
        if has_verification_certificates:
            verification_master_id, verification_template_path, master_validation = (
                _registered_verification_master_match(db, raw, suffix)
            )
        else:
            verification_master_id, verification_template_path, master_validation = (
                None,
                None,
                {"status": "no_aplicable", "method": "registered_master_fingerprint"},
            )
        filename_matches = expected.get(filename, [])
        certificate = filename_matches[0] if len(filename_matches) == 1 else None
        validation = (
            _validate_excel(
                raw,
                suffix,
                certificate,
                expected_template_path=(
                    verification_template_path
                    if certificate.certificate_type == "verification"
                    else None
                ),
            )
            if certificate
            else None
        )
        if certificate is None:
            certificate, validation = _match_uploaded_capture(
                raw,
                suffix,
                active_certificates,
                verification_template_path=verification_template_path,
            )
        if certificate and certificate.certificate_type == "verification":
            verification_context = dict(
                certificate.equipment.certificate_operational_context_snapshot or {}
            )
            requires_registered_final = bool(
                verification_context.get("initial_certificate_master_document_id")
                and not verification_context.get("final_certificate_master_document_id")
            )
            if verification_master_id is not None:
                freeze_selected_certificate_master(
                    db,
                    certificate.equipment,
                    verification_master_id,
                    user_id=user_id,
                    selection_source="capture_upload_fingerprint",
                )
                validation = _validate_excel(raw, suffix, certificate)
            if validation is not None:
                validation["master_registration"] = master_validation
            if (
                requires_registered_final
                and verification_master_id is None
            ) or (validation or {}).get("servicio", {}).get("status") != "coincide":
                certificate = None
                validation = None
        unidentified_validation = {
            "file": {
                "status": "no_identificado",
                "message": "No fue posible asociar de forma única el archivo real con un certificado/equipo por identidad y fingerprint",
            }
        }
        if has_verification_certificates:
            unidentified_validation["master_registration"] = master_validation
        record = CertificateCaptureFile(service_order_id=order.id, certificate_id=certificate.id if certificate else None,
            original_filename=filename, identification_status="identified" if certificate else "unidentified", uploaded_by_id=user_id)
        record.validation_results = validation if certificate else unidentified_validation
        db.add(record)
        db.flush()
        stored = save_validated_content(
            directory=f"capture/{order.id}", filename=f"{record.id}-{filename}",
            content=raw, original_filename=filename,
        )
        record.stored_path = stored.relative_path
        if certificate:
            _mark_capture_started(db, certificate, user_id=user_id, filename=filename)
        warning_keys, mismatch_keys = _validation_issue_keys(record.validation_results)
        results.append({
            "id": record.id,
            "certificate_id": record.certificate_id,
            "filename": filename,
            "status": record.identification_status,
            "certificate_status": certificate.status if certificate else None,
            "validation": record.validation_results,
            "checksum_sha256": stored.checksum_sha256,
            "warnings": warning_keys,
            "mismatches": mismatch_keys,
        })
    db.commit()
    return {
        "processed": results,
        "count": len(results),
        "summary": {
            "identified": sum(item["status"] == "identified" for item in results),
            "unidentified": sum(item["status"] == "unidentified" for item in results),
            "ignored_auxiliary": len(ignored_auxiliary),
            "ignored_unsupported": len(ignored_unsupported),
            "warnings": sum(len(item["warnings"]) for item in results),
            "mismatches": sum(len(item["mismatches"]) for item in results),
        },
        "ignored": {
            "auxiliary": ignored_auxiliary,
            "unsupported": ignored_unsupported,
        },
    }


def list_capture_files(db: Session, service_order_id: int) -> list[CertificateCaptureFile]:
    return list(db.scalars(select(CertificateCaptureFile).where(
        CertificateCaptureFile.service_order_id == service_order_id
    ).order_by(CertificateCaptureFile.created_at.desc(), CertificateCaptureFile.id.desc())).all())
