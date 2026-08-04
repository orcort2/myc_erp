from __future__ import annotations

import csv
import re
from datetime import date, datetime, timezone
from io import BytesIO, StringIO
from pathlib import Path
from uuid import uuid4

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session, selectinload

from app.models.client import Client, ClientCertificateProfile, ClientContact
from app.models.certificate import Certificate
from app.models.equipment import Equipment
from app.models.field_sheet import FieldSheet
from app.models.invoice import CreditNote, Invoice, InvoicePayment
from app.models.quotation import Quotation
from app.models.sat_catalog import SatCatalogRecord
from app.models.service_order import ServiceOrder
from app.schemas.client import (
    ClientCertificateProfileCreate,
    ClientCertificateProfileUpdate,
    ClientCreate,
    ClientDeleteEligibilityRead,
    ClientDeleteResultRead,
    ClientImportConfirm,
    ClientImportPreviewRead,
    ClientImportResultRead,
    ClientImportRowRead,
    ClientRestoreResultRead,
    ClientTaxConstancyPreviewRead,
    ClientUpdate,
)
from app.services.audit_logs import write_audit_log
from app.services.sat_catalogs.service import latest_version, get_catalog
from app.services.sat_catalogs.normalizers import normalize_search
from app.services.file_security import validate_upload
from app.services.storage_service import delete_if_unreferenced, save_validated_content


CLIENT_IMPORT_COLUMNS = [
    "tipo_cliente",
    "nombre_comercial",
    "razon_social",
    "curp",
    "nombres",
    "primer_apellido",
    "segundo_apellido",
    "rfc",
    "contacto",
    "telefono",
    "correo",
    "pais",
    "tipo_vialidad",
    "calle",
    "numero_exterior",
    "numero_interior",
    "colonia",
    "localidad",
    "municipio",
    "municipio_ciudad",
    "estado",
    "codigo_postal",
    "regimen_fiscal",
    "uso_cfdi",
    "estado_cliente",
]


def _normalize_key(value: str | None) -> str:
    text = (value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


RFC_PATTERN = re.compile(r"^(?:[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3}|XAXX010101000|XEXX010101000)$")
GENERIC_RFCS = {"XAXX010101000", "XEXX010101000"}


def _validate_sat_code(db: Session, catalog_code: str, value: str | None, label: str) -> None:
    """Reject a supplied fiscal code unless it exists in the active local SAT version."""
    if not value:
        return
    catalog = get_catalog(db, catalog_code)
    version = latest_version(db, catalog)
    if version is None:
        raise HTTPException(status_code=409, detail=f"El catálogo SAT de {label} no está disponible localmente")
    record = db.scalar(
        select(SatCatalogRecord.id).where(
            SatCatalogRecord.catalog_version_id == version.id,
            SatCatalogRecord.normalized_code == normalize_search(value),
            SatCatalogRecord.is_active.is_(True),
            (SatCatalogRecord.valid_from.is_(None) | (SatCatalogRecord.valid_from <= date.today())),
            (SatCatalogRecord.valid_until.is_(None) | (SatCatalogRecord.valid_until >= date.today())),
        )
    )
    if record is None:
        raise HTTPException(status_code=422, detail=f"Código SAT de {label} no válido: {value}")


def _validate_fiscal_profile(db: Session, values: dict) -> None:
    rfc = (values.get("rfc") or "").strip().upper()
    if rfc and not RFC_PATTERN.fullmatch(rfc):
        raise HTTPException(status_code=422, detail="RFC con estructura inválida")
    if "rfc" in values:
        values["rfc"] = rfc or None
    if values.get("fiscal_postal_code"):
        values["fiscal_postal_code"] = str(values["fiscal_postal_code"]).strip().zfill(5)
    _validate_sat_code(db, "fiscal_regimes", values.get("tax_regime"), "régimen fiscal")
    _validate_sat_code(db, "cfdi_uses", values.get("cfdi_use"), "uso CFDI")
    _validate_sat_code(db, "postal_codes", values.get("fiscal_postal_code"), "código postal fiscal")
    _validate_sat_code(db, "countries", values.get("fiscal_country_code"), "país fiscal")


def _requires_fiscal_review(values: dict) -> bool:
    required = ("rfc", "legal_name", "tax_regime", "cfdi_use", "fiscal_postal_code")
    return any(not (values.get(key) or "").strip() for key in required)


def _without_function_words(value: str | None) -> str:
    return " ".join(word for word in normalize_search(value).split() if word not in {"de"})


def _resolve_sat_import_value(
    db: Session,
    catalog_code: str,
    value: str | None,
    cache: dict | None = None,
) -> str | None:
    """Resolve a local SAT code or a unique normalized description without fuzzy matching."""
    raw = (value or "").strip()
    if not raw:
        return None
    cache = cache if cache is not None else {}
    normalized = normalize_search(raw)
    result_key = ("result", catalog_code, normalized)
    if result_key in cache:
        return cache[result_key]

    version_key = ("version", catalog_code)
    if version_key not in cache:
        catalog = get_catalog(db, catalog_code)
        cache[version_key] = latest_version(db, catalog)
    version = cache[version_key]
    if version is None:
        cache[result_key] = None
        return None

    exact = list(db.scalars(select(SatCatalogRecord).where(
        SatCatalogRecord.catalog_version_id == version.id,
        SatCatalogRecord.is_active.is_(True),
        or_(
            SatCatalogRecord.normalized_code == normalized,
            SatCatalogRecord.normalized_name == normalized,
        ),
    )).all())
    if len(exact) == 1:
        cache[result_key] = exact[0].code
        return cache[result_key]

    # Los códigos postales pueden tener decenas de miles de registros. Para ese
    # catálogo sólo se permite coincidencia exacta, nunca una carga completa.
    if catalog_code == "postal_codes":
        cache[result_key] = None
        return None

    functional = _without_function_words(raw)
    candidates_key = ("candidates", catalog_code)
    if candidates_key not in cache:
        cache[candidates_key] = list(db.scalars(select(SatCatalogRecord).where(
            SatCatalogRecord.catalog_version_id == version.id,
            SatCatalogRecord.is_active.is_(True),
        )).all())
    matches = [record for record in cache[candidates_key] if _without_function_words(record.name) == functional]
    cache[result_key] = matches[0].code if len(matches) == 1 else None
    return cache[result_key]


def _json_safe(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    return value


def _client_query():
    return (
        select(Client)
        .options(selectinload(Client.contacts), selectinload(Client.certificate_profiles))
        .order_by(Client.legal_name)
    )


def _certificate_profile_query(client_id: int):
    return (
        select(ClientCertificateProfile)
        .where(ClientCertificateProfile.client_id == client_id)
        .order_by(ClientCertificateProfile.is_default.desc(), ClientCertificateProfile.label)
    )


def list_client_certificate_profiles(
    db: Session, client_id: int, *, include_inactive: bool = False
) -> list[ClientCertificateProfile]:
    get_client(db, client_id, include_inactive=True)
    query = _certificate_profile_query(client_id)
    if not include_inactive:
        query = query.where(ClientCertificateProfile.is_active.is_(True))
    return list(db.scalars(query).all())


def _clear_default_certificate_profiles(db: Session, client_id: int, except_id: int | None = None) -> None:
    profiles = db.scalars(_certificate_profile_query(client_id)).all()
    for profile in profiles:
        if profile.is_active and profile.id != except_id:
            profile.is_default = False


def create_client_certificate_profile(
    db: Session,
    client_id: int,
    payload: ClientCertificateProfileCreate,
    *,
    user_id: int | None = None,
) -> ClientCertificateProfile:
    get_client(db, client_id)
    data = payload.model_dump()
    if data["is_default"]:
        _clear_default_certificate_profiles(db, client_id)
    profile = ClientCertificateProfile(client_id=client_id, **data)
    db.add(profile)
    db.flush()
    write_audit_log(
        db,
        action="client.certificate_profile.created",
        entity="client_certificate_profiles",
        entity_id=profile.id,
        user_id=user_id,
        new_values=_json_safe(data | {"client_id": client_id}),
    )
    db.commit()
    db.refresh(profile)
    return profile


def update_client_certificate_profile(
    db: Session,
    client_id: int,
    profile_id: int,
    payload: ClientCertificateProfileUpdate,
    *,
    user_id: int | None = None,
) -> ClientCertificateProfile:
    get_client(db, client_id, include_inactive=True)
    profile = db.scalar(
        _certificate_profile_query(client_id).where(ClientCertificateProfile.id == profile_id)
    )
    if profile is None or not profile.is_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dato para certificado no encontrado")
    updates = payload.model_dump(exclude_unset=True)
    previous = {key: getattr(profile, key) for key in updates}
    if updates.get("is_default"):
        _clear_default_certificate_profiles(db, client_id, except_id=profile.id)
    for key, value in updates.items():
        setattr(profile, key, value)
    write_audit_log(
        db,
        action="client.certificate_profile.updated",
        entity="client_certificate_profiles",
        entity_id=profile.id,
        user_id=user_id,
        previous_values=_json_safe(previous),
        new_values=_json_safe(updates),
    )
    db.commit()
    db.refresh(profile)
    return profile


def deactivate_client_certificate_profile(
    db: Session, client_id: int, profile_id: int, *, user_id: int | None = None
) -> None:
    profile = db.scalar(
        _certificate_profile_query(client_id).where(ClientCertificateProfile.id == profile_id)
    )
    if profile is None or not profile.is_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dato para certificado no encontrado")
    profile.is_active = False
    profile.is_default = False
    profile.deleted_at = datetime.now(timezone.utc)
    profile.deleted_by = user_id
    write_audit_log(
        db,
        action="client.certificate_profile.deactivated",
        entity="client_certificate_profiles",
        entity_id=profile.id,
        user_id=user_id,
        previous_values={"is_active": True},
        new_values={"is_active": False},
    )
    db.commit()


def _clean_text(value: str | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_client_type(value: str | None) -> str:
    normalized = _normalize_key(value)
    if normalized in {"personafisica", "fisica", "pf"}:
        return "persona_fisica"
    return "persona_moral"


def _compose_person_name(data: dict) -> str | None:
    parts = [
        _clean_text(data.get("first_name")),
        _clean_text(data.get("first_last_name")),
        _clean_text(data.get("second_last_name")),
    ]
    full_name = " ".join(part for part in parts if part)
    return full_name or None


def _normalize_client_data(data: dict, *, partial: bool = False) -> dict:
    normalized = dict(data)
    present_keys = set(normalized.keys())

    def has(key: str) -> bool:
        return key in present_keys or not partial

    if has("client_type"):
        normalized["client_type"] = _normalize_client_type(normalized.get("client_type"))
    if has("legal_name"):
        normalized["legal_name"] = _clean_text(normalized.get("legal_name"))
    if has("commercial_name"):
        normalized["commercial_name"] = _clean_text(normalized.get("commercial_name"))
    if has("rfc"):
        normalized["rfc"] = _clean_text(normalized.get("rfc"))
    if has("curp"):
        normalized["curp"] = _clean_text(normalized.get("curp"))
    if has("first_name"):
        normalized["first_name"] = _clean_text(normalized.get("first_name"))
    if has("first_last_name"):
        normalized["first_last_name"] = _clean_text(normalized.get("first_last_name"))
    if has("second_last_name"):
        normalized["second_last_name"] = _clean_text(normalized.get("second_last_name"))
    if has("street_type"):
        normalized["street_type"] = _clean_text(normalized.get("street_type"))
    if has("street"):
        normalized["street"] = _clean_text(normalized.get("street"))
    if has("exterior_number"):
        normalized["exterior_number"] = _clean_text(normalized.get("exterior_number"))
    if has("interior_number"):
        normalized["interior_number"] = _clean_text(normalized.get("interior_number"))
    if has("neighborhood"):
        normalized["neighborhood"] = _clean_text(normalized.get("neighborhood"))
    if has("locality"):
        normalized["locality"] = _clean_text(normalized.get("locality"))
    if has("municipality") or has("city"):
        normalized["municipality"] = _clean_text(normalized.get("municipality")) or _clean_text(normalized.get("city"))
        normalized["city"] = _clean_text(normalized.get("city")) or normalized.get("municipality")
    if has("state"):
        normalized["state"] = _clean_text(normalized.get("state"))
    if has("postal_code"):
        normalized["postal_code"] = _clean_text(normalized.get("postal_code"))
    if has("fiscal_postal_code") or has("postal_code"):
        normalized["fiscal_postal_code"] = _clean_text(normalized.get("fiscal_postal_code")) or normalized.get("postal_code")
    if has("country"):
        normalized["country"] = _clean_text(normalized.get("country")) or "Mexico"
    if has("fiscal_country_code"):
        normalized["fiscal_country_code"] = _clean_text(normalized.get("fiscal_country_code"))
    if has("tax_regime"):
        normalized["tax_regime"] = _clean_text(normalized.get("tax_regime"))
    if has("cfdi_use"):
        normalized["cfdi_use"] = _clean_text(normalized.get("cfdi_use"))

    client_type = normalized.get("client_type") if "client_type" in normalized else "persona_moral"
    if client_type == "persona_fisica":
        resolved_legal_name = _compose_person_name(normalized) or normalized.get("legal_name") or normalized.get("commercial_name")
    else:
        resolved_legal_name = normalized.get("legal_name") or normalized.get("commercial_name") or _compose_person_name(normalized)

    if resolved_legal_name or not partial:
        normalized["legal_name"] = resolved_legal_name or ""
    return normalized


def list_clients(
    db: Session,
    *,
    include_inactive: bool = False,
    search: str | None = None,
    status_filter: str | None = None,
) -> list[Client]:
    query = _client_query()
    if not include_inactive and status_filter != "inactive":
        query = query.where(Client.is_active.is_(True))
    if status_filter == "active":
        query = query.where(Client.is_active.is_(True))
    elif status_filter == "inactive":
        query = query.where(Client.is_active.is_(False))
    if search:
        term = f"%{search.strip()}%"
        query = query.where(
            or_(
                Client.legal_name.ilike(term),
                Client.commercial_name.ilike(term),
                Client.rfc.ilike(term),
                Client.email.ilike(term),
            )
        )
    return list(db.scalars(query).all())


def get_client(db: Session, client_id: int, *, include_inactive: bool = False) -> Client:
    client = db.scalar(_client_query().where(Client.id == client_id))
    if client is None or (not include_inactive and not client.is_active):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cliente no encontrado")
    return client


def _sync_contacts(client: Client, contacts_payload: list[dict] | None) -> None:
    if contacts_payload is None:
        return
    client.contacts.clear()
    client.contacts.extend(ClientContact(**contact) for contact in contacts_payload if contact.get("name"))


def create_client(
    db: Session,
    payload: ClientCreate,
    *,
    user_id: int | None = None,
    validate_fiscal: bool = True,
    fiscal_review_required: bool | None = None,
) -> Client:
    data = _normalize_client_data(payload.model_dump(exclude={"contacts"}))
    if validate_fiscal:
        _validate_fiscal_profile(db, data)
    client = Client(**data)
    client.fiscal_review_required = (
        _requires_fiscal_review(data)
        if fiscal_review_required is None
        else fiscal_review_required
    )
    _sync_contacts(client, [contact.model_dump() for contact in payload.contacts])
    db.add(client)
    db.flush()
    write_audit_log(
        db,
        action="client.created",
        entity="clients",
        entity_id=client.id,
        user_id=user_id,
        new_values={"legal_name": client.legal_name, "rfc": client.rfc, "commercial_name": client.commercial_name},
    )
    db.commit()
    db.refresh(client)
    return get_client(db, client.id)


def update_client(db: Session, client_id: int, payload: ClientUpdate, *, user_id: int | None = None) -> Client:
    client = get_client(db, client_id, include_inactive=True)
    updates = _normalize_client_data(payload.model_dump(exclude_unset=True, exclude={"contacts"}), partial=True)
    _validate_fiscal_profile(db, updates)
    previous_values = {key: getattr(client, key) for key in updates}
    previous_contacts = [{"name": item.name, "email": item.email, "phone": item.phone} for item in client.contacts]
    for key, value in updates.items():
        setattr(client, key, value)
    if any(key in updates for key in {"rfc", "legal_name", "tax_regime", "cfdi_use", "fiscal_postal_code"}):
        client.fiscal_review_required = _requires_fiscal_review({
            "rfc": client.rfc,
            "legal_name": client.legal_name,
            "tax_regime": client.tax_regime,
            "cfdi_use": client.cfdi_use,
            "fiscal_postal_code": client.fiscal_postal_code,
        })
    if payload.contacts is not None:
        _sync_contacts(client, [contact.model_dump() for contact in payload.contacts])
        previous_values["contacts"] = previous_contacts
        updates["contacts"] = [contact.model_dump() for contact in payload.contacts]
    write_audit_log(
        db,
        action="client.updated",
        entity="clients",
        entity_id=client.id,
        user_id=user_id,
        previous_values=_json_safe(previous_values),
        new_values=_json_safe(updates),
    )
    db.commit()
    db.refresh(client)
    return get_client(db, client.id)


def deactivate_client(db: Session, client_id: int, *, user_id: int | None = None) -> Client:
    return archive_client(db, client_id, user_id=user_id)


def get_client_delete_eligibility(db: Session, client_id: int) -> ClientDeleteEligibilityRead:
    """Single source of truth for client removal transitions.

    Counts include indirect records reachable through an ETS, so a historical
    equipment/certificate cannot be lost merely because it has no client FK.
    """
    get_client(db, client_id, include_inactive=True)
    invoice_filter = or_(Invoice.client_id == client_id, Invoice.fiscal_client_id == client_id)
    blocking_dependencies = {
        "quotations": int(db.scalar(select(func.count()).select_from(Quotation).where(Quotation.client_id == client_id)) or 0),
        "service_orders": int(db.scalar(select(func.count()).select_from(ServiceOrder).where(ServiceOrder.client_id == client_id)) or 0),
        "equipment": int(db.scalar(select(func.count()).select_from(Equipment).join(ServiceOrder).where(ServiceOrder.client_id == client_id)) or 0),
        "field_sheets": int(db.scalar(select(func.count()).select_from(FieldSheet).join(Equipment).join(ServiceOrder).where(ServiceOrder.client_id == client_id)) or 0),
        "certificates": int(db.scalar(select(func.count()).select_from(Certificate).join(ServiceOrder).where(ServiceOrder.client_id == client_id)) or 0),
        "invoices": int(db.scalar(select(func.count()).select_from(Invoice).where(invoice_filter)) or 0),
        "payments": int(db.scalar(select(func.count()).select_from(InvoicePayment).join(Invoice).where(invoice_filter)) or 0),
        "credit_notes": int(db.scalar(select(func.count()).select_from(CreditNote).join(Invoice).where(invoice_filter)) or 0),
    }
    cascade_dependencies = {
        "contacts": int(db.scalar(select(func.count()).select_from(ClientContact).where(ClientContact.client_id == client_id)) or 0),
        "certificate_profiles": int(db.scalar(select(func.count()).select_from(ClientCertificateProfile).where(ClientCertificateProfile.client_id == client_id)) or 0),
    }
    eligible = not any(blocking_dependencies.values())
    return ClientDeleteEligibilityRead(
        client_id=client_id,
        eligible_for_hard_delete=eligible,
        recommended_action="hard_delete" if eligible else "archive",
        blocking_dependencies=blocking_dependencies,
        cascade_dependencies=cascade_dependencies,
    )


def archive_client(db: Session, client_id: int, *, user_id: int | None = None) -> Client:
    client = get_client(db, client_id, include_inactive=True)
    if not client.is_active:
        return client
    client.is_active = False
    client.deleted_at = datetime.now(timezone.utc)
    client.deleted_by = user_id
    write_audit_log(
        db,
        action="client_archived",
        entity="clients",
        entity_id=client.id,
        user_id=user_id,
        previous_values={"is_active": True},
        new_values={"is_active": False, "mode": "archive"},
    )
    db.commit()
    db.refresh(client)
    return client


def restore_client(db: Session, client_id: int, *, user_id: int | None = None) -> Client:
    client = get_client(db, client_id, include_inactive=True)
    if client.is_active:
        return client
    normalized_rfc = (client.rfc or "").strip().upper()
    if normalized_rfc and normalized_rfc not in GENERIC_RFCS:
        active_match = db.scalar(
            select(Client.id).where(
                Client.id != client.id,
                Client.is_active.is_(True),
                func.upper(Client.rfc) == normalized_rfc,
            )
        )
        if active_match is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="No se puede restaurar: el RFC ya pertenece a otro cliente activo.",
            )
    client.is_active = True
    client.deleted_at = None
    client.deleted_by = None
    write_audit_log(
        db,
        action="client_restored",
        entity="clients",
        entity_id=client.id,
        user_id=user_id,
        previous_values={"is_active": False},
        new_values={"is_active": True, "mode": "restore"},
    )
    db.commit()
    db.refresh(client)
    return client


def delete_client_permanently(
    db: Session, client_id: int, *, user_id: int | None = None
) -> ClientDeleteResultRead:
    client = get_client(db, client_id, include_inactive=True)
    eligibility = get_client_delete_eligibility(db, client.id)
    if not client.is_active:
        return ClientDeleteResultRead(
            status="already_archived",
            delete_mode="archive",
            client_id=client.id,
            message="El cliente ya estaba archivado.",
            blocking_dependencies=eligibility.blocking_dependencies,
        )
    if not eligibility.eligible_for_hard_delete:
        write_audit_log(
            db,
            action="client_delete_blocked",
            entity="clients",
            entity_id=client.id,
            user_id=user_id,
            previous_values={"is_active": True, "rfc": client.rfc},
            new_values={"mode": "archive", "blocking_dependencies": eligibility.blocking_dependencies},
            comment="Eliminación física bloqueada por historial operativo.",
        )
        client.is_active = False
        client.deleted_at = datetime.now(timezone.utc)
        client.deleted_by = user_id
        write_audit_log(
            db,
            action="client_archived",
            entity="clients",
            entity_id=client.id,
            user_id=user_id,
            previous_values={"is_active": True},
            new_values={"is_active": False, "mode": "archive"},
        )
        db.commit()
        return ClientDeleteResultRead(
            status="archived",
            delete_mode="archive",
            client_id=client.id,
            message="El cliente conserva historial y fue archivado.",
            blocking_dependencies=eligibility.blocking_dependencies,
        )
    tax_constancy_path = client.tax_constancy_path
    tax_constancy_filename = client.tax_constancy_filename
    snapshot = {"legal_name": client.legal_name, "rfc": client.rfc, "commercial_name": client.commercial_name}
    write_audit_log(
        db,
        action="client_hard_deleted",
        entity="clients",
        entity_id=client.id,
        user_id=user_id,
        previous_values=snapshot,
        new_values={"deleted": True, "mode": "hard"},
    )
    db.delete(client)
    db.flush()
    delete_if_unreferenced(
        db,
        tax_constancy_path,
        user_id=user_id,
        module="Clientes",
        entity="clients",
        entity_id=client_id,
        filename=tax_constancy_filename,
    )
    db.commit()
    return ClientDeleteResultRead(
        status="deleted",
        delete_mode="hard",
        client_id=client_id,
        message="El cliente fue eliminado definitivamente.",
    )


def _extract_label_value(compact: str, start_label: str, end_label: str | None = None) -> str | None:
    pattern = (
        rf"{start_label}\s*:?\s*(.*?)(?=\s*{end_label}\s*:?)"
        if end_label
        else rf"{start_label}\s*:?\s*(.*)$"
    )
    match = re.search(pattern, compact, re.IGNORECASE | re.DOTALL)

    if not match:
        return None

    value = re.sub(r"\s+", " ", match.group(1)).strip().strip(":-")
    return value or None


def _extract_tax_regimes(compact: str) -> list[str]:
    section_match = re.search(
        r"Reg[ií]menes\s*:?\s*R[eé]gimen\s+Fecha Inicio\s+Fecha Fin\s+(.*?)(?=\s*Obligaciones\s*:|\s*Sus datos personales|\s*Cadena Original|$)",
        compact,
        re.IGNORECASE | re.DOTALL,
    )

    if not section_match:
        return []

    section = re.sub(r"\s+", " ", section_match.group(1)).strip()

    regimes = re.findall(
        r"(R[eé]gimen\s+.*?)(?=\s+\d{2}/\d{2}/\d{4}|$)",
        section,
        re.IGNORECASE,
    )

    values = []
    seen = set()

    for regime in regimes:
        cleaned = re.sub(r"\s+", " ", regime).strip().strip(":-")
        key = _normalize_key(cleaned)

        if cleaned and key not in seen:
            seen.add(key)
            values.append(cleaned)

    return values


def _extract_tax_constancy_fields_from_pdf_bytes(data: bytes) -> dict[str, str | list[str] | None]:
    try:
        reader = PdfReader(BytesIO(data))
        text = "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"No se pudo leer el PDF de constancia: {exc}") from exc

    if not text.strip():
        return {
            "client_type": None,
            "legal_name": None,
            "commercial_name": None,
            "rfc": None,
            "curp": None,
            "first_name": None,
            "first_last_name": None,
            "second_last_name": None,
            "fiscal_postal_code": None,
            "tax_regime": None,
            "tax_regimes": [],
            "street_type": None,
            "street": None,
            "exterior_number": None,
            "interior_number": None,
            "neighborhood": None,
            "locality": None,
            "municipality": None,
            "state": None,
        }

    compact = re.sub(r"[ \t]+", " ", text)
    rfc_match = re.search(
        r"(?:RFC|Registro Federal de Contribuyentes)\s*:?\s*([A-Z&Ñ]{3,4}\d{6}[A-Z0-9]{3})",
        compact,
        re.IGNORECASE,
    )
    curp_match = re.search(r"(?:CURP)\s*:?\s*([A-Z]{4}\d{6}[HM][A-Z]{5}[A-Z0-9]\d)", compact, re.IGNORECASE)
    legal_name = _extract_label_value(
        compact,
        r"(?:Denominaci[oó]n\/Raz[oó]n Social|Nombre, denominaci[oó]n o raz[oó]n social)",
        r"R[eé]gimen(?: de)? Capital",
    )
    first_name = _extract_label_value(
        compact,
        r"(?:Nombre\s*\(s\)|Nombre\s*\(s\)\s+del contribuyente)",
        r"Primer Apellido",
    )
    first_last_name = _extract_label_value(compact, r"Primer Apellido", r"Segundo Apellido")
    second_last_name = _extract_label_value(compact, r"Segundo Apellido", r"Fecha inicio de operaciones")
    commercial_name = _extract_label_value(
        compact,
        r"Nombre Comercial",
        r"(?:Fecha inicio de operaciones|Datos del domicilio registrado)",
    )
    capital_regime = _extract_label_value(compact, r"R[eé]gimen(?: de)? Capital", r"Nombre Comercial")
    postal_code = _extract_label_value(compact, r"(?:C[oó]digo Postal|\bCP\b)", r"Tipo de Vialidad")
    street_type = _extract_label_value(compact, r"Tipo de Vialidad", r"Nombre de Vialidad")
    street = _extract_label_value(compact, r"Nombre de Vialidad", r"N[uú]mero Exterior")
    exterior_number = _extract_label_value(compact, r"N[uú]mero Exterior", r"N[uú]mero Interior")
    interior_number = _extract_label_value(compact, r"N[uú]mero Interior", r"Nombre de la Colonia")
    neighborhood = _extract_label_value(compact, r"(?:Nombre de la Colonia|Colonia)", r"Nombre de la Localidad")
    locality = _extract_label_value(compact, r"(?:Nombre de la Localidad|Localidad)", r"Nombre del Municipio o Demarcaci[oó]n Territorial")
    municipality = _extract_label_value(
        compact,
        r"Nombre del Municipio o Demarcaci[oó]n Territorial",
        r"Nombre de la Entidad Federativa",
    )
    state = _extract_label_value(compact, r"Nombre de la Entidad Federativa", r"Entre Calle")
    tax_regimes = _extract_tax_regimes(compact)
    client_type = None
    if curp_match and first_name and first_last_name:
        client_type = "persona_fisica"
    elif legal_name and capital_regime:
        client_type = "persona_moral"
    elif curp_match:
        client_type = "persona_fisica"
    elif legal_name:
        client_type = "persona_moral"
    tax_regime = tax_regimes[0] if len(tax_regimes) == 1 else None
    return {
        "client_type": client_type,
        "legal_name": legal_name or None,
        "commercial_name": commercial_name or None,
        "rfc": rfc_match.group(1).upper() if rfc_match else None,
        "curp": curp_match.group(1).upper() if curp_match else None,
        "first_name": first_name or None,
        "first_last_name": first_last_name or None,
        "second_last_name": second_last_name or None,
        "fiscal_postal_code": postal_code or None,
        "tax_regime": tax_regime or None,
        "tax_regimes": tax_regimes,
        "street_type": street_type or None,
        "street": street or None,
        "exterior_number": exterior_number or None,
        "interior_number": interior_number or None,
        "neighborhood": neighborhood or None,
        "locality": locality or None,
        "municipality": municipality or None,
        "state": state or None,
    }


def preview_tax_constancy(upload: UploadFile) -> ClientTaxConstancyPreviewRead:
    validated = validate_upload(upload, "tax_constancy")
    original_name = validated.original_filename
    extension = validated.extension
    data = validated.content
    if extension != ".pdf":
        return ClientTaxConstancyPreviewRead(
            available=False,
            filename=original_name,
            message="La extracción automática de constancia aún no está disponible para este tipo de archivo.",
        )

    extracted = _extract_tax_constancy_fields_from_pdf_bytes(data)
    if not any(extracted.values()):
        return ClientTaxConstancyPreviewRead(
            available=False,
            filename=original_name,
            message="No se pudieron extraer datos fiscales del PDF de constancia de forma automática.",
        )

    return ClientTaxConstancyPreviewRead(
        available=True,
        filename=original_name,
        message="Se detectaron datos fiscales en la constancia. Revisa y confirma antes de guardar.",
        extracted_client_type=extracted["client_type"],
        extracted_legal_name=extracted["legal_name"],
        extracted_commercial_name=extracted["commercial_name"],
        extracted_rfc=extracted["rfc"],
        extracted_curp=extracted["curp"],
        extracted_first_name=extracted["first_name"],
        extracted_first_last_name=extracted["first_last_name"],
        extracted_second_last_name=extracted["second_last_name"],
        extracted_fiscal_postal_code=extracted["fiscal_postal_code"],
        extracted_tax_regime=extracted["tax_regime"],
        extracted_tax_regimes=extracted["tax_regimes"],
        extracted_street_type=extracted["street_type"],
        extracted_street=extracted["street"],
        extracted_exterior_number=extracted["exterior_number"],
        extracted_interior_number=extracted["interior_number"],
        extracted_neighborhood=extracted["neighborhood"],
        extracted_locality=extracted["locality"],
        extracted_municipality=extracted["municipality"],
        extracted_state=extracted["state"],
    )


def upload_tax_constancy(
    db: Session,
    client_id: int,
    upload: UploadFile,
    *,
    user_id: int | None = None,
) -> Client:
    client = get_client(db, client_id)
    validated = validate_upload(upload, "tax_constancy")
    original_name = validated.original_filename
    extension = validated.extension
    filename = f"constancia_fiscal_{uuid4().hex}{extension}"
    stored_file = save_validated_content(
        directory=Path("clientes") / f"cliente_{client.id}",
        filename=filename,
        content=validated.content,
        original_filename=original_name,
    )
    previous = {
        "tax_constancy_filename": client.tax_constancy_filename,
        "tax_constancy_path": client.tax_constancy_path,
    }
    client.tax_constancy_filename = original_name
    client.tax_constancy_path = stored_file.relative_path
    client.tax_constancy_uploaded_at = datetime.now(timezone.utc)
    write_audit_log(
        db,
        action="client.tax_constancy_uploaded",
        entity="clients",
        entity_id=client.id,
        user_id=user_id,
        previous_values=previous,
        new_values={
            "tax_constancy_filename": client.tax_constancy_filename,
            "tax_constancy_path": client.tax_constancy_path,
        },
    )
    delete_if_unreferenced(
        db,
        previous["tax_constancy_path"],
        user_id=user_id,
        module="Clientes",
        entity="clients",
        entity_id=client.id,
        filename=previous["tax_constancy_filename"],
    )
    db.commit()
    db.refresh(client)
    return get_client(db, client.id)


def _read_tabular_file(upload: UploadFile) -> tuple[list[str], list[dict[str, str]]]:
    validated = validate_upload(upload, "client_import")
    suffix = validated.extension
    data = validated.content

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], []
        columns = [str(value).strip() if value is not None else "" for value in rows[0]]
        parsed_rows = []
        for raw_row in rows[1:]:
            if raw_row is None:
                continue
            values = ["" if value is None else str(value).strip() for value in raw_row]
            if not any(values):
                continue
            parsed_rows.append({columns[index]: values[index] if index < len(values) else "" for index in range(len(columns))})
        return columns, parsed_rows
    if suffix == ".xls":
        raise HTTPException(status_code=400, detail="El formato .xls no está soportado. Usa .xlsx o CSV.")

    text = data.decode("utf-8-sig", errors="ignore")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(StringIO(text), dialect=dialect)
    columns = reader.fieldnames or []
    rows = [{key: (value or "").strip() for key, value in row.items()} for row in reader if row]
    return columns, [row for row in rows if any(str(value).strip() for value in row.values())]


def _get_row_value(row: dict[str, str], names: list[str]) -> str:
    normalized = {_normalize_key(key): value for key, value in row.items()}
    for name in names:
        value = normalized.get(_normalize_key(name))
        if value is not None:
            return str(value).strip()
    return ""


def _import_fiscal_warnings(db: Session, row: dict[str, str], cache: dict | None = None) -> list[str]:
    warnings = []
    rfc = _get_row_value(row, ["rfc", "RFC"]).strip().upper()
    if not rfc:
        warnings.append("RFC ausente; cliente importado y marcado para revisión fiscal.")
    elif not RFC_PATTERN.fullmatch(rfc):
        warnings.append("RFC con formato no reconocido; cliente importado y marcado para revisión fiscal.")
    for catalog, names, label in (
        ("fiscal_regimes", ["regimen_fiscal", "Regimen fiscal", "Régimen fiscal"], "Régimen fiscal"),
        ("cfdi_uses", ["uso_cfdi", "Uso CFDI"], "Uso CFDI"),
        ("postal_codes", ["codigo_postal", "Codigo postal", "Código postal"], "Código postal fiscal"),
    ):
        value = _get_row_value(row, names)
        if not value:
            warnings.append(f"{label} ausente; cliente importado y marcado para revisión fiscal.")
        elif _resolve_sat_import_value(db, catalog, value, cache) is None:
            warnings.append(f"{label} no reconocido ({value}); cliente importado y marcado para revisión fiscal.")
    return warnings


def _build_import_preview(
    db: Session,
    rows: list[dict[str, str]],
    existing_clients: list[Client],
    resolver_cache: dict | None = None,
) -> ClientImportPreviewRead:
    resolver_cache = resolver_cache if resolver_cache is not None else {}
    active_clients = [client for client in existing_clients if client.is_active]
    inactive_by_rfc: dict[str, list[Client]] = {}
    for client in existing_clients:
        if not client.is_active and client.rfc:
            inactive_by_rfc.setdefault(_normalize_key(client.rfc), []).append(client)
    existing_rfc = {_normalize_key(client.rfc) for client in active_clients if client.rfc}
    existing_email = {_normalize_key(client.email) for client in active_clients if client.email}
    existing_name = {
        _normalize_key(client.commercial_name or client.legal_name)
        for client in active_clients
        if client.commercial_name or client.legal_name
    }

    seen_rfc: set[str] = set()
    seen_email: set[str] = set()
    seen_name: set[str] = set()

    preview_rows: list[ClientImportRowRead] = []

    for index, row in enumerate(rows):
        name = _get_row_value(row, ["nombre_comercial", "Nombre comercial", "nombre", "Cliente"])
        legal_name = _get_row_value(row, ["razon_social", "Razon social", "Razón social"])
        first_name = _get_row_value(row, ["nombres", "Nombre(s)", "Nombre"])
        first_last_name = _get_row_value(row, ["primer_apellido", "Primer apellido"])
        second_last_name = _get_row_value(row, ["segundo_apellido", "Segundo apellido"])
        contact_name = _get_row_value(row, ["contacto", "Contacto principal", "Contacto"])
        rfc = _get_row_value(row, ["rfc", "RFC"])
        email = _get_row_value(row, ["correo", "Correo", "Email"])

        person_name = " ".join(
            part for part in [first_name.strip(), first_last_name.strip(), second_last_name.strip()] if part.strip()
        )

        display_name = name.strip() or legal_name.strip() or person_name.strip() or contact_name.strip()

        name_key = _normalize_key(display_name)
        rfc_key = _normalize_key(rfc)
        email_key = _normalize_key(email)

        errors: list[str] = []
        duplicates: list[str] = []
        warnings = _import_fiscal_warnings(db, row, resolver_cache) if display_name else []

        if not display_name:
            errors.append("Nombre comercial, razón social o nombre del cliente obligatorio")

        if rfc_key and (rfc_key in existing_rfc or rfc_key in seen_rfc):
            duplicates.append("RFC")
        if email_key and (email_key in existing_email or email_key in seen_email):
            duplicates.append("Correo")
        if name_key and (name_key in existing_name or name_key in seen_name):
            duplicates.append("Nombre")

        if rfc_key:
            seen_rfc.add(rfc_key)
        if email_key:
            seen_email.add(email_key)
        if name_key:
            seen_name.add(name_key)

        archived_matches = inactive_by_rfc.get(rfc_key, [])
        ambiguous = False
        if rfc_key in GENERIC_RFCS and archived_matches:
            errors.append("RFC genérico: no se restaura automáticamente porque no identifica inequívocamente al cliente.")
            ambiguous = True
        elif len(archived_matches) > 1:
            errors.append("Hay más de un cliente archivado con el mismo RFC; requiere revisión manual.")
            ambiguous = True
        elif len(archived_matches) == 1 and not duplicates:
            warnings.append("Cliente archivado encontrado por RFC; acción propuesta: restaurar y actualizar.")

        row_status = "ambiguous" if ambiguous else "error" if errors else "duplicate" if duplicates else "warning" if warnings else "valid"

        preview_rows.append(
            ClientImportRowRead(
                id=f"{index}-{name_key or 'cliente'}",
                name=display_name,
                rfc=rfc or "-",
                email=email or "-",
                status=row_status,
                errors=errors,
                duplicates=duplicates,
                warnings=warnings,
                raw=row,
            )
        )

    return ClientImportPreviewRead(
        columns=list(rows[0].keys()) if rows else [],
        rows=preview_rows,
        valid_count=len([row for row in preview_rows if row.status in {"valid", "warning"}]),
        duplicate_count=len([row for row in preview_rows if row.status == "duplicate"]),
        error_count=len([row for row in preview_rows if row.status in {"error", "ambiguous"}]),
        warning_count=len([row for row in preview_rows if row.status == "warning"]),
    )

def preview_client_import(db: Session, upload: UploadFile) -> ClientImportPreviewRead:
    columns, rows = _read_tabular_file(upload)
    if not columns:
        return ClientImportPreviewRead(columns=[], rows=[], valid_count=0, duplicate_count=0, error_count=0)
    preview = _build_import_preview(db, rows, list_clients(db, include_inactive=True))
    preview.columns = columns
    return preview

def _clean_import_email(value: str | None) -> str:
    email = (value or "").strip().strip("<>").strip()
    email = email.replace("mailto:", "").strip()
    email = email.rstrip(";,>").strip()

    if re.fullmatch(r"^[^\s@<>;,]+@[^\s@<>;,]+\.[^\s@<>;,]+$", email):
        return email

    return ""

def _row_to_client_payload(
    db: Session,
    row: dict[str, str],
    resolver_cache: dict | None = None,
) -> tuple[ClientCreate, list[str]]:
    name = _get_row_value(row, ["nombre_comercial", "Nombre comercial", "nombre", "Cliente"]).strip()
    legal_name = _get_row_value(row, ["razon_social", "Razon social", "Razón social"]).strip()
    first_name = _get_row_value(row, ["nombres", "Nombre(s)", "Nombre"]).strip()
    first_last_name = _get_row_value(row, ["primer_apellido", "Primer apellido"]).strip()
    second_last_name = _get_row_value(row, ["segundo_apellido", "Segundo apellido"]).strip()
    client_type = _normalize_client_type(_get_row_value(row, ["tipo_cliente", "Tipo de cliente"]))
    email = _get_row_value(row, ["correo", "Correo", "Email"]).strip()
    phone = _get_row_value(row, ["telefono", "Telefono", "Teléfono"]).strip()
    contact_name = _get_row_value(row, ["contacto", "Contacto principal", "Contacto"]).strip()

    fallback_name = (
        name
        or legal_name
        or " ".join(part for part in [first_name, first_last_name, second_last_name] if part).strip()
        or contact_name
        or ""
    )

    valid_email = _clean_import_email(email)

    postal_code = _get_row_value(row, ["codigo_postal", "Codigo postal", "Código postal"]).strip()
    resolved_tax_regime = _resolve_sat_import_value(db, "fiscal_regimes", _get_row_value(row, ["regimen_fiscal", "Regimen fiscal", "Régimen fiscal"]), resolver_cache)
    resolved_cfdi_use = _resolve_sat_import_value(db, "cfdi_uses", _get_row_value(row, ["uso_cfdi", "Uso CFDI"]), resolver_cache)
    resolved_postal_code = _resolve_sat_import_value(db, "postal_codes", postal_code, resolver_cache)
    warnings = _import_fiscal_warnings(db, row, resolver_cache)

    return ClientCreate(
        client_type=client_type,
        commercial_name=name or fallback_name,
        legal_name=legal_name or fallback_name,
        rfc=_get_row_value(row, ["rfc", "RFC"]).strip().upper() or None,
        curp=_get_row_value(row, ["curp", "CURP"]).strip().upper() or None,
        first_name=first_name or None,
        first_last_name=first_last_name or None,
        second_last_name=second_last_name or None,
        email=valid_email or None,
        phone=phone or None,
        tax_regime=resolved_tax_regime,
        cfdi_use=resolved_cfdi_use,
        street_type=_get_row_value(row, ["tipo_vialidad", "Tipo de vialidad"]).strip() or None,
        street=_get_row_value(row, ["calle", "Calle"]).strip() or None,
        exterior_number=_get_row_value(row, ["numero_exterior", "Numero exterior", "Número exterior"]).strip() or None,
        interior_number=_get_row_value(row, ["numero_interior", "Numero interior", "Número interior"]).strip() or None,
        neighborhood=_get_row_value(row, ["colonia", "Colonia"]).strip() or None,
        locality=_get_row_value(row, ["localidad", "Localidad"]).strip() or None,
        municipality=_get_row_value(row, ["municipio", "Municipio", "municipio_ciudad", "Municipio / Ciudad", "Ciudad"]).strip() or None,
        city=_get_row_value(row, ["municipio_ciudad", "Municipio / Ciudad", "Ciudad", "municipio", "Municipio"]).strip() or None,
        state=_get_row_value(row, ["estado", "Estado"]).strip() or None,
        postal_code=postal_code if postal_code.isdigit() else None,
        fiscal_postal_code=resolved_postal_code,
        country=_get_row_value(row, ["pais", "Pais", "País"]).strip() or "Mexico",
        contacts=[
            {
                "name": contact_name,
                "email": valid_email or None,
                "phone": phone or None,
                "position": None,
            }
        ]
        if contact_name
        else [],
    ), warnings


def confirm_client_import(
    db: Session,
    payload: ClientImportConfirm,
    *,
    user_id: int | None = None,
) -> ClientImportResultRead:
    resolver_cache: dict = {}
    preview = _build_import_preview(db, payload.rows, list_clients(db, include_inactive=True), resolver_cache)
    imported_ids: list[int] = []
    errors: list[dict] = []
    warnings: list[dict] = []
    duplicate_count = len([row for row in preview.rows if row.status == "duplicate"])
    error_count = len([row for row in preview.rows if row.status in {"error", "ambiguous"}])
    omitted_count = duplicate_count + error_count

    for index, row in enumerate(preview.rows, start=1):
        if row.status not in {"valid", "warning"}:
            continue
        try:
            client_payload, row_warnings = _row_to_client_payload(db, row.raw, resolver_cache)
            rfc_key = _normalize_key(client_payload.rfc)
            archived_matches = (
                list(db.scalars(select(Client).where(func.upper(Client.rfc) == client_payload.rfc, Client.is_active.is_(False))).all())
                if rfc_key and rfc_key not in GENERIC_RFCS
                else []
            )
            if len(archived_matches) == 1:
                archived = archived_matches[0]
                restored = restore_client(db, archived.id, user_id=user_id)
                updates = _normalize_client_data(client_payload.model_dump(exclude={"contacts"}), partial=True)
                for key, value in updates.items():
                    if value is not None:
                        setattr(restored, key, value)
                restored.fiscal_review_required = bool(row_warnings)
                if not restored.contacts and client_payload.contacts:
                    _sync_contacts(restored, [contact.model_dump() for contact in client_payload.contacts])
                write_audit_log(
                    db,
                    action="client_import_restored",
                    entity="clients",
                    entity_id=restored.id,
                    user_id=user_id,
                    previous_values={"is_active": False, "rfc": restored.rfc},
                    new_values={"is_active": True, "source": "import"},
                )
                db.commit()
                imported_ids.append(restored.id)
                row_warnings = [*row_warnings, "Cliente archivado restaurado y actualizado por coincidencia exacta de RFC."]
            else:
                created = create_client(
                    db,
                    client_payload,
                    user_id=user_id,
                    validate_fiscal=False,
                    fiscal_review_required=bool(row_warnings),
                )
                imported_ids.append(created.id)
                write_audit_log(
                    db,
                    action="client_import_created",
                    entity="clients",
                    entity_id=created.id,
                    user_id=user_id,
                    new_values={"rfc": created.rfc, "source": "import"},
                )
                db.commit()
            warnings.extend({"row": index, "name": row.name, "message": warning} for warning in row_warnings)
        except Exception as exc:  # noqa: BLE001
            db.rollback()
            omitted_count += 1
            errors.append({"row": row.raw, "message": str(exc)})

    return ClientImportResultRead(
        imported_count=len(imported_ids),
        omitted_count=omitted_count,
        duplicate_count=duplicate_count,
        error_count=error_count + len(errors),
        imported_ids=imported_ids,
        errors=errors,
        total_rows=len(preview.rows),
        imported_with_warnings_count=len([row for row in preview.rows if row.status == "warning"]),
        warning_count=len(warnings),
        warnings=warnings,
    )


def export_clients_workbook(
    db: Session,
    *,
    include_inactive: bool = True,
    search: str | None = None,
    status_filter: str | None = None,
) -> tuple[bytes, str]:
    clients = list_clients(db, include_inactive=include_inactive, search=search, status_filter=status_filter)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clientes"
    sheet.append(CLIENT_IMPORT_COLUMNS)

    for client in clients:
        contact = client.contacts[0] if client.contacts else None
        sheet.append(
            [
                "Persona Física" if client.client_type == "persona_fisica" else "Persona Moral",
                client.commercial_name or "",
                client.legal_name or "",
                client.curp or "",
                client.first_name or "",
                client.first_last_name or "",
                client.second_last_name or "",
                client.rfc or "",
                contact.name if contact else "",
                client.phone or (contact.phone if contact else "") or "",
                client.email or (contact.email if contact else "") or "",
                client.country or "",
                client.street_type or "",
                client.street or "",
                client.exterior_number or "",
                client.interior_number or "",
                client.neighborhood or "",
                client.locality or "",
                client.municipality or client.city or "",
                client.city or "",
                client.state or "",
                client.postal_code or "",
                client.tax_regime or "",
                client.cfdi_use or "",
                "Activo" if client.is_active else "Inactivo",
            ]
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), "clientes_myc_export.xlsx"
