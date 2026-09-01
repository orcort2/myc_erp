from __future__ import annotations

import re
import unicodedata
from io import BytesIO

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import and_, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.models.lab_client import LabClient
from app.models.user import User
from app.schemas.lab_client import LabClientCreate, LabClientImportSummary
from app.services.audit_logs import write_audit_log


def normalize_lab_client_identity(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value.strip())
    without_marks = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", without_marks.casefold()).strip()


def _scope_clause(operator_client_id: int | None):
    return (
        LabClient.operator_client_id.is_(None)
        if operator_client_id is None
        else LabClient.operator_client_id == operator_client_id
    )


def list_lab_clients(
    db: Session, *, operator_client_id: int | None, search: str | None = None
) -> list[LabClient]:
    query = select(LabClient).where(_scope_clause(operator_client_id))
    if search and search.strip():
        value = f"%{search.strip()}%"
        query = query.where(
            or_(
                LabClient.company.ilike(value),
                LabClient.address.ilike(value),
                LabClient.attention.ilike(value),
            )
        )
    return list(db.scalars(query.order_by(LabClient.company, LabClient.attention)).all())


def get_lab_client(db: Session, client_id: int, *, operator_client_id: int | None) -> LabClient:
    client = db.scalar(
        select(LabClient).where(
            LabClient.id == client_id,
            _scope_clause(operator_client_id),
        )
    )
    if client is None:
        raise HTTPException(status_code=404, detail="Cliente LAB no encontrado")
    return client


def create_lab_client(
    db: Session,
    payload: LabClientCreate,
    user: User,
    *,
    operator_client_id: int | None,
) -> LabClient:
    values = {
        "company": payload.company.strip(),
        "address": payload.address.strip(),
        "attention": payload.attention.strip(),
    }
    normalized = {
        f"normalized_{key}": normalize_lab_client_identity(value)
        for key, value in values.items()
    }
    existing = db.scalar(
        select(LabClient).where(
            _scope_clause(operator_client_id),
            LabClient.normalized_company == normalized["normalized_company"],
            LabClient.normalized_address == normalized["normalized_address"],
            LabClient.normalized_attention == normalized["normalized_attention"],
        )
    )
    if existing is not None:
        return existing
    client = LabClient(
        **values,
        **normalized,
        operator_client_id=operator_client_id,
        created_by_user_id=user.id,
    )
    db.add(client)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="El cliente LAB ya existe") from exc
    write_audit_log(
        db,
        action="lab_client.created_inline",
        entity="lab_clients",
        entity_id=client.id,
        user_id=user.id,
        new_values={**values, "operator_client_id": operator_client_id},
    )
    db.commit()
    db.refresh(client)
    return client


async def import_lab_clients_xlsx(
    db: Session, upload: UploadFile, user: User
) -> LabClientImportSummary:
    if not (upload.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="La importación requiere un archivo XLSX")
    content = await upload.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="El XLSX excede 10 MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="El XLSX no es válido") from exc
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if headers is None:
        raise HTTPException(status_code=422, detail="El XLSX está vacío")
    positions = {str(value or "").strip().upper(): index for index, value in enumerate(headers)}
    required = {"CLIENTE", "CONTACTO", "DIRECCIÓN"}
    if not required.issubset(positions):
        raise HTTPException(status_code=422, detail="Se requieren CLIENTE, CONTACTO y DIRECCIÓN")

    existing_keys = set(
        db.execute(
            select(
                LabClient.normalized_company,
                LabClient.normalized_address,
                LabClient.normalized_attention,
            ).where(LabClient.operator_client_id.is_(None))
        ).all()
    )
    new_count = skipped = invalid = 0
    errors: list[dict] = []
    for row_number, row in enumerate(rows, start=2):
        values = {
            "company": str(row[positions["CLIENTE"]] or "").strip(),
            "address": str(row[positions["DIRECCIÓN"]] or "").strip(),
            "attention": str(row[positions["CONTACTO"]] or "").strip(),
        }
        if not values["company"]:
            invalid += 1
            errors.append({"row": row_number, "reason": "Falta Empresa"})
            continue
        key = tuple(normalize_lab_client_identity(values[name]) for name in ("company", "address", "attention"))
        if key in existing_keys:
            skipped += 1
            continue
        client = LabClient(
            **values,
            normalized_company=key[0],
            normalized_address=key[1],
            normalized_attention=key[2],
            operator_client_id=None,
            created_by_user_id=user.id,
        )
        db.add(client)
        db.flush()
        write_audit_log(
            db,
            action="lab_client.imported",
            entity="lab_clients",
            entity_id=client.id,
            user_id=user.id,
            new_values={**values, "source_row": row_number},
        )
        existing_keys.add(key)
        new_count += 1
    db.commit()
    return LabClientImportSummary(new=new_count, skipped=skipped, invalid=invalid, errors=errors)
