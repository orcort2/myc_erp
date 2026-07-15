import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from app.services.sat_catalogs.normalizers import normalize_header


class SatCatalogParseError(ValueError):
    pass


SUPPORTED_EXTENSIONS = {".csv", ".json", ".xlsx"}


def parse_file(path: Path) -> list[dict[str, object]]:
    extension = path.suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise SatCatalogParseError("Formato no compatible. Usa CSV, JSON o XLSX.")
    if extension == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    if extension == ".json":
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if isinstance(payload, dict):
            payload = payload.get("records") or payload.get("items")
        if not isinstance(payload, list) or not all(isinstance(row, dict) for row in payload):
            raise SatCatalogParseError("El JSON debe ser una lista de objetos o contener records/items.")
        return payload
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration as error:
        raise SatCatalogParseError("El XLSX está vacío.") from error
    normalized_headers = [normalize_header(header) for header in headers]
    if not any(normalized_headers):
        raise SatCatalogParseError("El XLSX no contiene encabezados.")
    return [dict(zip(normalized_headers, values)) for values in rows if any(value is not None for value in values)]
