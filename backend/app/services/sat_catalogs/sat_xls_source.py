"""Read official SAT Excel catalog releases without executing workbook content."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook

from app.services.sat_catalogs.normalizers import normalize_header


class SatXlsSourceError(ValueError):
    pass


@dataclass(frozen=True)
class CatalogSheetDefinition:
    sheets: tuple[str, ...]
    code_headers: tuple[str, ...]
    name_headers: tuple[str, ...] = ("descripcion", "nombre")


CATALOG_SHEETS: dict[str, CatalogSheetDefinition] = {
    "payment_forms": CatalogSheetDefinition(("c_FormaPago",), ("c_formapago",)),
    "currencies": CatalogSheetDefinition(("c_Moneda",), ("c_moneda",)),
    "voucher_types": CatalogSheetDefinition(("c_TipoDeComprobante",), ("c_tipodecomprobante",)),
    "exports": CatalogSheetDefinition(("c_Exportacion",), ("c_exportacion",)),
    "payment_methods": CatalogSheetDefinition(("c_MetodoPago",), ("c_metodopago",)),
    "postal_codes": CatalogSheetDefinition(("c_CodigoPostal_Parte_1", "c_CodigoPostal_Parte_2"), ("c_codigopostal",)),
    "relation_types": CatalogSheetDefinition(("c_TipoRelacion",), ("c_tiporelacion",)),
    "fiscal_regimes": CatalogSheetDefinition(("c_RegimenFiscal",), ("c_regimenfiscal",)),
    "countries": CatalogSheetDefinition(("c_Pais",), ("c_pais",)),
    "cfdi_uses": CatalogSheetDefinition(("c_UsoCFDI",), ("c_usocfdi",)),
    "products_services": CatalogSheetDefinition(("c_ClaveProdServ",), ("c_claveprodserv",)),
    "units": CatalogSheetDefinition(("c_ClaveUnidad",), ("c_claveunidad",), ("nombre", "descripcion")),
    "tax_objects": CatalogSheetDefinition(("c_ObjetoImp",), ("c_objetoimp",)),
    "taxes": CatalogSheetDefinition(("c_Impuesto",), ("c_impuesto",)),
    "factor_types": CatalogSheetDefinition(("c_TipoFactor",), ("c_tipofactor",), ()),
    "tax_rates": CatalogSheetDefinition(("c_TasaOCuota",), ("c_tasaocuota",), ()),
}

DATE_FROM_HEADERS = ("fecha_inicio_de_vigencia", "fecha_inicio_vigencia", "fechainiciovigencia", "fecha_de_inicio_de_vigencia")
DATE_UNTIL_HEADERS = ("fecha_fin_de_vigencia", "fecha_fin_vigencia", "fechafinvigencia", "fecha_de_fin_de_vigencia")
SUPPORTED_EXTENSIONS = {".xls", ".xlsx"}


@dataclass(frozen=True)
class _Cell:
    value: object
    number_format: str = "General"


def _ensure_source(path: str | Path) -> Path:
    source = Path(path).expanduser()
    if not source.is_file():
        raise SatXlsSourceError(f"Fuente Excel SAT no encontrada: {source}")
    if source.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise SatXlsSourceError("La fuente SAT debe tener extensión .xls o .xlsx.")
    return source


def _format_number(value: int | float, number_format: str) -> str:
    pattern = (number_format or "General").split(";")[0]
    pattern = re.sub(r'"[^"]*"|\\.', "", pattern)
    integer_pattern = re.fullmatch(r"0+", pattern)
    if integer_pattern:
        return str(int(value)).zfill(len(pattern))
    decimal_match = re.search(r"\.([0]+)", pattern)
    if decimal_match:
        return f"{float(value):.{len(decimal_match.group(1))}f}"
    if float(value).is_integer():
        return str(int(value))
    return format(float(value), "f").rstrip("0").rstrip(".")


def _cell_text(cell: _Cell) -> str | None:
    value = cell.value
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if isinstance(value, (int, float)):
        return _format_number(value, cell.number_format)
    text = str(value).strip()
    return text or None


def _xlsx_rows(sheet, *, start_row: int = 1, max_row: int | None = None, max_col: int | None = None) -> Iterable[list[_Cell]]:
    for row in sheet.iter_rows(min_row=start_row, max_row=max_row, max_col=max_col):
        yield [_Cell(cell.value, cell.number_format) for cell in row]


def _xls_rows(book, sheet, *, start_row: int = 1, max_row: int | None = None, max_col: int | None = None) -> Iterable[list[_Cell]]:
    import xlrd

    limit = min(sheet.nrows, max_row) if max_row else sheet.nrows
    width = min(sheet.ncols, max_col) if max_col else sheet.ncols
    for index in range(start_row - 1, limit):
        row: list[_Cell] = []
        for column in range(width):
            cell = sheet.cell(index, column)
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
            number_format = "General"
            if cell.xf_index is not None and cell.xf_index < len(book.xf_list):
                format_key = book.xf_list[cell.xf_index].format_key
                number_format = book.format_map.get(format_key, type("Format", (), {"format_str": "General"})()).format_str
            row.append(_Cell(value, number_format))
        yield row


def _open_workbook(path: Path):
    if path.suffix.lower() == ".xlsx":
        try:
            return "xlsx", load_workbook(path, read_only=True, data_only=True, keep_links=False)
        except Exception as error:  # pragma: no cover - library-specific corrupt file errors
            raise SatXlsSourceError(f"No fue posible abrir el XLSX oficial: {error}") from error
    try:
        import xlrd
    except ImportError as error:  # pragma: no cover - exercised in deployed dependency environment
        raise SatXlsSourceError("La lectura de .xls requiere xlrd==2.0.1.") from error
    try:
        return "xls", xlrd.open_workbook(path, on_demand=True, formatting_info=True)
    except Exception as error:  # pragma: no cover - library-specific corrupt file errors
        raise SatXlsSourceError(f"No fue posible abrir el XLS oficial: {error}") from error


def _sheet_names(kind: str, workbook) -> list[str]:
    return list(workbook.sheetnames) if kind == "xlsx" else list(workbook.sheet_names())


def _sheet_rows(kind: str, workbook, name: str, **kwargs) -> Iterable[list[_Cell]]:
    sheet = workbook[name] if kind == "xlsx" else workbook.sheet_by_name(name)
    return _xlsx_rows(sheet, **kwargs) if kind == "xlsx" else _xls_rows(workbook, sheet, **kwargs)


def _close_workbook(kind: str, workbook) -> None:
    if kind == "xlsx":
        workbook.close()
    else:
        workbook.release_resources()


def _find_header(kind: str, workbook, sheet_name: str, definition: CatalogSheetDefinition) -> tuple[int, list[str], int]:
    for row_index, row in enumerate(_sheet_rows(kind, workbook, sheet_name, start_row=1, max_row=35), start=1):
        headers = [normalize_header(_cell_text(cell)) for cell in row]
        if not any(header in definition.code_headers for header in headers):
            continue
        width = max((index + 1 for index, header in enumerate(headers) if header), default=0)
        if width:
            return row_index, headers[:width], width
    raise SatXlsSourceError(f"No se localizó el encabezado de clave para la hoja {sheet_name}.")


def _unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for index, header in enumerate(headers, start=1):
        base = header or f"column_{index}"
        seen[base] = seen.get(base, 0) + 1
        result.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return result


def _first_index(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    return next((index for index, header in enumerate(headers) if header in candidates), None)


def _extract_sheet_rows(kind: str, workbook, sheet_name: str, definition: CatalogSheetDefinition) -> tuple[dict[str, object], list[dict[str, object]]]:
    header_row, raw_headers, width = _find_header(kind, workbook, sheet_name, definition)
    is_tax_rate_sheet = sheet_name == "c_TasaOCuota"
    if is_tax_rate_sheet:
        secondary_row = next(_sheet_rows(kind, workbook, sheet_name, start_row=header_row + 1, max_row=header_row + 1, max_col=width))
        secondary_headers = [normalize_header(_cell_text(cell)) for cell in secondary_row]
        raw_headers = [
            secondary_headers[index] if secondary_headers[index] else header
            for index, header in enumerate(raw_headers)
        ]
    headers = _unique_headers(raw_headers)
    code_index = _first_index(raw_headers, definition.code_headers)
    name_index = _first_index(raw_headers, definition.name_headers)
    valid_from_index = _first_index(raw_headers, DATE_FROM_HEADERS)
    valid_until_index = _first_index(raw_headers, DATE_UNTIL_HEADERS)
    if code_index is None and not is_tax_rate_sheet:
        raise SatXlsSourceError(f"La hoja {sheet_name} no contiene columna de clave.")
    rows: list[dict[str, object]] = []
    start_row = header_row + 2 if is_tax_rate_sheet else header_row + 1
    for row in _sheet_rows(kind, workbook, sheet_name, start_row=start_row, max_col=width):
        values = [_cell_text(cell) for cell in row]
        if is_tax_rate_sheet:
            minimum = values[1] if len(values) > 1 else None
            maximum = values[2] if len(values) > 2 else None
            rate_kind = values[0] if values else None
            code = maximum if rate_kind == "Fijo" else f"{minimum}..{maximum}" if minimum and maximum else None
        else:
            code = values[code_index] if code_index < len(values) else None
        if not code:
            continue
        data = {headers[index]: value for index, value in enumerate(values) if value is not None}
        if is_tax_rate_sheet:
            data["c_tasaocuota"] = code
        rows.append({
            "code": code,
            "name": (
                f"{values[0] or ''} · {values[3] or ''} · {values[4] or ''}".strip(" ·")
                if is_tax_rate_sheet
                else values[name_index] if name_index is not None and name_index < len(values) else None
            ),
            "valid_from": values[valid_from_index] if valid_from_index is not None and valid_from_index < len(values) else None,
            "valid_until": values[valid_until_index] if valid_until_index is not None and valid_until_index < len(values) else None,
            **data,
        })
    descriptor = {
        "sheet": sheet_name,
        "header_row": header_row,
        "columns": [header for header in raw_headers if header],
        "record_count": len(rows),
    }
    return descriptor, rows


def extract_catalog_rows(path: str | Path, catalog_code: str) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Return normalized rows and source sheet details for one internal catalog."""
    source = _ensure_source(path)
    definition = CATALOG_SHEETS.get(catalog_code)
    if definition is None:
        raise SatXlsSourceError(f"No hay mapeo Excel SAT para el catálogo {catalog_code}.")
    kind, workbook = _open_workbook(source)
    try:
        available = set(_sheet_names(kind, workbook))
        missing = set(definition.sheets) - available
        if missing:
            raise SatXlsSourceError(f"Faltan hojas requeridas para {catalog_code}: {', '.join(sorted(missing))}")
        all_rows: list[dict[str, object]] = []
        sheets: list[dict[str, object]] = []
        seen: set[str] = set()
        for sheet_name in definition.sheets:
            descriptor, rows = _extract_sheet_rows(kind, workbook, sheet_name, definition)
            for row in rows:
                code = str(row["code"])
                if code in seen:
                    if catalog_code == "tax_rates":
                        suffix = " | ".join(str(row.get(key) or "-") for key in ("impuesto", "factor", "traslado", "retencion"))
                        row["code"] = f"{code} | {suffix}"
                        row["c_tasaocuota"] = code
                        code = str(row["code"])
                    if code not in seen:
                        seen.add(code)
                        continue
                    raise SatXlsSourceError(f"Clave duplicada en Excel para {catalog_code}: {code}")
                seen.add(code)
            sheets.append(descriptor)
            all_rows.extend(rows)
        return all_rows, sheets
    finally:
        _close_workbook(kind, workbook)


def inspect_source(path: str | Path) -> dict[str, object]:
    source = _ensure_source(path)
    kind, workbook = _open_workbook(source)
    try:
        sheet_names = _sheet_names(kind, workbook)
    finally:
        _close_workbook(kind, workbook)
    catalogs: dict[str, object] = {}
    for catalog_code in CATALOG_SHEETS:
        try:
            rows, sheets = extract_catalog_rows(source, catalog_code)
            catalogs[catalog_code] = {"sheets": sheets, "record_count": len(rows), "status": "ready"}
        except SatXlsSourceError as error:
            catalogs[catalog_code] = {"status": "error", "error": str(error)}
    return {"filename": source.name, "file_format": source.suffix.lower(), "sheets": sheet_names, "catalogs": catalogs}


def file_checksum(path: str | Path) -> str:
    source = _ensure_source(path)
    digest = sha256()
    with source.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def catalog_checksum(path: str | Path, catalog_code: str) -> str:
    digest = sha256(file_checksum(path).encode("ascii"))
    digest.update(catalog_code.encode("utf-8"))
    return digest.hexdigest()


def detect_version(path: str | Path) -> str | None:
    match = re.search(r"(?<!\d)(20\d{6})(?!\d)", Path(path).name)
    return match.group(1) if match else None


def source_metadata(path: str | Path, *, version: str, publication_date: date | None) -> dict[str, object]:
    source = _ensure_source(path)
    return {
        "source_type": "sat_official_xls",
        "origin": "SAT",
        "filename": source.name,
        "file_format": source.suffix.lower(),
        "size_bytes": source.stat().st_size,
        "sha256": file_checksum(source),
        "version": version,
        "data_date": publication_date.isoformat() if publication_date else None,
    }
